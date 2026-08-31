"""Tests for the backup tool: sanitize, interface parsing, workbook, tool run."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path

import pytest
from openpyxl import load_workbook

from bunnyauto.backup.collect import HostBackup
from bunnyauto.backup.interfaces import build_interface_rows
from bunnyauto.backup.sanitize import SanitizationError, sanitize_running_config
from bunnyauto.backup.workbook import create_interface_workbook
from bunnyauto.context import Settings
from bunnyauto.reporting import Reporter
from bunnyauto.result import Status
from bunnyauto.tools import backup
from bunnyauto.tools.backup import TOOL

# ---------------------------------------------------------------------------
# sanitize
# ---------------------------------------------------------------------------

_DIRTY = """\
hostname edge01
enable secret 9 $9$abcdef0123456789
username admin privilege 15 secret 5 $1$xxxx$yyyy
snmp-server community S3cr3tRO RO
snmp-server community S3cr3tRW RW
tacacs-server key 7 070C285F4D06
crypto isakmp key MyPreSharedKey address 10.0.0.1
ntp authentication-key 1 md5 070C285F
line vty 0 4
 password 7 06120A2D48
 transport input ssh
!
"""


def test_sanitize_redacts_known_secrets():
    out = sanitize_running_config(_DIRTY)
    assert out.startswith("! SANITIZED BACKUP")
    for secret in (
        "$9$abcdef0123456789",
        "$1$xxxx$yyyy",
        "S3cr3tRO",
        "S3cr3tRW",
        "070C285F4D06",
        "MyPreSharedKey",
        "06120A2D48",
    ):
        assert secret not in out, secret
    # structure is preserved
    assert "hostname edge01" in out
    assert "transport input ssh" in out
    assert "<removed enable secret>" in out
    assert "<removed SNMP community>" in out


def test_sanitize_is_idempotent():
    once = sanitize_running_config(_DIRTY)
    twice = sanitize_running_config(once)
    assert once == twice


def test_sanitize_redacts_url_credentials():
    out = sanitize_running_config("ntp server https://bob:hunter2@ntp.example.com\n")
    assert "hunter2" not in out
    assert "bob" not in out


def test_sanitize_fails_closed_on_unterminated_private_key():
    text = "crypto pki certificate chain X\n-----BEGIN RSA PRIVATE KEY-----\nAAAA\n"
    with pytest.raises(SanitizationError):
        sanitize_running_config(text)


def test_sanitize_drops_private_key_block():
    text = (
        "banner motd ^C hi ^C\n"
        "-----BEGIN OPENSSH PRIVATE KEY-----\n"
        "b3BlbnNzaC1rZXk=\n"
        "-----END OPENSSH PRIVATE KEY-----\n"
        "end\n"
    )
    out = sanitize_running_config(text)
    assert "b3BlbnNzaC1rZXk=" not in out
    assert "<removed private key block>" in out
    assert "end" in out


def test_sanitize_rejects_empty():
    with pytest.raises(SanitizationError):
        sanitize_running_config("   \n")


# ---------------------------------------------------------------------------
# interface parsing
# ---------------------------------------------------------------------------


def test_build_interface_rows_joins_by_interface():
    outputs = {
        "ip_brief": (
            "Interface              IP-Address      OK? Method Status                Protocol\n"
            "GigabitEthernet1/0/1   10.1.1.1        YES NVRAM  up                    up\n"
            "GigabitEthernet1/0/2   unassigned      YES unset  administratively down down\n"
        ),
        "interfaces": (
            "GigabitEthernet1/0/1 is up, line protocol is up\n"
            "  Description: uplink-core\n"
            "  Full-duplex, 1000Mb/s, media type is RJ45\n"
            "     1000 packets input, 640000 bytes\n"
            "     900 packets output, 570000 bytes\n"
        ),
    }
    rows = build_interface_rows(outputs)
    by_name = {row["Interface"]: row for row in rows}
    assert set(by_name) == {"Gi1/0/1", "Gi1/0/2"}
    assert by_name["Gi1/0/1"]["Description"] == "uplink-core"
    assert by_name["Gi1/0/1"]["Duplex"] == "Full-duplex"
    assert by_name["Gi1/0/1"]["RX Packets"] == 1000
    assert by_name["Gi1/0/2"]["Protocol State"] == "down"


# ---------------------------------------------------------------------------
# workbook
# ---------------------------------------------------------------------------


def test_create_interface_workbook(tmp_path: Path):
    rows = build_interface_rows(
        {
            "ip_brief": (
                "GigabitEthernet1/0/1   10.1.1.1   YES NVRAM  up   up\n"
                "GigabitEthernet1/0/2   unassigned YES unset  down down\n"
            )
        }
    )
    path = tmp_path / "sw1_interfaces.xlsx"
    create_interface_workbook("sw1", rows, {"show interfaces trunk": "no output"}, path)

    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Summary", "Interfaces"]
    sheet = workbook["Interfaces"]
    assert sheet.cell(4, 1).value == "Interface"
    assert sheet.cell(5, 1).value == "Gi1/0/1"
    assert sheet.cell(6, 1).value == "Gi1/0/2"


def test_create_interface_workbook_no_rows(tmp_path: Path):
    path = tmp_path / "empty.xlsx"
    create_interface_workbook("sw1", [], {}, path)
    assert path.is_file()


# ---------------------------------------------------------------------------
# the tool
# ---------------------------------------------------------------------------


class _Item:
    def __init__(self, result=None, exception=None):
        self.result = result
        self.exception = exception


class _Multi(list):
    def __init__(self, items, *, failed=False):
        super().__init__(items)
        self.failed = failed


class _Targets:
    def __init__(self, hosts, run_result):
        self.inventory = argparse.Namespace(hosts=hosts)
        self._run_result = run_result
        self.run_kwargs = None

    def run(self, **kwargs):
        self.run_kwargs = kwargs
        return self._run_result


@dataclass
class _Ctx:
    settings: Settings
    reporter: Reporter
    _nr: object = None

    def nornir(self):
        return self._nr


def _ctx() -> _Ctx:
    return _Ctx(
        settings=Settings(
            environment="test",
            nb_url="https://nb",
            config_file="config.yaml",
            target_tag="nornirtest",
            read_timeout=180.0,
        ),
        reporter=Reporter(json_mode=True),
    )


def _args(*, raw: bool = False, output_dir: str = "./bk") -> argparse.Namespace:
    return argparse.Namespace(raw=raw, output_dir=output_dir)


def _patch(monkeypatch, hosts, run_result):
    targets = _Targets(hosts, run_result)
    monkeypatch.setattr(backup, "filter_by_tag", lambda nr, tag: targets)
    return targets


def test_backup_no_devices(monkeypatch):
    _patch(monkeypatch, {}, {})
    result = TOOL.run(_ctx(), _args())
    assert result.status is Status.OK
    assert "nornirtest" in result.summary


def test_backup_all_succeed(monkeypatch, tmp_path):
    rec1 = HostBackup(host="sw1", directory=tmp_path / "sw1", config_path=tmp_path / "sw1/sw1.cfg")
    rec2 = HostBackup(host="sw2", directory=tmp_path / "sw2", config_path=tmp_path / "sw2/sw2.cfg")
    run_result = {"sw1": _Multi([_Item(result=rec1)]), "sw2": _Multi([_Item(result=rec2)])}
    targets = _patch(monkeypatch, {"sw1": 1, "sw2": 1}, run_result)

    result = TOOL.run(_ctx(), _args())

    assert result.status is Status.OK
    assert result.data["sanitized"] is True
    assert targets.run_kwargs["sanitize"] is True
    assert len(result.artifacts) == 2


def test_backup_raw_flag_disables_sanitize(monkeypatch, tmp_path):
    rec = HostBackup(host="sw1", sanitized=False, directory=tmp_path / "sw1")
    targets = _patch(monkeypatch, {"sw1": 1}, {"sw1": _Multi([_Item(result=rec)])})
    result = TOOL.run(_ctx(), _args(raw=True))
    assert targets.run_kwargs["sanitize"] is False
    assert result.data["sanitized"] is False
    assert "(raw)" in result.summary


def test_backup_partial_and_error(monkeypatch, tmp_path):
    good = HostBackup(host="sw1", directory=tmp_path / "sw1")
    bad = HostBackup(host="sw2", error="sanitization failed — config not written: leak at line 4")
    run_result = {
        "sw1": _Multi([_Item(result=good)]),
        "sw2": _Multi([_Item(result=bad)], failed=True),
    }
    _patch(monkeypatch, {"sw1": 1, "sw2": 1}, run_result)
    result = TOOL.run(_ctx(), _args())

    assert result.status is Status.PARTIAL
    assert result.data["devices"]["sw2"]["ok"] is False
    assert "sanitization failed" in result.data["devices"]["sw2"]["error"]
