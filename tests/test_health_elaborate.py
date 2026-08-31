"""Tests for health-elaborate: EtherChannel/quality/security parsers, workbook, tool."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path

import pytest
from openpyxl import load_workbook

from bunnyauto.context import Settings
from bunnyauto.errors import ToolError
from bunnyauto.health.elaborate_collect import (
    _duration_seconds,
    _interface_key,
    enrich_etherchannels,
    parse_err_disabled_interfaces,
    parse_etherchannels,
    parse_interface_quality,
    parse_port_security_issues,
    parse_security_control_exceptions,
)
from bunnyauto.health.elaborate_workbook import (
    build_collection_notes,
    build_issue_rows,
    create_elaborate_health_workbook,
)
from bunnyauto.reporting import Reporter
from bunnyauto.result import Status
from bunnyauto.tools import health_elaborate
from bunnyauto.tools.health_elaborate import TOOL

# ---------------------------------------------------------------------------
# small helpers
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("text", "expected"),
    [
        ("00:12:34", 754),
        ("1:00:00", 3600),
        ("2w3d", 2 * 604_800 + 3 * 86_400),
        ("never", None),
    ],
)
def test_duration_seconds(text, expected):
    assert _duration_seconds(text) == expected


def test_interface_key_normalizes():
    assert _interface_key("GigabitEthernet1/0/1") == _interface_key("Gi1/0/1")


# ---------------------------------------------------------------------------
# etherchannels
# ---------------------------------------------------------------------------

_ETHERCHANNEL_SUMMARY = """\
Flags:  D - down        P - bundled in port-channel
        s - suspended   H - Hot-standby (LACP only)
Number of channel-groups in use: 2
Group  Port-channel  Protocol    Ports
------+-------------+-----------+-----------------------------------------------
1      Po1(SU)         LACP        Gi1/0/1(P)   Gi1/0/2(P)
2      Po2(SD)         LACP        Gi1/0/3(s)   Gi1/0/4(s)
"""


def test_parse_etherchannels_states():
    channels = parse_etherchannels(_ETHERCHANNEL_SUMMARY)
    by_group = {c.group: c for c in channels}
    assert by_group[1].state == "Up"
    assert by_group[1].bundled_members == ["Gi1/0/1", "Gi1/0/2"]
    assert by_group[2].state == "Down"  # SD flag => not up
    assert by_group[2].protocol.upper() == "LACP"


def test_enrich_etherchannels_min_links():
    channels = parse_etherchannels(_ETHERCHANNEL_SUMMARY)
    detail = "Group: 1\n  Minimum Links: 3\n"
    enrich_etherchannels(channels, detail)
    po1 = next(c for c in channels if c.group == 1)
    assert po1.min_links == 3
    assert po1.state == "Down"  # only 2 bundled, needs 3
    assert any("minimum links" in r for r in po1.risk_reasons)


# ---------------------------------------------------------------------------
# other parsers
# ---------------------------------------------------------------------------


def test_parse_port_security_issues():
    text = (
        "Secure Port  MaxSecureAddr  CurrentAddr  SecurityViolation  Security Action\n"
        "Gi1/0/5                2              1                  4          Shutdown\n"
        "Gi1/0/6                2              1                  0          Restrict\n"
    )
    issues = parse_port_security_issues(text)
    assert len(issues) == 1
    assert issues[0].interface == "Gi1/0/5"
    assert issues[0].violation_count == 4


def test_parse_err_disabled_interfaces():
    text = "Gi1/0/7   err-disabled psecure-violation\n"
    issues = parse_err_disabled_interfaces(text)
    assert issues == [{"interface": "Gi1/0/7", "reason": "psecure-violation"}]


def test_parse_interface_quality_counters():
    text = (
        "GigabitEthernet1/0/1 is up, line protocol is up\n"
        "  1234 input errors, 12 CRC, 0 frame, 0 overrun, 0 ignored\n"
        "  5 interface resets\n"
    )
    issues = parse_interface_quality(text)
    assert len(issues) == 1
    issue = issues[0]
    assert issue.counters["input_errors"] == 1234
    assert issue.counters["crc_errors"] == 12
    assert issue.interface_resets == 5
    assert issue.unstable is True


def test_parse_security_control_exceptions_copp():
    outputs = {
        "control_plane": (
            "Class-map: CoPP-CRITICAL (match-any)\n  1000 packets, 64000 bytes\n  drop packets 42\n"
        )
    }
    exceptions = parse_security_control_exceptions(outputs)
    assert len(exceptions) == 1
    assert exceptions[0].count == 42
    assert exceptions[0].category == "Control Plane"


# ---------------------------------------------------------------------------
# notes / issue rows
# ---------------------------------------------------------------------------


def test_build_collection_notes_etherchannel_down():
    record = {
        "reachable": True,
        "notes": [],
        "etherchannels_total": 2,
        "etherchannels_down": 2,
        "etherchannels": [
            {
                "name": "Po1",
                "state": "Down",
                "members": ["Gi1"],
                "bundled_members": [],
                "risk_reasons": [],
            },
            {
                "name": "Po2",
                "state": "Down",
                "members": ["Gi2"],
                "bundled_members": [],
                "risk_reasons": [],
            },
        ],
    }
    notes = build_collection_notes(record)
    assert any("critical outage threshold" in n for n in notes)


def test_build_issue_rows_unreachable_only():
    rows = build_issue_rows([{"hostname": "sw1", "reachable": False, "notes": []}])
    assert len(rows) == 1
    assert rows[0][1] == "Critical"
    assert rows[0][2] == "Reachability"


def test_build_issue_rows_port_security_is_info():
    record = {
        "hostname": "sw1",
        "reachable": True,
        "notes": [],
        "port_security_issues": [
            {"interface": "Gi1/0/5", "violation_count": 4, "action": "Shutdown"}
        ],
    }
    rows = build_issue_rows([record])
    ps_rows = [r for r in rows if r[2] == "Port Security"]
    assert ps_rows and ps_rows[0][1] == "Info"  # informational, never a penalty


# ---------------------------------------------------------------------------
# workbook
# ---------------------------------------------------------------------------


def test_create_elaborate_workbook(tmp_path: Path):
    records = [
        {"hostname": "sw1", "reachable": True, "firmware": "17.9", "etherchannels_total": 1},
        {"hostname": "sw2", "reachable": False, "notes": ["unreachable"]},
    ]
    path = tmp_path / "elab.xlsx"
    create_elaborate_health_workbook(records, "nornirtest", path)

    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Engineer Health", "Issue Details", "Scoring"]
    overview = workbook["Engineer Health"]
    assert overview.cell(4, 2).value == "sw1"
    assert overview.cell(16, 2).value == "17.9"  # Current Firmware row
    assert workbook["Scoring"].sheet_state == "veryHidden"


# ---------------------------------------------------------------------------
# the tool
# ---------------------------------------------------------------------------


class _Targets:
    def __init__(self, hosts, run_result):
        self.inventory = argparse.Namespace(hosts=hosts)
        self._run_result = run_result

    def run(self, **kwargs):
        return self._run_result


@dataclass
class _Ctx:
    settings: Settings
    reporter: Reporter
    _nr: object = None

    def nornir(self):
        return self._nr


def _ctx() -> _Ctx:
    return _Ctx(
        settings=Settings(
            environment="test",
            nb_url="https://nb",
            config_file="config.yaml",
            target_tag="nornirtest",
            read_timeout=180.0,
        ),
        reporter=Reporter(json_mode=True),
    )


def test_output_must_be_xlsx():
    with pytest.raises(ToolError):
        TOOL.run(_ctx(), argparse.Namespace(output="report.txt"))


def test_tool_writes_workbook(monkeypatch, tmp_path):
    hosts = {"sw1": object(), "sw2": object()}
    monkeypatch.setattr(health_elaborate, "filter_by_tag", lambda nr, tag: _Targets(hosts, {}))
    monkeypatch.setattr(
        health_elaborate,
        "extract_records",
        lambda results, hosts_: [
            {"hostname": "sw1", "reachable": True},
            {"hostname": "sw2", "reachable": True},
        ],
    )
    out = tmp_path / "e.xlsx"
    result = TOOL.run(_ctx(), argparse.Namespace(output=str(out)))
    assert result.status is Status.OK
    assert out.is_file()
    assert result.artifacts == [out]


def test_tool_no_devices(monkeypatch):
    monkeypatch.setattr(health_elaborate, "filter_by_tag", lambda nr, tag: _Targets({}, {}))
    result = TOOL.run(_ctx(), argparse.Namespace(output=None))
    assert result.status is Status.OK
