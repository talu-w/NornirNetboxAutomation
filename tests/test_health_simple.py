"""Tests for health-simple: collection parsers, workbook, tool."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path

import pytest
from openpyxl import load_workbook

from bunnyauto.context import Settings
from bunnyauto.errors import ToolError
from bunnyauto.health import collect
from bunnyauto.health.collect import (
    command_profile,
    count_environment_alerts,
    extract_records,
    parse_cpu_pct,
    parse_firmware,
    parse_interface_summary,
    run_first_supported,
)
from bunnyauto.health.simple_workbook import build_collection_notes, create_health_workbook
from bunnyauto.reporting import Reporter
from bunnyauto.result import Status
from bunnyauto.tools import health_simple
from bunnyauto.tools.health_simple import TOOL

# ---------------------------------------------------------------------------
# parsers
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("text", "expected"),
    [
        ("Cisco IOS XE Software, Version 17.09.04a, RELEASE", "17.09.04a"),
        ("  NXOS: version 9.3(10)", "9.3(10)"),
        ("nothing useful here", "Unknown"),
    ],
)
def test_parse_firmware(text, expected):
    assert parse_firmware(text) == expected


@pytest.mark.parametrize(
    ("text", "expected"),
    [
        ("CPU utilization for five seconds: 23%/0%; one minute: 19%", 23.0),
        ("  91.5% idle", pytest.approx(8.5)),
        ("no cpu data", None),
    ],
)
def test_parse_cpu_pct(text, expected):
    assert parse_cpu_pct(text) == expected


def test_parse_interface_summary_switch_style():
    text = (
        "Port      Name    Status       Vlan       Duplex  Speed Type\n"
        "Gi1/0/1           connected    10         a-full  a-1000 10/100/1000BaseTX\n"
        "Gi1/0/2           notconnect   20         auto    auto   10/100/1000BaseTX\n"
        "Gi1/0/3           err-disabled 30         auto    auto   10/100/1000BaseTX\n"
    )
    summary = parse_interface_summary(text)
    assert summary.total == 3
    assert summary.connected == 1
    assert summary.err_disabled == 1
    assert summary.not_connected == 2


def test_count_environment_alerts():
    text = "PSU1 Status: OK\nPSU2 Status: FAILED\nFan tray 1: normal\nTemp sensor: CRITICAL\n"
    assert count_environment_alerts(text) == 2


def test_command_profile_nxos_overrides():
    profile = command_profile("cisco_nxos")
    assert profile["cpu"] == ["show system resources"]
    assert profile["interfaces"] == ["show interface status"]


# ---------------------------------------------------------------------------
# run_first_supported
# ---------------------------------------------------------------------------


class _R:
    def __init__(self, result="", *, failed=False, exception=None):
        self.result = result
        self.failed = failed
        self.exception = exception


class _Task:
    def __init__(self, responses):
        self.responses = responses

    def run(self, *, name, task, command_string, read_timeout):
        response = self.responses[command_string]
        if isinstance(response, Exception):
            raise response
        return [response]


def test_run_first_supported_falls_through():
    task = _Task(
        {
            "show interface status": _R("% Invalid input detected"),
            "show ip interface brief": _R("Gi0/1 up up"),
        }
    )
    output, command = run_first_supported(
        task, "Interfaces", ["show interface status", "show ip interface brief"]
    )
    assert command == "show ip interface brief"
    assert "Gi0/1" in output


def test_run_first_supported_all_fail():
    task = _Task({"show version": _R("", failed=True, exception=RuntimeError("timeout"))})
    with pytest.raises(RuntimeError):
        run_first_supported(task, "Firmware", ["show version"])


# ---------------------------------------------------------------------------
# metadata / records / notes
# ---------------------------------------------------------------------------


class _Host:
    def __init__(self, data, platform="ios", hostname="10.0.0.1"):
        self.data = data
        self.platform = platform
        self.hostname = hostname


def test_netbox_metadata():
    host = _Host(
        {
            "site": {"name": "HQ"},
            "role": {"name": "access"},
            "device_type": {"model": "C9300-48P"},
            "status": "active",
            "serial": "FCW123",
        }
    )
    meta = collect.netbox_metadata(host)
    assert meta["site"] == "HQ"
    assert meta["model"] == "C9300-48P"
    assert meta["serial_number"] == "FCW123"


def test_extract_records_placeholder_for_missing():
    class _Results:
        def get(self, name):
            return None

    hosts = {"sw1": _Host({})}
    records = extract_records(_Results(), hosts)
    assert records[0]["hostname"] == "sw1"
    assert "did not return a collection result" in records[0]["notes"][0]


def test_build_collection_notes_unreachable():
    notes = build_collection_notes({"reachable": False, "notes": []})
    assert any("unreachable" in n for n in notes)


def test_build_collection_notes_cpu_and_env():
    notes = build_collection_notes(
        {"reachable": True, "cpu_pct": 95.0, "environment_alerts": 2, "notes": []}
    )
    assert any("critical threshold" in n for n in notes)
    assert any("environmental alert" in n for n in notes)


# ---------------------------------------------------------------------------
# workbook
# ---------------------------------------------------------------------------


def test_create_health_workbook(tmp_path: Path):
    records = [
        {"hostname": "sw1", "reachable": True, "cpu_pct": 12.0, "firmware": "17.9"},
        {"hostname": "sw2", "reachable": False, "notes": ["unreachable"]},
    ]
    path = tmp_path / "report.xlsx"
    create_health_workbook(records, "nornirtest", path)

    workbook = load_workbook(path)
    sheet = workbook["Network Health"]
    assert sheet.cell(4, 2).value == "sw1"
    assert sheet.cell(4, 3).value == "sw2"
    assert sheet.cell(16, 2).value == "17.9"  # Current Firmware row


# ---------------------------------------------------------------------------
# the tool
# ---------------------------------------------------------------------------


class _Targets:
    def __init__(self, hosts, run_result):
        self.inventory = argparse.Namespace(hosts=hosts)
        self._run_result = run_result

    def run(self, **kwargs):
        return self._run_result


@dataclass
class _Ctx:
    settings: Settings
    reporter: Reporter
    _nr: object = None

    def nornir(self):
        return self._nr


def _ctx() -> _Ctx:
    return _Ctx(
        settings=Settings(
            environment="test",
            nb_url="https://nb",
            config_file="config.yaml",
            target_tag="nornirtest",
            read_timeout=180.0,
        ),
        reporter=Reporter(json_mode=True),
    )


def test_output_must_be_xlsx():
    with pytest.raises(ToolError):
        TOOL.run(_ctx(), argparse.Namespace(output="report.pdf"))


def test_tool_writes_workbook(monkeypatch, tmp_path):
    hosts = {"sw1": _Host({}), "sw2": _Host({})}

    class _Results(dict):
        pass

    run_result = _Results()
    monkeypatch.setattr(health_simple, "filter_by_tag", lambda nr, tag: _Targets(hosts, run_result))
    monkeypatch.setattr(
        health_simple,
        "extract_records",
        lambda results, hosts_: [
            {"hostname": "sw1", "reachable": True},
            {"hostname": "sw2", "reachable": False, "notes": []},
        ],
    )

    out = tmp_path / "r.xlsx"
    result = TOOL.run(_ctx(), argparse.Namespace(output=str(out)))

    assert result.status is Status.PARTIAL  # sw2 unreachable
    assert out.is_file()
    assert result.data["reachable"] == 1
    assert result.artifacts == [out]


def test_tool_no_devices(monkeypatch):
    monkeypatch.setattr(health_simple, "filter_by_tag", lambda nr, tag: _Targets({}, {}))
    result = TOOL.run(_ctx(), argparse.Namespace(output=None))
    assert result.status is Status.OK
    assert "nornirtest" in result.summary
