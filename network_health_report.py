"""Create a management network-health workbook from Nornir and NetBox inventory.

The script uses the repository's existing Nornir ``config.yaml``. NetBox devices
are loaded by that configuration, normalized, and filtered with Nornir's ``F``
filter. No device configuration is changed.
"""

from __future__ import annotations

import argparse
import os
import re
from collections.abc import Mapping, Sequence
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from nornir import InitNornir
from nornir.core.filter import F
from nornir.core.inventory import ConnectionOptions, Host
from nornir.core.task import Result, Task
from nornir_netmiko.tasks import netmiko_send_command

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "ERROR: This script requires openpyxl. Install it with: pip install openpyxl"
    ) from exc

try:
    from rich.console import Console
except ImportError as exc:
    raise SystemExit(
        "ERROR: This script requires Rich. Install it with: pip install rich"
    ) from exc


DEFAULT_CONFIG_FILE = "config.yaml"
DEFAULT_TARGET_TAG = "nornirtest"
DEFAULT_OUTPUT_FILE = "network_health_report.xlsx"

# Internal scoring values. They are stored in hidden worksheet rows so the
# front-facing report stays clean while its formulas remain consistent.
CPU_WARNING_PCT = 75.0
CPU_CRITICAL_PCT = 90.0
MIN_DATA_COVERAGE_PCT = 60.0
WARNING_PENALTY = 10
CRITICAL_PENALTY = 25
ENVIRONMENT_ALERT_PENALTY = 10
ERR_DISABLED_PENALTY = 10
MAX_COUNT_PENALTY = 30
HEALTHY_SCORE = 85
WATCH_SCORE = 70
HEALTH_COMPONENT_COUNT = 4

NETMIKO_EXTRAS = {
    "conn_timeout": 30,
    "banner_timeout": 60,
    "auth_timeout": 60,
    "fast_cli": False,
}

PHYSICAL_INTERFACE_RE = re.compile(
    r"^(?:Fa|FastEthernet|Gi|GigabitEthernet|Te|TenGigabitEthernet|"
    r"Tw|TwoGigabitEthernet|Twe|TwentyFiveGigE|Fo|FortyGigabitEthernet|"
    r"Hu|HundredGig(?:E|abitEthernet)|Eth|Ethernet|Et)\d",
    re.IGNORECASE,
)

INTERFACE_STATUS_RE = re.compile(
    r"\b(err-disabled|notconnect(?:ed)?|notconnec|connected|disabled|inactive|"
    r"monitoring|sfpAbsent|xcvrAbsent)\b",
    re.IGNORECASE,
)

INVALID_COMMAND_MARKERS = (
    "% invalid input",
    "% incomplete command",
    "% ambiguous command",
    "invalid command",
    "unknown command",
)

console = Console()


@dataclass(slots=True)
class InterfaceSummary:
    connected: int | None = None
    not_connected: int | None = None
    total: int | None = None
    err_disabled: int | None = None


@dataclass(slots=True)
class DeviceHealth:
    hostname: str
    location: str = ""
    site: str = ""
    role: str = ""
    model: str = ""
    platform: str = ""
    netbox_status: str = ""
    primary_ip: str = ""
    serial_number: str = ""
    firmware: str = ""
    connected_interfaces: int | None = None
    not_connected_interfaces: int | None = None
    total_interfaces: int | None = None
    cpu_pct: float | None = None
    environment_alerts: int | None = None
    err_disabled_interfaces: int | None = None
    reachable: bool = False
    collected_at_utc: datetime | None = None
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Collect management health data from NetBox-tagged devices with Nornir "
            "and create one Excel scorecard."
        )
    )
    parser.add_argument(
        "--config",
        default=os.getenv("NORNIR_CONFIG_FILE", DEFAULT_CONFIG_FILE),
        help="Existing Nornir config file (default: config.yaml)",
    )
    parser.add_argument(
        "--tag",
        default=os.getenv("NORNIR_TARGET_TAG", DEFAULT_TARGET_TAG),
        help="NetBox tag to select (default: nornirtest)",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT_FILE,
        help="Excel output path (default: network_health_report.xlsx)",
    )
    return parser.parse_args(argv)


def normalize_tags(tags: Sequence[Any] | None) -> list[str]:
    """Convert NetBox tag strings, dictionaries, or objects to lowercase values."""

    normalized: set[str] = set()
    for tag in tags or []:
        if isinstance(tag, str):
            values = [tag]
        elif isinstance(tag, Mapping):
            values = [tag.get("slug"), tag.get("name")]
        else:
            values = [getattr(tag, "slug", None), getattr(tag, "name", None)]

        for value in values:
            if value:
                normalized.add(str(value).casefold())
    return sorted(normalized)


def _nested_value(data: Mapping[str, Any], *paths: str) -> Any:
    for path in paths:
        value: Any = data
        for part in path.split("."):
            if isinstance(value, Mapping):
                value = value.get(part)
            else:
                value = getattr(value, part, None)
            if value is None:
                break
        if value not in (None, ""):
            return value
    return None


def _display_value(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, Mapping):
        for key in ("display", "label", "name", "model", "address", "value", "slug"):
            if value.get(key) not in (None, ""):
                return str(value[key])
        return ""
    for attribute in ("display", "label", "name", "model", "address", "value", "slug"):
        candidate = getattr(value, attribute, None)
        if candidate not in (None, ""):
            return str(candidate)
    return str(value)


def netbox_metadata(host: Host) -> dict[str, str]:
    """Read useful management fields already supplied by the NetBox inventory."""

    data = host.data
    site = _display_value(_nested_value(data, "site", "site_name"))
    location = _display_value(_nested_value(data, "location", "location_name")) or site
    return {
        "location": location,
        "site": site,
        "role": _display_value(_nested_value(data, "role", "device_role")),
        "model": _display_value(_nested_value(data, "device_type.model", "device_type", "model")),
        "platform": _display_value(host.platform or _nested_value(data, "platform")),
        "netbox_status": _display_value(_nested_value(data, "status")),
        "primary_ip": _display_value(
            _nested_value(data, "primary_ip4.address", "primary_ip4", "primary_ip.address", "primary_ip")
        )
        or str(host.hostname or ""),
        "serial_number": _display_value(_nested_value(data, "serial", "serial_number")),
    }


def _is_invalid_command(output: str) -> bool:
    lowered = output.casefold()
    return any(marker in lowered for marker in INVALID_COMMAND_MARKERS)


def run_first_supported(
    task: Task,
    label: str,
    commands: Sequence[str],
    read_timeout: int = 120,
) -> tuple[str, str]:
    """Run commands in order and return the first usable output and command."""

    errors: list[str] = []
    for command in commands:
        try:
            results = task.run(
                name=f"{label}: {command}",
                task=netmiko_send_command,
                command_string=command,
                read_timeout=read_timeout,
            )
            result = results[-1]
        except Exception as exc:  # noqa: BLE001 - plugins raise platform-specific errors.
            errors.append(f"{command}: {exc}")
            continue

        if result.failed:
            errors.append(f"{command}: {result.exception or result.result}")
            continue

        output = str(result.result or "")
        if output.strip() and not _is_invalid_command(output):
            return output, command
        errors.append(f"{command}: unsupported or empty output")

    raise RuntimeError("; ".join(errors) or f"No {label} command succeeded")


def command_profile(platform: str) -> dict[str, list[str]]:
    platform_name = platform.casefold()
    profile = {
        "version": ["show version"],
        "interfaces": ["show interfaces status", "show ip interface brief"],
        "cpu": ["show processes cpu | include CPU utilization", "show processes cpu"],
        "environment": ["show environment all", "show environment"],
    }

    if "nxos" in platform_name or "nx-os" in platform_name:
        profile.update(
            {
                "interfaces": ["show interface status"],
                "cpu": ["show system resources"],
                "environment": ["show environment"],
            }
        )
    elif "eos" in platform_name or "arista" in platform_name:
        profile.update(
            {
                "interfaces": ["show interfaces status"],
                "cpu": ["show processes top once"],
                "environment": ["show system environment all"],
            }
        )
    return profile


def parse_firmware(output: str) -> str:
    patterns = (
        r"Cisco IOS XE Software, Version\s+([^,\s]+)",
        r"Cisco IOS Software.*?Version\s+([^,\s]+)",
        r"NXOS:\s+version\s+([^\s]+)",
        r"system:\s+version\s+([^\s]+)",
        r"Software image version:\s*([^\s]+)",
        r"Arista .*? version\s+([^\s]+)",
    )
    for pattern in patterns:
        match = re.search(pattern, output, re.IGNORECASE | re.DOTALL)
        if match:
            return match.group(1).strip()
    return "Unknown"


def parse_cpu_pct(output: str) -> float | None:
    match = re.search(
        r"CPU utilization for five seconds:\s*(\d+(?:\.\d+)?)%",
        output,
        re.IGNORECASE,
    )
    if match:
        return float(match.group(1))

    idle_match = re.search(
        r"(\d+(?:\.\d+)?)%\s*(?:idle|id)\b", output, re.IGNORECASE
    )
    if idle_match:
        return max(0.0, min(100.0, 100.0 - float(idle_match.group(1))))

    user_match = re.search(r"(\d+(?:\.\d+)?)%\s*user", output, re.IGNORECASE)
    kernel_match = re.search(r"(\d+(?:\.\d+)?)%\s*kernel", output, re.IGNORECASE)
    if user_match and kernel_match:
        return min(100.0, float(user_match.group(1)) + float(kernel_match.group(1)))
    return None


def parse_interface_summary(output: str) -> InterfaceSummary:
    connected = 0
    total = 0
    err_disabled = 0

    for line in output.splitlines():
        tokens = line.split()
        if not tokens or not PHYSICAL_INTERFACE_RE.match(tokens[0]):
            continue
        status_match = INTERFACE_STATUS_RE.search(line)
        if not status_match:
            continue
        status = status_match.group(1).casefold()
        total += 1
        connected += int(status == "connected")
        err_disabled += int(status == "err-disabled")

    if total:
        return InterfaceSummary(
            connected=connected,
            not_connected=total - connected,
            total=total,
            err_disabled=err_disabled,
        )

    # Router-style fallback for "show ip interface brief".
    brief_pattern = re.compile(
        r"^\s*(?P<interface>\S+)\s+\S+\s+\S+\s+\S+\s+"
        r"(?P<status>administratively down|up|down)\s+(?P<protocol>up|down)\s*$",
        re.IGNORECASE,
    )
    for line in output.splitlines():
        match = brief_pattern.match(line)
        if not match or not PHYSICAL_INTERFACE_RE.match(match.group("interface")):
            continue
        total += 1
        connected += int(
            match.group("status").casefold() == "up"
            and match.group("protocol").casefold() == "up"
        )

    return InterfaceSummary(
        connected=connected if total else None,
        not_connected=(total - connected) if total else None,
        total=total if total else None,
        err_disabled=0 if total else None,
    )


def count_environment_alerts(output: str) -> int:
    alert_terms = re.compile(
        r"\b(fail(?:ed|ure)?|critical|alarm|shutdown|overtemp)\b", re.IGNORECASE
    )
    healthy_terms = re.compile(
        r"\b(no alarms?|normal|ok|good|passed|not present)\b", re.IGNORECASE
    )
    return sum(
        1
        for line in output.splitlines()
        if alert_terms.search(line) and not healthy_terms.search(line)
    )


def collect_device_health(task: Task) -> Result:
    metadata = netbox_metadata(task.host)
    record = DeviceHealth(hostname=task.host.name, **metadata)
    profile = command_profile(record.platform)

    try:
        version_output, _ = run_first_supported(task, "Firmware", profile["version"])
    except Exception as exc:  # noqa: BLE001 - collection errors must remain in the report.
        record.notes.append(f"Connection or show version failed: {exc}")
        return Result(host=task.host, changed=False, result=record.to_dict())

    record.reachable = True
    record.firmware = parse_firmware(version_output)
    record.collected_at_utc = datetime.now(UTC).replace(tzinfo=None)

    try:
        interface_output, _ = run_first_supported(task, "Interfaces", profile["interfaces"])
        summary = parse_interface_summary(interface_output)
        record.connected_interfaces = summary.connected
        record.not_connected_interfaces = summary.not_connected
        record.total_interfaces = summary.total
        record.err_disabled_interfaces = summary.err_disabled
        if summary.total is None:
            record.notes.append("Interface status output was returned but no physical ports were parsed.")
    except Exception as exc:  # noqa: BLE001 - collection errors must remain in the report.
        record.notes.append(f"Interface statistics unavailable: {exc}")

    try:
        cpu_output, _ = run_first_supported(task, "CPU", profile["cpu"])
        record.cpu_pct = parse_cpu_pct(cpu_output)
        if record.cpu_pct is None:
            record.notes.append("CPU output was returned but utilization could not be parsed.")
    except Exception as exc:  # noqa: BLE001 - collection errors must remain in the report.
        record.notes.append(f"CPU utilization unavailable: {exc}")

    try:
        environment_output, _ = run_first_supported(
            task, "Environment", profile["environment"]
        )
        record.environment_alerts = count_environment_alerts(environment_output)
    except Exception as exc:  # noqa: BLE001 - collection errors must remain in the report.
        record.notes.append(f"Environment status unavailable: {exc}")

    return Result(host=task.host, changed=False, result=record.to_dict())


def _extract_records(results: Any, hosts: Mapping[str, Host]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for hostname, host in hosts.items():
        record: dict[str, Any] | None = None
        multi_result = results.get(hostname)
        if multi_result is not None:
            for item in multi_result:
                if isinstance(item.result, dict) and item.result.get("hostname") == hostname:
                    record = item.result
                    break
        if record is None:
            record = DeviceHealth(
                hostname=hostname,
                **netbox_metadata(host),
                notes=["Nornir did not return a collection result for this device."],
            ).to_dict()
        records.append(record)
    return sorted(records, key=lambda item: item["hostname"].casefold())


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def build_collection_notes(record: Mapping[str, Any]) -> list[str]:
    """Combine collection warnings with the conditions that reduced health."""

    notes = [str(note) for note in record.get("notes", []) if str(note).strip()]

    def add(note: str) -> None:
        if note not in notes:
            notes.append(note)

    if not record.get("reachable"):
        add("Health impact: device was unreachable; health score is forced to 0.")
        return notes

    cpu_pct = record.get("cpu_pct")
    if isinstance(cpu_pct, (int, float)):
        if cpu_pct >= CPU_CRITICAL_PCT:
            add(
                f"Health impact: CPU utilization {cpu_pct:.1f}% met or exceeded the "
                f"critical threshold of {CPU_CRITICAL_PCT:.0f}%; score reduced by "
                f"{CRITICAL_PENALTY} points."
            )
        elif cpu_pct >= CPU_WARNING_PCT:
            add(
                f"Health impact: CPU utilization {cpu_pct:.1f}% met or exceeded the "
                f"warning threshold of {CPU_WARNING_PCT:.0f}%; score reduced by "
                f"{WARNING_PENALTY} points."
            )

    environment_alerts = record.get("environment_alerts")
    if isinstance(environment_alerts, (int, float)) and environment_alerts > 0:
        penalty = min(
            int(environment_alerts) * ENVIRONMENT_ALERT_PENALTY, MAX_COUNT_PENALTY
        )
        add(
            f"Health impact: {int(environment_alerts)} environmental alert"
            f"{'s were' if environment_alerts != 1 else ' was'} detected; score reduced "
            f"by {penalty} points."
        )

    err_disabled = record.get("err_disabled_interfaces")
    if isinstance(err_disabled, (int, float)) and err_disabled > 0:
        penalty = min(int(err_disabled) * ERR_DISABLED_PENALTY, MAX_COUNT_PENALTY)
        add(
            f"Health impact: {int(err_disabled)} physical interface"
            f"{'s are' if err_disabled != 1 else ' is'} err-disabled; score reduced by "
            f"{penalty} points."
        )

    available_components = 1 + sum(
        isinstance(record.get(field), (int, float))
        for field in ("cpu_pct", "environment_alerts", "err_disabled_interfaces")
    )
    coverage_pct = available_components / HEALTH_COMPONENT_COUNT * 100
    if coverage_pct < MIN_DATA_COVERAGE_PCT:
        add(
            f"Health impact: only {coverage_pct:.0f}% of health inputs were available; "
            "status is Insufficient Data."
        )

    return notes


def create_health_workbook(
    records: Sequence[Mapping[str, Any]],
    target_tag: str,
    output_path: Path,
) -> None:
    """Create one leadership scorecard with a device in each horizontal column."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Network Health"
    sheet.sheet_view.showGridLines = False

    navy = "17365D"
    teal = "147D74"
    green = "C6EFCE"
    amber = "FFEB9C"
    red = "FFC7CE"
    pale_blue = "D9EAF7"
    white = "FFFFFF"
    thin_gray = Side(style="thin", color="D9E1F2")

    last_column = max(2, len(records) + 1)
    last_column_letter = get_column_letter(last_column)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet["A1"] = "Networking Infrastructure Health — Leadership Overview"
    sheet["A1"].font = Font(size=18, bold=True, color=white)
    sheet["A1"].fill = _fill(navy)
    sheet["A1"].alignment = Alignment(vertical="center", wrap_text=True, shrink_to_fit=True)
    sheet.row_dimensions[1].height = 42

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    generated = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    sheet["A2"] = (
        f"Scope: devices from the existing Nornir/NetBox inventory tagged "
        f"'{target_tag}' | Generated {generated}"
    )
    sheet["A2"].font = Font(italic=True, color="44546A")
    sheet["A2"].fill = _fill(pale_blue)
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True, shrink_to_fit=True)
    sheet.row_dimensions[2].height = 32

    metric_rows = {
        "Health Status": 5,
        "Health Score": 6,
        "Data Coverage": 7,
        "Location": 8,
        "Site": 9,
        "Role": 10,
        "Model": 11,
        "Platform": 12,
        "NetBox Status": 13,
        "Primary IP": 14,
        "Serial Number": 15,
        "Current Firmware": 16,
        "Interface Usage": 17,
        "Connected Interfaces": 18,
        "Not Connected Interfaces": 19,
        "Total Interfaces": 20,
        "CPU Utilization": 21,
        "Environment Alerts": 22,
        "Err-disabled Interfaces": 23,
        "Reachable": 24,
        "Collected UTC": 25,
        "Collection Notes": 26,
    }

    sheet["A4"] = "Metric"
    for label, row in metric_rows.items():
        sheet.cell(row, 1, label)

    for row in range(4, 27):
        cell = sheet.cell(row, 1)
        cell.fill = _fill(navy)
        cell.font = Font(bold=True, color=white)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    field_map = {
        "Location": "location",
        "Site": "site",
        "Role": "role",
        "Model": "model",
        "Platform": "platform",
        "NetBox Status": "netbox_status",
        "Primary IP": "primary_ip",
        "Serial Number": "serial_number",
        "Current Firmware": "firmware",
        "Connected Interfaces": "connected_interfaces",
        "Not Connected Interfaces": "not_connected_interfaces",
        "Total Interfaces": "total_interfaces",
        "Environment Alerts": "environment_alerts",
        "Err-disabled Interfaces": "err_disabled_interfaces",
        "Collected UTC": "collected_at_utc",
    }

    for index, record in enumerate(records, start=2):
        column = get_column_letter(index)
        sheet.cell(4, index, record.get("hostname", ""))
        for label, field_name in field_map.items():
            sheet.cell(metric_rows[label], index, record.get(field_name))

        cpu_pct = record.get("cpu_pct")
        sheet.cell(21, index, cpu_pct / 100 if isinstance(cpu_pct, (int, float)) else None)
        sheet.cell(24, index, "Yes" if record.get("reachable") else "No")
        sheet.cell(26, index, " | ".join(build_collection_notes(record)))

        sheet.cell(17, index, f'=IF({column}20=0,"",{column}18/{column}20)')
        sheet.cell(7, index, f"=(1+COUNT({column}21:{column}23))/4")
        sheet.cell(
            6,
            index,
            (
                f'=IF({column}24<>"Yes",0,MAX(0,100'
                f'-IF(ISNUMBER({column}21),IF({column}21>=$B$32,$B$35,'
                f'IF({column}21>=$B$31,$B$34,0)),0)'
                f'-IF(ISNUMBER({column}22),MIN({column}22*$B$36,$B$38),0)'
                f'-IF(ISNUMBER({column}23),MIN({column}23*$B$37,$B$38),0)))'
            ),
        )
        sheet.cell(
            5,
            index,
            (
                f'=IF({column}24<>"Yes","Unreachable",'
                f'IF({column}7<$B$33,"Insufficient Data",'
                f'IF({column}6>=$B$39,"Healthy",'
                f'IF({column}6>=$B$40,"Watch","Critical"))))'
            ),
        )

        header = sheet.cell(4, index)
        header.fill = _fill(teal)
        header.font = Font(bold=True, color=white)
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[column].width = 24

        for row in range(5, 27):
            cell = sheet.cell(row, index)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin_gray)

    sheet["A29"] = "Health scoring method"
    sheet.merge_cells("A29:B29")
    sheet["A29"].fill = _fill(navy)
    sheet["A29"].font = Font(bold=True, color=white)
    sheet["A30"] = "Parameter"
    sheet["B30"] = "Value"
    settings = [
        ("CPU warning", CPU_WARNING_PCT / 100),
        ("CPU critical", CPU_CRITICAL_PCT / 100),
        ("Minimum data coverage", MIN_DATA_COVERAGE_PCT / 100),
        ("Warning penalty", WARNING_PENALTY),
        ("Critical penalty", CRITICAL_PENALTY),
        ("Environment alert penalty", ENVIRONMENT_ALERT_PENALTY),
        ("Err-disabled penalty", ERR_DISABLED_PENALTY),
        ("Maximum count penalty", MAX_COUNT_PENALTY),
        ("Healthy score", HEALTHY_SCORE),
        ("Watch score", WATCH_SCORE),
    ]
    for row, (label, value) in enumerate(settings, start=31):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
    for cell in sheet[30][:2]:
        cell.fill = _fill(teal)
        cell.font = Font(bold=True, color=white)
    for row in range(31, 34):
        sheet.cell(row, 2).number_format = "0%"

    sheet["A42"] = "Definitions"
    sheet["A42"].font = Font(bold=True)
    sheet.merge_cells(start_row=42, start_column=2, end_row=42, end_column=last_column)
    sheet.merge_cells(start_row=43, start_column=2, end_row=43, end_column=last_column)
    sheet["B42"] = (
        "Interface Usage = connected physical interfaces divided by all parsed physical "
        "interfaces. Unconnected, disabled, and err-disabled ports count as not connected."
    )
    sheet["B43"] = (
        "Health Score begins at 100 and applies visible CPU, environment, and "
        "err-disabled penalties. Missing metrics reduce Data Coverage instead of being "
        "treated as healthy or unhealthy."
    )
    sheet["B42"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet["B43"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[42].height = 58
    sheet.row_dimensions[43].height = 72

    sheet["A45"] = "Fleet summary"
    sheet.merge_cells("A45:B45")
    sheet["A45"].fill = _fill(navy)
    sheet["A45"].font = Font(bold=True, color=white)
    summary_labels = [
        "Devices",
        "Healthy",
        "Watch",
        "Critical",
        "Unreachable",
        "Insufficient Data",
        "Average Health Score",
        "Average Interface Usage",
    ]
    for row, label in enumerate(summary_labels, start=46):
        sheet.cell(row, 1, label)

    device_header_range = f"B4:{last_column_letter}4"
    status_range = f"B5:{last_column_letter}5"
    score_range = f"B6:{last_column_letter}6"
    usage_range = f"B17:{last_column_letter}17"
    sheet["B46"] = f"=COUNTA({device_header_range})"
    for row, status in zip(range(47, 52), summary_labels[1:6], strict=True):
        sheet.cell(row, 2, f'=COUNTIF({status_range},"{status}")')
    sheet["B52"] = f'=IFERROR(AVERAGE({score_range}),"")'
    sheet["B53"] = f'=IFERROR(AVERAGE({usage_range}),"")'
    sheet["B53"].number_format = "0%"

    sheet.column_dimensions["A"].width = 29
    sheet.column_dimensions["B"].width = max(sheet.column_dimensions["B"].width or 0, 24)
    sheet.row_dimensions[4].height = 30
    sheet.row_dimensions[26].height = 160
    sheet.freeze_panes = "A5"

    # Keep scoring mechanics available to formulas without showing them to
    # management users. Hidden rows collapse the method/definition section so
    # Fleet Summary follows the scorecard when scrolling.
    for row in range(29, 45):
        sheet.row_dimensions[row].hidden = True

    for column in range(2, last_column + 1):
        sheet.cell(6, column).number_format = "0"
        sheet.cell(7, column).number_format = "0%"
        sheet.cell(17, column).number_format = "0.0%"
        sheet.cell(21, column).number_format = "0.0%"
        sheet.cell(25, column).number_format = "yyyy-mm-dd hh:mm"

    status_cells = f"B5:{last_column_letter}5"
    for status, color in {
        "Healthy": green,
        "Watch": amber,
        "Critical": red,
        "Unreachable": red,
        "Insufficient Data": amber,
    }.items():
        sheet.conditional_formatting.add(
            status_cells,
            FormulaRule(formula=[f'B5="{status}"'], fill=_fill(color)),
        )
    sheet.conditional_formatting.add(
        f"B6:{last_column_letter}6",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=70,
            mid_color="FFEB84",
            end_type="num",
            end_value=100,
            end_color="63BE7B",
        ),
    )

    sheet["B17"].comment = Comment(
        "Connected physical interfaces divided by all parsed physical interfaces.",
        "Network Automation",
    )

    for row in range(29, 54):
        for column in range(1, min(last_column, 2) + 1):
            sheet.cell(row, column).alignment = Alignment(vertical="center", wrap_text=True)

    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_title_rows = "1:4"
    sheet.print_area = f"A1:{last_column_letter}26"

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    config_path = Path(args.config).expanduser()
    output_path = Path(args.output).expanduser()
    target_tag = str(args.tag).strip().casefold()

    if not config_path.is_file():
        console.print(
            f"[bold red]ERROR:[/] Existing Nornir config was not found: {config_path.resolve()}"
        )
        return 1
    if output_path.suffix.casefold() != ".xlsx":
        console.print("[bold red]ERROR:[/] --output must end in .xlsx")
        return 1
    if not target_tag:
        console.print("[bold red]ERROR:[/] --tag cannot be empty")
        return 1

    username = os.getenv("NORNIR_USERNAME")
    password = os.getenv("NORNIR_PASSWORD")
    if bool(username) != bool(password):
        console.print(
            "[bold red]ERROR:[/] Set both NORNIR_USERNAME and NORNIR_PASSWORD, or neither."
        )
        return 1

    try:
        nr = InitNornir(config_file=str(config_path))
    except Exception as exc:  # noqa: BLE001 - inventory plugins raise varied errors.
        console.print(f"[bold red]ERROR:[/] Could not initialize Nornir: {exc}")
        return 1

    if username and password:
        nr.inventory.defaults.username = username
        nr.inventory.defaults.password = password

    for host in nr.inventory.hosts.values():
        host.data["tag_slugs"] = normalize_tags(host.data.get("tags", []))
        if "netmiko" not in host.connection_options:
            host.connection_options["netmiko"] = ConnectionOptions(extras=dict(NETMIKO_EXTRAS))
        else:
            existing_extras = host.connection_options["netmiko"].extras or {}
            for key, value in NETMIKO_EXTRAS.items():
                existing_extras.setdefault(key, value)
            host.connection_options["netmiko"].extras = existing_extras

    targets = nr.filter(F(tag_slugs__contains=target_tag))
    console.print(f"Target tag: [bold]{target_tag}[/]")
    console.print(f"Matched devices: [bold]{len(targets.inventory.hosts)}[/]")

    if not targets.inventory.hosts:
        console.print("No devices matched the requested NetBox tag. No workbook was created.")
        return 0

    console.print("Collecting firmware, interface state, CPU, and environment health...")
    results = targets.run(name="Collect network health", task=collect_device_health)
    records = _extract_records(results, targets.inventory.hosts)
    create_health_workbook(records, target_tag, output_path)

    reachable = sum(bool(record.get("reachable")) for record in records)
    console.print(f"[bold green]Created:[/] {output_path.resolve()}")
    console.print(f"Reachable devices: {reachable}/{len(records)}")
    if reachable < len(records):
        console.print(
            "Unreachable devices remain in the workbook with their NetBox information and notes."
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
