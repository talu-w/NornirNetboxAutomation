"""Render one device's parsed interface rows into a formatted Excel workbook.

Ported verbatim from ``perform_backup.py``.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from bunnyauto.backup.interfaces import new_interface_row


def create_interface_workbook(
    hostname: str,
    rows: list[dict[str, Any]],
    command_errors: dict[str, str],
    output_path: Path,
) -> None:
    """Create a formatted Excel interface inventory for one device."""
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    interfaces = workbook.create_sheet("Interfaces")
    headers = list(new_interface_row("").keys())
    data_start_row = 5
    data_end_row = max(data_start_row, data_start_row + len(rows) - 1)

    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:D1")
    summary["A1"] = f"{hostname} Interface Backup"
    summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="17365D")
    summary["A1"].alignment = Alignment(horizontal="left")
    summary["A3"] = "Backup time"
    summary["B3"] = datetime.now()
    summary["B3"].number_format = "yyyy-mm-dd hh:mm:ss"
    summary["A5"] = "Metric"
    summary["B5"] = "Count"
    summary["A6"] = "Total interfaces"
    summary["B6"] = f"=COUNTA('Interfaces'!A{data_start_row}:A{data_end_row})"
    summary["A7"] = "Operationally up"
    summary["B7"] = f"=COUNTIF('Interfaces'!D{data_start_row}:D{data_end_row},\"up\")"
    summary["A8"] = "Operationally down"
    summary["B8"] = f"=COUNTIF('Interfaces'!D{data_start_row}:D{data_end_row},\"down\")"
    summary["A10"] = "Collection warnings"
    summary["A10"].font = Font(bold=True)
    if command_errors:
        for row_number, (command, error) in enumerate(command_errors.items(), start=11):
            summary.cell(row_number, 1, command)
            summary.cell(row_number, 2, error)
    else:
        summary["A11"] = "None"

    for cell in summary[5]:
        if cell.column <= 2:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F75B5")
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 42
    summary.freeze_panes = "A5"

    interfaces.sheet_view.showGridLines = False
    interfaces.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    interfaces.cell(1, 1, f"{hostname} Interface Details")
    interfaces.cell(1, 1).font = Font(size=18, bold=True, color="FFFFFF")
    interfaces.cell(1, 1).fill = PatternFill("solid", fgColor="17365D")
    interfaces.cell(2, 1, "Collected")
    interfaces.cell(2, 2, datetime.now())
    interfaces.cell(2, 2).number_format = "yyyy-mm-dd hh:mm:ss"

    for column, header in enumerate(headers, start=1):
        cell = interfaces.cell(4, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F75B5")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_number, row in enumerate(rows, start=data_start_row):
        for column, header in enumerate(headers, start=1):
            value = row[header]
            interfaces.cell(row_number, column, "N/A" if value is None or value == "" else value)

    if rows:
        table = Table(displayName="InterfaceInventory", ref=f"A4:Y{data_end_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        interfaces.add_table(table)

        interfaces.conditional_formatting.add(
            f"D{data_start_row}:D{data_end_row}",
            FormulaRule(
                formula=[f'LOWER(D{data_start_row})="up"'],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )
        interfaces.conditional_formatting.add(
            f"D{data_start_row}:D{data_end_row}",
            FormulaRule(
                formula=[f'LOWER(D{data_start_row})="down"'],
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )

    interfaces.freeze_panes = "A5"
    interfaces.auto_filter.ref = f"A4:Y{data_end_row}"
    widths = {
        "A": 20,
        "B": 16,
        "C": 22,
        "D": 16,
        "E": 32,
        "F": 14,
        "G": 14,
        "H": 18,
        "I": 18,
        "J": 18,
        "K": 18,
        "L": 16,
        "M": 16,
        "N": 16,
        "O": 16,
        "P": 18,
        "Q": 18,
        "R": 14,
        "S": 22,
        "T": 20,
        "U": 14,
        "V": 14,
        "W": 24,
        "X": 24,
        "Y": 28,
    }
    for column, width in widths.items():
        interfaces.column_dimensions[column].width = width

    for row in interfaces.iter_rows(min_row=data_start_row, max_row=data_end_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top")
    for column in range(8, 16):
        for row_number in range(data_start_row, data_end_row + 1):
            interfaces.cell(row_number, column).number_format = "#,##0"

    workbook.save(output_path)
