"""The management-facing health scorecard. Ported from ``network_simple_health_report.py``.

The Health Score is computed by Excel formulas (not Python) so the front-facing
report stays consistent; Data Coverage is tracked separately so missing metrics
are not scored as healthy or unhealthy.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Internal scoring values, stored in hidden worksheet rows.
CPU_WARNING_PCT = 75.0
CPU_CRITICAL_PCT = 90.0
MIN_DATA_COVERAGE_PCT = 60.0
WARNING_PENALTY = 10
CRITICAL_PENALTY = 25
ENVIRONMENT_ALERT_PENALTY = 10
ERR_DISABLED_PENALTY = 10
MAX_COUNT_PENALTY = 30
HEALTHY_SCORE = 85
WATCH_SCORE = 70
HEALTH_COMPONENT_COUNT = 4


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def build_collection_notes(record: Mapping[str, Any]) -> list[str]:
    """Collection warnings plus the conditions that reduced health."""
    notes = [str(note) for note in record.get("notes", []) if str(note).strip()]

    def add(note: str) -> None:
        if note not in notes:
            notes.append(note)

    if not record.get("reachable"):
        add("Health impact: device was unreachable; live health data could not be collected.")
        return notes

    cpu_pct = record.get("cpu_pct")
    if isinstance(cpu_pct, int | float):
        if cpu_pct >= CPU_CRITICAL_PCT:
            add(
                f"Health impact: CPU utilization {cpu_pct:.1f}% met or exceeded the "
                f"critical threshold of {CPU_CRITICAL_PCT:.0f}%."
            )
        elif cpu_pct >= CPU_WARNING_PCT:
            add(
                f"Health impact: CPU utilization {cpu_pct:.1f}% met or exceeded the "
                f"warning threshold of {CPU_WARNING_PCT:.0f}%."
            )

    environment_alerts = record.get("environment_alerts")
    if isinstance(environment_alerts, int | float) and environment_alerts > 0:
        add(
            f"Health impact: {int(environment_alerts)} environmental alert"
            f"{'s were' if environment_alerts != 1 else ' was'} detected."
        )

    err_disabled = record.get("err_disabled_interfaces")
    if isinstance(err_disabled, int | float) and err_disabled > 0:
        add(
            f"Health impact: {int(err_disabled)} physical interface"
            f"{'s are' if err_disabled != 1 else ' is'} err-disabled."
        )

    available_components = 1 + sum(
        isinstance(record.get(name), int | float)
        for name in ("cpu_pct", "environment_alerts", "err_disabled_interfaces")
    )
    coverage_pct = available_components / HEALTH_COMPONENT_COUNT * 100
    if coverage_pct < MIN_DATA_COVERAGE_PCT:
        add(
            f"Health impact: only {coverage_pct:.0f}% of health inputs were available; "
            "health data is incomplete."
        )
    return notes


def create_health_workbook(
    records: Sequence[Mapping[str, Any]],
    target_tag: str,
    output_path: Path,
) -> None:
    """Create one leadership scorecard with a device in each column."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Network Health"
    sheet.sheet_view.showGridLines = False

    navy = "17365D"
    teal = "147D74"
    green = "C6EFCE"
    amber = "FFEB9C"
    red = "FFC7CE"
    pale_blue = "D9EAF7"
    white = "FFFFFF"
    thin_gray = Side(style="thin", color="D9E1F2")

    last_column = max(2, len(records) + 1)
    last_column_letter = get_column_letter(last_column)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet["A1"] = "Nornir Networking Report - Networking Infrastructure Health"
    sheet["A1"].font = Font(size=18, bold=True, color=white)
    sheet["A1"].fill = _fill(navy)
    sheet["A1"].alignment = Alignment(vertical="center", wrap_text=True, shrink_to_fit=True)
    sheet.row_dimensions[1].height = 42

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    generated = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    sheet["A2"] = (
        f"Scope: devices from the existing Nornir/NetBox inventory tagged "
        f"'{target_tag}' | Generated {generated}"
    )
    sheet["A2"].font = Font(italic=True, color="44546A")
    sheet["A2"].fill = _fill(pale_blue)
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True, shrink_to_fit=True)
    sheet.row_dimensions[2].height = 32

    metric_rows = {
        "Health Status": 5,
        "Health Score": 6,
        "Data Coverage": 7,
        "Location": 8,
        "Site": 9,
        "Role": 10,
        "Model": 11,
        "Platform": 12,
        "NetBox Status": 13,
        "Primary IP": 14,
        "Serial Number": 15,
        "Current Firmware": 16,
        "Interface Usage": 17,
        "Connected Interfaces": 18,
        "Not Connected Interfaces": 19,
        "Total Interfaces": 20,
        "CPU Utilization": 21,
        "Environment Alerts": 22,
        "Err-disabled Interfaces": 23,
        "Reachable": 24,
        "Collected UTC": 25,
        "Collection Notes": 26,
    }

    sheet["A4"] = "Metric"
    for label, row in metric_rows.items():
        sheet.cell(row, 1, label)

    for row in range(4, 27):
        cell = sheet.cell(row, 1)
        cell.fill = _fill(navy)
        cell.font = Font(bold=True, color=white)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    field_map = {
        "Location": "location",
        "Site": "site",
        "Role": "role",
        "Model": "model",
        "Platform": "platform",
        "NetBox Status": "netbox_status",
        "Primary IP": "primary_ip",
        "Serial Number": "serial_number",
        "Current Firmware": "firmware",
        "Connected Interfaces": "connected_interfaces",
        "Not Connected Interfaces": "not_connected_interfaces",
        "Total Interfaces": "total_interfaces",
        "Environment Alerts": "environment_alerts",
        "Err-disabled Interfaces": "err_disabled_interfaces",
        "Collected UTC": "collected_at_utc",
    }

    for index, record in enumerate(records, start=2):
        column = get_column_letter(index)
        sheet.cell(4, index, record.get("hostname", ""))
        for label, field_name in field_map.items():
            sheet.cell(metric_rows[label], index, record.get(field_name))

        cpu_pct = record.get("cpu_pct")
        sheet.cell(21, index, cpu_pct / 100 if isinstance(cpu_pct, int | float) else None)
        sheet.cell(24, index, "Yes" if record.get("reachable") else "No")
        sheet.cell(26, index, " | ".join(build_collection_notes(record)))

        sheet.cell(17, index, f'=IF({column}20=0,"",{column}18/{column}20)')
        sheet.cell(7, index, f"=(1+COUNT({column}21:{column}23))/4")
        sheet.cell(
            6,
            index,
            (
                f'=IF({column}24<>"Yes",0,MAX(0,100'
                f"-IF(ISNUMBER({column}21),IF({column}21>=$B$32,$B$35,"
                f"IF({column}21>=$B$31,$B$34,0)),0)"
                f"-IF(ISNUMBER({column}22),MIN({column}22*$B$36,$B$38),0)"
                f"-IF(ISNUMBER({column}23),MIN({column}23*$B$37,$B$38),0)))"
            ),
        )
        sheet.cell(
            5,
            index,
            (
                f'=IF({column}24<>"Yes","Unreachable",'
                f'IF({column}7<$B$33,"Insufficient Data",'
                f'IF({column}6>=$B$39,"Healthy",'
                f'IF({column}6>=$B$40,"Watch","Critical"))))'
            ),
        )

        header = sheet.cell(4, index)
        header.fill = _fill(teal)
        header.font = Font(bold=True, color=white)
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[column].width = 24

        for row in range(5, 27):
            cell = sheet.cell(row, index)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin_gray)

    sheet["A29"] = "Health scoring method"
    sheet.merge_cells("A29:B29")
    sheet["A29"].fill = _fill(navy)
    sheet["A29"].font = Font(bold=True, color=white)
    sheet["A30"] = "Parameter"
    sheet["B30"] = "Value"
    scoring_settings = [
        ("CPU warning", CPU_WARNING_PCT / 100),
        ("CPU critical", CPU_CRITICAL_PCT / 100),
        ("Minimum data coverage", MIN_DATA_COVERAGE_PCT / 100),
        ("Warning penalty", WARNING_PENALTY),
        ("Critical penalty", CRITICAL_PENALTY),
        ("Environment alert penalty", ENVIRONMENT_ALERT_PENALTY),
        ("Err-disabled penalty", ERR_DISABLED_PENALTY),
        ("Maximum count penalty", MAX_COUNT_PENALTY),
        ("Healthy score", HEALTHY_SCORE),
        ("Watch score", WATCH_SCORE),
    ]
    for row, (label, value) in enumerate(scoring_settings, start=31):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
    for cell in sheet[30][:2]:
        cell.fill = _fill(teal)
        cell.font = Font(bold=True, color=white)
    for row in range(31, 34):
        sheet.cell(row, 2).number_format = "0%"

    sheet["A42"] = "Definitions"
    sheet["A42"].font = Font(bold=True)
    sheet.merge_cells(start_row=42, start_column=2, end_row=42, end_column=last_column)
    sheet.merge_cells(start_row=43, start_column=2, end_row=43, end_column=last_column)
    sheet["B42"] = (
        "Interface Usage = connected physical interfaces divided by all parsed physical "
        "interfaces. Unconnected, disabled, and err-disabled ports count as not connected."
    )
    sheet["B43"] = (
        "Health Score begins at 100 and applies visible CPU, environment, and "
        "err-disabled penalties. Missing metrics reduce Data Coverage instead of being "
        "treated as healthy or unhealthy."
    )
    sheet["B42"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet["B43"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[42].height = 58
    sheet.row_dimensions[43].height = 72

    sheet["A45"] = "Fleet summary"
    sheet.merge_cells("A45:B45")
    sheet["A45"].fill = _fill(navy)
    sheet["A45"].font = Font(bold=True, color=white)
    summary_labels = [
        "Devices",
        "Healthy",
        "Watch",
        "Critical",
        "Unreachable",
        "Insufficient Data",
        "Average Health Score",
        "Average Interface Usage",
    ]
    for row, label in enumerate(summary_labels, start=46):
        sheet.cell(row, 1, label)

    device_header_range = f"B4:{last_column_letter}4"
    status_range = f"B5:{last_column_letter}5"
    score_range = f"B6:{last_column_letter}6"
    usage_range = f"B17:{last_column_letter}17"
    sheet["B46"] = f"=COUNTA({device_header_range})"
    for row, status in zip(range(47, 52), summary_labels[1:6], strict=True):
        sheet.cell(row, 2, f'=COUNTIF({status_range},"{status}")')
    sheet["B52"] = f'=IFERROR(AVERAGE({score_range}),"")'
    sheet["B53"] = f'=IFERROR(AVERAGE({usage_range}),"")'
    sheet["B53"].number_format = "0%"

    sheet.column_dimensions["A"].width = 29
    sheet.column_dimensions["B"].width = max(sheet.column_dimensions["B"].width or 0, 24)
    sheet.row_dimensions[4].height = 30
    sheet.row_dimensions[26].height = 160
    sheet.freeze_panes = "A5"

    for row in range(29, 45):
        sheet.row_dimensions[row].hidden = True

    for column in range(2, last_column + 1):
        sheet.cell(6, column).number_format = "0"
        sheet.cell(7, column).number_format = "0%"
        sheet.cell(17, column).number_format = "0.0%"
        sheet.cell(21, column).number_format = "0.0%"
        sheet.cell(25, column).number_format = "yyyy-mm-dd hh:mm"

    status_cells = f"B5:{last_column_letter}5"
    for status, color in {
        "Healthy": green,
        "Watch": amber,
        "Critical": red,
        "Unreachable": red,
        "Insufficient Data": amber,
    }.items():
        sheet.conditional_formatting.add(
            status_cells, FormulaRule(formula=[f'B5="{status}"'], fill=_fill(color))
        )
    sheet.conditional_formatting.add(
        f"B6:{last_column_letter}6",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=70,
            mid_color="FFEB84",
            end_type="num",
            end_value=100,
            end_color="63BE7B",
        ),
    )

    sheet["B17"].comment = Comment(
        "Connected physical interfaces divided by all parsed physical interfaces.",
        "Network Automation",
    )

    for row in range(29, 54):
        for column in range(1, min(last_column, 2) + 1):
            sheet.cell(row, column).alignment = Alignment(vertical="center", wrap_text=True)

    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_title_rows = "1:4"
    sheet.print_area = f"A1:{last_column_letter}26"

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
