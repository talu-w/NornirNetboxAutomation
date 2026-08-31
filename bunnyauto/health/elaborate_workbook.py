"""The engineer-facing health workbook. Ported from ``network_elaborate_health_report.py``.

Two sheets: an Engineer Health overview (device per column, Excel-formula
scoring, hover-note detail cells) and an Issue Details table (one row per
actionable finding), plus a very-hidden Scoring sheet the formulas reference.
"""

from __future__ import annotations

import re
from collections.abc import Mapping, Sequence
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.table import Table, TableStyleInfo

# Internal scoring values, stored in the very-hidden "Scoring" sheet.
# NOTE: ERR_DISABLED_PENALTY is 2 here but 10 in health-simple. Left as-is
# pending the owner's call (CLAUDE.md issue #2) — do not "fix" silently.
CPU_WARNING_PCT = 75.0
CPU_CRITICAL_PCT = 90.0
MIN_DATA_COVERAGE_PCT = 60.0
WARNING_PENALTY = 10
CRITICAL_PENALTY = 25
ENVIRONMENT_ALERT_PENALTY = 10
ERR_DISABLED_PENALTY = 2
DEGRADED_ETHERCHANNEL_PENALTY = 10
DOWN_ETHERCHANNEL_PENALTY = 25
MAX_COUNT_PENALTY = 30
MAX_ETHERCHANNEL_PENALTY = 50
INTERFACE_QUALITY_PENALTY = 1
INTERFACE_INSTABILITY_PENALTY = 2
MAX_INTERFACE_PENALTY = 10
HEALTHY_SCORE = 85
WATCH_SCORE = 70
HEALTH_COMPONENT_COUNT = 6
ETHERCHANNEL_CRITICAL_DOWN_RATIO = 0.75


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _compact_detail_preview(value: Any, max_chars: int = 110) -> str:
    text = " ".join(str(value or "").split())
    if len(text) <= max_chars:
        return text
    suffix = "... [see note]"
    available = max(20, max_chars - len(suffix))
    preview = text[:available].rsplit(" ", 1)[0].rstrip(" ,;:")
    return f"{preview}{suffix}"


def _set_compact_detail_cell(cell: Any, full_text: Any, *, max_chars: int = 110) -> None:
    """Show a short cell preview and keep the full text in an Excel note."""
    text = str(full_text or "").strip()
    cell.value = _compact_detail_preview(text, max_chars=max_chars)
    if not text:
        return
    note_text = re.sub(r"\s*\|\s*", "\n", text)
    note = Comment(
        "Full unabridged details:\n\n"
        f"{note_text}\n\n"
        "Tip: hover over or select this noted cell in Excel to review the full text.",
        "Network Automation",
    )
    note.width = 620
    note.height = 360
    cell.comment = note


def _etherchannel_down_ratio(record: Mapping[str, Any]) -> float | None:
    total = record.get("etherchannels_total")
    down = record.get("etherchannels_down")
    if not isinstance(total, int | float) or not isinstance(down, int | float):
        channels = record.get("etherchannels", [])
        if not isinstance(channels, Sequence):
            return None
        reported = [channel for channel in channels if isinstance(channel, Mapping)]
        total = len(reported)
        down = sum(channel.get("state") == "Down" for channel in reported)
    if total <= 0:
        return None
    return max(0.0, float(down)) / float(total)


def _etherchannel_outage_is_critical(record: Mapping[str, Any]) -> bool:
    ratio = _etherchannel_down_ratio(record)
    return ratio is not None and ratio >= ETHERCHANNEL_CRITICAL_DOWN_RATIO


def build_collection_notes(record: Mapping[str, Any]) -> list[str]:
    """Collection warnings plus the conditions that reduced health."""
    notes = [str(note) for note in record.get("notes", []) if str(note).strip()]

    def add(note: str) -> None:
        if note not in notes:
            notes.append(note)

    if not record.get("reachable"):
        add("Health impact: device was unreachable; live health data could not be collected.")
        return notes

    cpu_pct = record.get("cpu_pct")
    if isinstance(cpu_pct, int | float):
        if cpu_pct >= CPU_CRITICAL_PCT:
            add(
                f"Health impact: CPU utilization {cpu_pct:.1f}% met or exceeded the "
                f"critical threshold of {CPU_CRITICAL_PCT:.0f}%."
            )
        elif cpu_pct >= CPU_WARNING_PCT:
            add(
                f"Health impact: CPU utilization {cpu_pct:.1f}% met or exceeded the "
                f"warning threshold of {CPU_WARNING_PCT:.0f}%."
            )

    environment_alerts = record.get("environment_alerts")
    if isinstance(environment_alerts, int | float) and environment_alerts > 0:
        add(
            f"Health impact: {int(environment_alerts)} environmental alert"
            f"{'s were' if environment_alerts != 1 else ' was'} detected."
        )

    err_disabled = record.get("err_disabled_interfaces", [])
    if isinstance(err_disabled, Sequence) and err_disabled:
        details = ", ".join(
            f"{issue.get('interface', 'Unknown')} ({issue.get('reason', 'Reason not reported')})"
            for issue in err_disabled
            if isinstance(issue, Mapping)
        )
        add(f"Health impact: err-disabled: {details}.")

    port_security_issues = record.get("port_security_issues", [])
    if isinstance(port_security_issues, Sequence) and port_security_issues:
        details = ", ".join(
            f"{issue.get('interface', 'Unknown')} "
            f"({issue.get('violation_count', 0)} violations / {issue.get('action', 'Unknown')})"
            for issue in port_security_issues
            if isinstance(issue, Mapping)
        )
        add(f"Informational: port security: {details}.")

    interface_quality = record.get("interface_quality_issues", [])
    if isinstance(interface_quality, Sequence) and interface_quality:
        quality_details: list[str] = []
        for issue in interface_quality[:8]:
            if not isinstance(issue, Mapping):
                continue
            parts: list[str] = []
            error_total = issue.get("error_total")
            if isinstance(error_total, int | float) and error_total > 0:
                parts.append(f"{int(error_total):,} cumulative error counters")
            if issue.get("interface_resets"):
                parts.append(f"interface resets {issue.get('interface_resets')}")
            if issue.get("carrier_transitions"):
                parts.append(f"carrier transitions {issue.get('carrier_transitions')}")
            if issue.get("recent_link_flap"):
                parts.append(f"last link flap {issue.get('last_link_flapped', 'recent')} ago")
            quality_details.append(f"{issue.get('interface', 'Unknown')} ({', '.join(parts)})")
        if len(interface_quality) > 8:
            quality_details.append(
                f"{len(interface_quality) - 8} additional interfaces; see Issue Details"
            )
        add("Health impact: interface quality/instability: " + "; ".join(quality_details) + ".")

    etherchannels = record.get("etherchannels", [])
    if isinstance(etherchannels, Sequence):
        down_ratio = _etherchannel_down_ratio(record)
        if _etherchannel_outage_is_critical(record) and down_ratio is not None:
            add(
                "Health impact: "
                f"{record.get('etherchannels_down', 0)}/"
                f"{record.get('etherchannels_total', 0)} EtherChannels are down "
                f"({down_ratio:.0%}), meeting the critical outage threshold of "
                f"{ETHERCHANNEL_CRITICAL_DOWN_RATIO:.0%}."
            )
        for channel in etherchannels:
            if not isinstance(channel, Mapping) or channel.get("state") == "Up":
                continue
            members = channel.get("members", [])
            bundled = channel.get("bundled_members", [])
            problem = channel.get("problem_members", [])
            detail = (
                f"Health impact: {channel.get('name', 'Unknown channel')} "
                f"{channel.get('state', 'Unknown')}; {len(bundled)}/{len(members)} members bundled"
            )
            if problem:
                detail += f"; affected: {', '.join(map(str, problem))}"
            risks = channel.get("risk_reasons", [])
            if risks:
                concise_risks = [
                    "minimum links not met" if "minimum links" in str(risk) else str(risk)
                    for risk in risks
                ]
                detail += f"; {', '.join(dict.fromkeys(concise_risks))}"
            add(f"{detail}.")

    security_exceptions = record.get("security_control_exceptions", [])
    if isinstance(security_exceptions, Sequence) and security_exceptions:
        object_names = [
            f"{issue.get('category', 'Security')} / {issue.get('object_name', 'Unknown')}"
            for issue in security_exceptions[:6]
            if isinstance(issue, Mapping)
        ]
        drop_total = sum(
            int(issue.get("count", 0))
            for issue in security_exceptions
            if isinstance(issue, Mapping) and isinstance(issue.get("count", 0), int | float)
        )
        add(
            f"Informational: {len(security_exceptions)} security/control-plane "
            f"exceptions with {drop_total:,} cumulative drops: {', '.join(object_names)}."
        )

    available_components = 1 + sum(
        isinstance(record.get(name), int | float)
        for name in (
            "cpu_pct",
            "environment_alerts",
            "err_disabled_count",
            "etherchannels_total",
            "interfaces_with_errors",
        )
    )
    coverage_pct = available_components / HEALTH_COMPONENT_COUNT * 100
    if coverage_pct < MIN_DATA_COVERAGE_PCT:
        add(
            f"Health impact: only {coverage_pct:.0f}% of health inputs were available; "
            "health data is incomplete."
        )
    return notes


# ---------------------------------------------------------------------------
# detail-cell text builders
# ---------------------------------------------------------------------------


def _interface_issue_text(record: Mapping[str, Any]) -> str:
    issues = record.get("err_disabled_interfaces", [])
    if not isinstance(issues, Sequence):
        return ""
    return ", ".join(
        f"{issue.get('interface', 'Unknown')} ({issue.get('reason', 'Reason not reported')})"
        for issue in issues
        if isinstance(issue, Mapping)
    )


def _port_security_text(record: Mapping[str, Any]) -> str:
    issues = record.get("port_security_issues", [])
    if not isinstance(issues, Sequence):
        return ""
    return ", ".join(
        f"{issue.get('interface', 'Unknown')}: "
        f"{issue.get('violation_count', 0)} violations / {issue.get('action', 'Unknown')}"
        for issue in issues
        if isinstance(issue, Mapping)
    )


def _interface_quality_text(record: Mapping[str, Any]) -> str:
    issues = record.get("interface_quality_issues", [])
    if not isinstance(issues, Sequence):
        return ""
    details: list[str] = []
    for issue in issues:
        if not isinstance(issue, Mapping):
            continue
        parts: list[str] = []
        counters = issue.get("counters", {})
        if isinstance(counters, Mapping):
            values = [
                f"{name.replace('_', ' ')}={value}"
                for name, value in counters.items()
                if isinstance(value, int | float) and value > 0
            ]
            if values:
                parts.append(", ".join(values))
        if issue.get("interface_resets"):
            parts.append(f"resets={issue.get('interface_resets')}")
        if issue.get("carrier_transitions"):
            parts.append(f"carrier transitions={issue.get('carrier_transitions')}")
        if issue.get("recent_link_flap"):
            parts.append(f"last flap={issue.get('last_link_flapped', 'recent')} ago")
        details.append(f"{issue.get('interface', 'Unknown')}: {'; '.join(parts)}")
    return " | ".join(details)


def _etherchannel_issue_text(record: Mapping[str, Any]) -> str:
    channels = record.get("etherchannels", [])
    if not isinstance(channels, Sequence):
        return ""
    details: list[str] = []
    for channel in channels:
        if not isinstance(channel, Mapping) or channel.get("state") == "Up":
            continue
        members = channel.get("members", [])
        bundled = channel.get("bundled_members", [])
        problem = channel.get("problem_members", [])
        text = (
            f"{channel.get('name', 'Unknown')} {channel.get('state', 'Unknown')} "
            f"({len(bundled)}/{len(members)} bundled)"
        )
        if problem:
            text += f": {', '.join(map(str, problem))}"
        standby = channel.get("standby_members", [])
        if standby:
            text += f"; hot standby: {', '.join(map(str, standby))}"
        if channel.get("min_links") is not None:
            text += f"; minimum links: {channel.get('min_links')}"
        risks = channel.get("risk_reasons", [])
        if risks:
            text += f"; risks: {', '.join(map(str, risks))}"
        details.append(text)
    return " | ".join(details)


def _etherchannel_member_state_text(record: Mapping[str, Any]) -> str:
    channels = record.get("etherchannels", [])
    if not isinstance(channels, Sequence):
        return ""
    details: list[str] = []
    for channel in channels:
        if not isinstance(channel, Mapping):
            continue
        members = channel.get("member_details", [])
        member_text = ", ".join(
            f"{member.get('interface', 'Unknown')}={member.get('state', 'Unknown')}"
            for member in members
            if isinstance(member, Mapping)
        )
        text = (
            f"{channel.get('name', 'Unknown')} [{channel.get('protocol', 'Unknown')}] "
            f"{channel.get('state', 'Unknown')}"
        )
        if channel.get("min_links") is not None:
            text += f", min-links {channel.get('min_links')}"
        if member_text:
            text += f": {member_text}"
        details.append(text)
    return " | ".join(details)


def _security_control_text(record: Mapping[str, Any]) -> str:
    issues = record.get("security_control_exceptions", [])
    if not isinstance(issues, Sequence):
        return ""
    return " | ".join(
        f"{issue.get('category', 'Security')} / "
        f"{issue.get('object_name', 'Unknown')}: {issue.get('count', 0)} — "
        f"{issue.get('details', '')}"
        for issue in issues
        if isinstance(issue, Mapping)
    )


def build_issue_rows(records: Sequence[Mapping[str, Any]]) -> list[list[Any]]:
    """One engineer-actionable row for every detected issue."""
    rows: list[list[Any]] = []
    for record in records:
        hostname = str(record.get("hostname", "Unknown"))
        location = str(record.get("location", ""))
        primary_ip = str(record.get("primary_ip", ""))
        collected = record.get("collected_at_utc")

        def add(
            severity: str,
            category: str,
            object_name: str,
            state: str,
            details: str,
            *,
            _hostname: str = hostname,
            _location: str = location,
            _primary_ip: str = primary_ip,
            _collected: Any = collected,
        ) -> None:
            rows.append(
                [
                    _hostname,
                    severity,
                    category,
                    object_name,
                    state,
                    details,
                    _location,
                    _primary_ip,
                    _collected,
                ]
            )

        if not record.get("reachable"):
            add(
                "Critical",
                "Reachability",
                hostname,
                "Unreachable",
                "SSH collection did not complete.",
            )
            continue

        cpu_pct = record.get("cpu_pct")
        if isinstance(cpu_pct, int | float) and cpu_pct >= CPU_WARNING_PCT:
            severity = "Critical" if cpu_pct >= CPU_CRITICAL_PCT else "Warning"
            add(
                severity,
                "CPU",
                hostname,
                f"{cpu_pct:.1f}%",
                "CPU utilization exceeded the configured health threshold.",
            )

        environment_alerts = record.get("environment_alerts")
        if isinstance(environment_alerts, int | float) and environment_alerts > 0:
            add(
                "Warning",
                "Environment",
                hostname,
                f"{int(environment_alerts)} alerts",
                "Environmental command output contained active alert terms.",
            )

        for issue in record.get("err_disabled_interfaces", []):
            if isinstance(issue, Mapping):
                add(
                    "Warning",
                    "Err-disabled",
                    str(issue.get("interface", "Unknown")),
                    "Err-disabled",
                    f"Reason: {issue.get('reason', 'Reason not reported')}",
                )

        for issue in record.get("port_security_issues", []):
            if isinstance(issue, Mapping):
                add(
                    "Info",
                    "Port Security",
                    str(issue.get("interface", "Unknown")),
                    str(issue.get("action", "Violation")),
                    f"Violation counter: {issue.get('violation_count', 0)}",
                )

        for issue in record.get("interface_quality_issues", []):
            if not isinstance(issue, Mapping):
                continue
            error_total = issue.get("error_total", 0)
            error_total = error_total if isinstance(error_total, int | float) else 0
            states: list[str] = []
            detail_parts: list[str] = []
            counters = issue.get("counters", {})
            if error_total > 0:
                states.append("Quality errors")
            if isinstance(counters, Mapping):
                detail_parts.extend(
                    f"{name.replace('_', ' ')}={value}"
                    for name, value in counters.items()
                    if isinstance(value, int | float) and value > 0
                )
            if issue.get("interface_resets"):
                states.append("Instability")
                detail_parts.append(f"resets={issue.get('interface_resets')}")
            if issue.get("carrier_transitions"):
                if "Instability" not in states:
                    states.append("Instability")
                detail_parts.append(f"carrier transitions={issue.get('carrier_transitions')}")
            if issue.get("recent_link_flap"):
                if "Instability" not in states:
                    states.append("Instability")
                detail_parts.append(
                    f"last link flap={issue.get('last_link_flapped', 'recent')} ago"
                )
            add(
                "Warning",
                "Interface Quality",
                str(issue.get("interface", "Unknown")),
                " / ".join(states) or "Exception",
                "; ".join(detail_parts) + "; counters are cumulative unless cleared on the device",
            )

        etherchannel_outage_is_critical = _etherchannel_outage_is_critical(record)
        for channel in record.get("etherchannels", []):
            if not isinstance(channel, Mapping) or channel.get("state") == "Up":
                continue
            state = str(channel.get("state", "Unknown"))
            problem = channel.get("problem_members", [])
            bundled = channel.get("bundled_members", [])
            standby = channel.get("standby_members", [])
            members = channel.get("members", [])
            details = (
                f"Protocol {channel.get('protocol', 'Unknown')}; "
                f"{len(bundled)}/{len(members)} members bundled"
            )
            if problem:
                details += f"; affected: {', '.join(map(str, problem))}"
            if standby:
                details += f"; hot standby: {', '.join(map(str, standby))}"
            if channel.get("min_links") is not None:
                details += f"; minimum links: {channel.get('min_links')}"
            risks = channel.get("risk_reasons", [])
            if risks:
                details += f"; risks: {', '.join(map(str, risks))}"
            add(
                "Critical" if state == "Down" and etherchannel_outage_is_critical else "Warning",
                "EtherChannel",
                str(channel.get("name", "Unknown")),
                state,
                details,
            )

        for issue in record.get("security_control_exceptions", []):
            if isinstance(issue, Mapping):
                add(
                    "Warning",
                    str(issue.get("category", "Security / Control Plane")),
                    str(issue.get("object_name", "Unknown")),
                    f"{issue.get('count', 0)} drops",
                    str(issue.get("details", "Cumulative exception counter")),
                )

        for note in record.get("notes", []):
            add("Warning", "Telemetry", hostname, "Collection note", str(note))
    return rows


def create_elaborate_health_workbook(
    records: Sequence[Mapping[str, Any]],
    target_tag: str,
    output_path: Path,
) -> None:
    """Create an engineer overview plus a normalized issue-detail worksheet."""
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Engineer Health"
    overview.sheet_view.showGridLines = False
    details = workbook.create_sheet("Issue Details")
    details.sheet_view.showGridLines = False
    scoring = workbook.create_sheet("Scoring")

    navy = "17365D"
    teal = "147D74"
    green = "C6EFCE"
    amber = "FFEB9C"
    red = "FFC7CE"
    pale_blue = "D9EAF7"
    pale_teal = "DDEBF7"
    white = "FFFFFF"
    gray = "44546A"
    thin_gray = Side(style="thin", color="D9E1F2")

    last_column = max(2, len(records) + 1)
    last_column_letter = get_column_letter(last_column)
    generated = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")

    overview.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    overview["A1"] = "Nornir Engineer Report — Network Infrastructure Health"
    overview["A1"].font = Font(size=18, bold=True, color=white)
    overview["A1"].fill = _fill(navy)
    overview["A1"].alignment = Alignment(vertical="center", wrap_text=True, shrink_to_fit=True)
    overview.row_dimensions[1].height = 42

    overview.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    overview["A2"] = f"Scope: Nornir/NetBox devices tagged '{target_tag}' | Generated {generated}"
    overview["A2"].font = Font(italic=True, color=gray)
    overview["A2"].fill = _fill(pale_blue)
    overview["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    overview.row_dimensions[2].height = 30

    metric_rows = {
        "Health Status": 5,
        "Health Score": 6,
        "Data Coverage": 7,
        "Location": 8,
        "Site": 9,
        "Role": 10,
        "Model": 11,
        "Platform": 12,
        "NetBox Status": 13,
        "Primary IP": 14,
        "Serial Number": 15,
        "Current Firmware": 16,
        "Reachable": 17,
        "Collected UTC": 18,
        "Interface State / Use": 20,
        "Interface Usage": 21,
        "Connected Interfaces": 22,
        "Not Connected Interfaces": 23,
        "Total Interfaces": 24,
        "Err-disabled Interfaces": 25,
        "Err-disabled Details": 26,
        "Port-security Interfaces": 27,
        "Port-security Violations": 28,
        "Port-security Details": 29,
        "EtherChannel Health": 31,
        "EtherChannels Total": 32,
        "EtherChannels Healthy": 33,
        "EtherChannels Degraded": 34,
        "EtherChannels Down": 35,
        "Member Availability": 36,
        "Bundled Members": 37,
        "Total Members": 38,
        "Degraded / Down Details": 39,
        "Min-link Risks": 40,
        "Hot-standby Members": 41,
        "Member State Details": 42,
        "System Environmental Health": 44,
        "CPU Utilization": 45,
        "Environment Alerts": 46,
        "Interface Health": 48,
        "Interfaces with Errors": 49,
        "Error Counter Sum": 50,
        "Instability Indicators": 51,
        "Interface Quality Details": 52,
        "Security Details": 54,
        "Security / Control Plane Details": 55,
        "Engineering Notes": 57,
    }
    section_rows = {20, 31, 44, 48, 54}

    overview["A4"] = "Metric"
    for label, row in metric_rows.items():
        overview.cell(row, 1, label)

    for row in range(4, 58):
        cell = overview.cell(row, 1)
        cell.fill = _fill(teal) if row in section_rows else _fill(navy)
        cell.font = Font(bold=True, color=white)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in section_rows:
        for column in range(2, last_column + 1):
            cell = overview.cell(row, column)
            cell.fill = _fill(pale_teal)
            cell.border = Border(bottom=thin_gray)

    field_map = {
        "Location": "location",
        "Site": "site",
        "Role": "role",
        "Model": "model",
        "Platform": "platform",
        "NetBox Status": "netbox_status",
        "Primary IP": "primary_ip",
        "Serial Number": "serial_number",
        "Current Firmware": "firmware",
        "Collected UTC": "collected_at_utc",
        "Connected Interfaces": "connected_interfaces",
        "Not Connected Interfaces": "not_connected_interfaces",
        "Total Interfaces": "total_interfaces",
        "Err-disabled Interfaces": "err_disabled_count",
        "Port-security Interfaces": "port_security_interfaces",
        "Port-security Violations": "port_security_violations",
        "EtherChannels Total": "etherchannels_total",
        "EtherChannels Healthy": "etherchannels_healthy",
        "EtherChannels Degraded": "etherchannels_degraded",
        "EtherChannels Down": "etherchannels_down",
        "Bundled Members": "etherchannel_members_bundled",
        "Total Members": "etherchannel_members_total",
        "Environment Alerts": "environment_alerts",
        "Interfaces with Errors": "interfaces_with_errors",
        "Error Counter Sum": "interface_error_total",
        "Instability Indicators": "interface_instability_count",
        "Min-link Risks": "etherchannel_min_link_risks",
        "Hot-standby Members": "etherchannel_members_standby",
    }

    for index, record in enumerate(records, start=2):
        column = get_column_letter(index)
        overview.cell(4, index, record.get("hostname", ""))
        for label, field_name in field_map.items():
            overview.cell(metric_rows[label], index, record.get(field_name))

        overview.cell(17, index, "Yes" if record.get("reachable") else "No")
        for row, full_text, max_chars in (
            (26, _interface_issue_text(record), 100),
            (29, _port_security_text(record), 100),
            (39, _etherchannel_issue_text(record), 120),
            (42, _etherchannel_member_state_text(record), 120),
            (52, _interface_quality_text(record), 120),
            (55, _security_control_text(record), 120),
            (57, "\n".join(build_collection_notes(record)), 150),
        ):
            _set_compact_detail_cell(overview.cell(row, index), full_text, max_chars=max_chars)

        cpu_pct = record.get("cpu_pct")
        overview.cell(45, index, cpu_pct / 100 if isinstance(cpu_pct, int | float) else None)
        overview.cell(21, index, f'=IF({column}24=0,"",{column}22/{column}24)')
        overview.cell(36, index, f'=IF({column}38=0,"",{column}37/{column}38)')
        overview.cell(
            7,
            index,
            (
                f'=(IF({column}17="Yes",1,0)+'
                f"COUNT({column}45,{column}46,{column}25,"
                f"{column}32,{column}49))/"
                f"{HEALTH_COMPONENT_COUNT}"
            ),
        )
        overview.cell(
            6,
            index,
            (
                f'=IF({column}17<>"Yes",0,MAX(0,100'
                f"-IF(ISNUMBER({column}45),IF({column}45>='Scoring'!$B$3,"
                f"'Scoring'!$B$6,IF({column}45>='Scoring'!$B$2,"
                f"'Scoring'!$B$5,0)),0)"
                f"-IF(ISNUMBER({column}46),MIN({column}46*'Scoring'!$B$7,"
                f"'Scoring'!$B$11),0)"
                f"-MIN(IF(ISNUMBER({column}25),{column}25*'Scoring'!$B$8,0)"
                f"+IF(ISNUMBER({column}49),{column}49*'Scoring'!$B$15,0)"
                f"+IF(ISNUMBER({column}51),{column}51*'Scoring'!$B$16,0),"
                f"'Scoring'!$B$17)"
                f"-MIN(IF(ISNUMBER({column}34),{column}34*'Scoring'!$B$9,0)"
                f"+IF(ISNUMBER({column}35),{column}35*'Scoring'!$B$10,0),"
                f"'Scoring'!$B$12)))"
            ),
        )
        overview.cell(
            5,
            index,
            (
                f'=IF({column}17<>"Yes","Unreachable",'
                f"IF({column}7<'Scoring'!$B$4,\"Insufficient Data\","
                f"IF(OR(AND(ISNUMBER({column}45),"
                f"{column}45>='Scoring'!$B$3),"
                f"AND({column}32>0,{column}35>="
                f"{column}32*'Scoring'!$B$18)),\"Critical\","
                f"IF(OR(AND(ISNUMBER({column}45),"
                f"{column}45>='Scoring'!$B$2),{column}46>0,"
                f"{column}25>0,{column}34>0,{column}35>0,"
                f"{column}40>0,{column}6<'Scoring'!$B$13),"
                f'"Watch","Healthy"))))'
            ),
        )

        header = overview.cell(4, index)
        header.fill = _fill(teal)
        header.font = Font(bold=True, color=white)
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        overview.column_dimensions[column].width = 29

        for row in range(5, 58):
            if row in section_rows:
                continue
            cell = overview.cell(row, index)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin_gray)

    overview.column_dimensions["A"].width = 31
    overview.row_dimensions[4].height = 32
    for row in (26, 29):
        overview.row_dimensions[row].height = 45
    for row in (39, 42, 52, 55):
        overview.row_dimensions[row].height = 60
    overview.row_dimensions[57].height = 72
    overview.freeze_panes = "A5"

    for column in range(2, last_column + 1):
        overview.cell(6, column).number_format = "0"
        overview.cell(7, column).number_format = "0%"
        overview.cell(18, column).number_format = "yyyy-mm-dd hh:mm"
        overview.cell(21, column).number_format = "0.0%"
        overview.cell(36, column).number_format = "0.0%"
        overview.cell(45, column).number_format = "0.0%"

    status_cells = f"B5:{last_column_letter}5"
    for status, color in {
        "Healthy": green,
        "Watch": amber,
        "Critical": red,
        "Unreachable": red,
        "Insufficient Data": amber,
    }.items():
        overview.conditional_formatting.add(
            status_cells, FormulaRule(formula=[f'B5="{status}"'], fill=_fill(color))
        )
    overview.conditional_formatting.add(
        f"B6:{last_column_letter}6",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=70,
            mid_color="FFEB84",
            end_type="num",
            end_value=100,
            end_color="63BE7B",
        ),
    )
    overview.conditional_formatting.add(
        f"B35:{last_column_letter}35",
        FormulaRule(
            formula=["AND(B32>0,B35>=B32*'Scoring'!$B$18)"],
            fill=_fill(red),
            stopIfTrue=True,
        ),
    )
    for row, color in (
        (25, amber),
        (34, amber),
        (35, amber),
        (46, amber),
        (49, amber),
        (51, amber),
        (40, amber),
    ):
        overview.conditional_formatting.add(
            f"B{row}:{last_column_letter}{row}",
            FormulaRule(formula=[f"B{row}>0"], fill=_fill(color)),
        )

    overview["A21"].comment = Comment(
        "Connected physical interfaces divided by all parsed physical interfaces.",
        "Network Automation",
    )
    overview["A36"].comment = Comment(
        "Bundled EtherChannel members divided by all parsed channel members.",
        "Network Automation",
    )
    overview["A50"].comment = Comment(
        "Sum of parsed cumulative error/drop counters; use Issue Details to see "
        "the counter type and interface.",
        "Network Automation",
    )
    overview["A57"].comment = Comment(
        "Device cells in detail and Engineering Notes rows show a short preview. "
        "Hover over or select the noted device cell to read its complete text.",
        "Network Automation",
    )

    summary_start = 60
    overview[f"A{summary_start}"] = "Fleet Summary"
    overview.merge_cells(
        start_row=summary_start, start_column=1, end_row=summary_start, end_column=2
    )
    overview[f"A{summary_start}"].fill = _fill(navy)
    overview[f"A{summary_start}"].font = Font(bold=True, color=white)
    summary_labels = [
        "Devices",
        "Healthy",
        "Watch",
        "Critical",
        "Unreachable",
        "Insufficient Data",
        "Average Health Score",
        "Total EtherChannels",
        "Degraded EtherChannels",
        "Down EtherChannels",
        "Err-disabled Interfaces",
        "Port-security Interfaces",
        "Interfaces with Quality Errors",
        "Instability Indicators",
        "EtherChannel Min-link Risks",
    ]
    for row, label in enumerate(summary_labels, start=summary_start + 1):
        overview.cell(row, 1, label)

    status_range = f"B5:{last_column_letter}5"
    overview.cell(summary_start + 1, 2, f"=COUNTA(B4:{last_column_letter}4)")
    for offset, status in enumerate(summary_labels[1:6], start=2):
        overview.cell(summary_start + offset, 2, f'=COUNTIF({status_range},"{status}")')
    overview.cell(summary_start + 7, 2, f'=IFERROR(AVERAGE(B6:{last_column_letter}6),"")')
    for offset, source_row in enumerate((32, 34, 35, 25, 27, 49, 51, 40), start=8):
        overview.cell(
            summary_start + offset,
            2,
            f"=SUM(B{source_row}:{last_column_letter}{source_row})",
        )
    overview.cell(summary_start + 7, 2).number_format = "0"
    for row in range(summary_start + 1, summary_start + len(summary_labels) + 1):
        overview.cell(row, 1).font = Font(bold=True)
        overview.cell(row, 1).fill = _fill(pale_blue)
        overview.cell(row, 1).alignment = Alignment(wrap_text=True)
        overview.cell(row, 2).alignment = Alignment(horizontal="right")

    overview.sheet_properties.pageSetUpPr.fitToPage = True
    overview.page_setup.orientation = "landscape"
    overview.page_setup.fitToWidth = 1
    overview.page_setup.fitToHeight = 0
    overview.print_title_rows = "1:4"
    overview.print_area = f"A1:{last_column_letter}{summary_start + len(summary_labels)}"
    overview.row_breaks.append(Break(id=56))
    overview.row_breaks.append(Break(id=summary_start - 1))

    details.merge_cells("A1:I1")
    details["A1"] = "Engineer Issue Details"
    details["A1"].font = Font(size=18, bold=True, color=white)
    details["A1"].fill = _fill(navy)
    details["A1"].alignment = Alignment(vertical="center")
    details.row_dimensions[1].height = 42
    details.merge_cells("A2:I2")
    details["A2"] = (
        "One row per actionable device, interface, security, "
        f"EtherChannel, or telemetry issue | Generated {generated}"
    )
    details["A2"].font = Font(italic=True, color=gray)
    details["A2"].fill = _fill(pale_blue)
    details["A2"].alignment = Alignment(wrap_text=True)
    details.row_dimensions[2].height = 30

    headers = [
        "Device",
        "Severity",
        "Category",
        "Object",
        "State",
        "Details",
        "Location",
        "Primary IP",
        "Collected UTC",
        "Full Details",
    ]
    for column, header in enumerate(headers, start=1):
        details.cell(4, column, header)
    issue_rows = build_issue_rows(records)
    if not issue_rows:
        issue_rows = [
            [
                "Fleet",
                "Info",
                "Health",
                "All devices",
                "No active issues",
                "No engineer-actionable conditions were detected.",
                "",
                "",
                None,
            ]
        ]
    for row_index, row in enumerate(issue_rows, start=5):
        full_detail = str(row[5] or "")
        for column_index, value in enumerate(row, start=1):
            cell = details.cell(row_index, column_index)
            if column_index == 6:
                _set_compact_detail_cell(cell, full_detail, max_chars=105)
            else:
                cell.value = value
        details.cell(row_index, 10, full_detail)
        details.row_dimensions[row_index].height = 36

    detail_last_row = 4 + len(issue_rows)
    detail_table = Table(displayName="EngineerIssues", ref=f"A4:J{detail_last_row}")
    detail_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    details.add_table(detail_table)
    details.freeze_panes = "A5"
    for letter, width in {
        "A": 24,
        "B": 12,
        "C": 18,
        "D": 20,
        "E": 18,
        "F": 52,
        "G": 20,
        "H": 18,
        "I": 20,
        "J": 90,
    }.items():
        details.column_dimensions[letter].width = width
    details.column_dimensions["J"].hidden = True
    for row in range(5, detail_last_row + 1):
        details.cell(row, 9).number_format = "yyyy-mm-dd hh:mm"
    for row in details.iter_rows(min_row=5, max_row=detail_last_row, max_col=10):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    details["F4"].comment = Comment(
        "Visible Details cells are compact previews. Hover over or select a noted "
        "cell for the complete text. The same unabridged value is retained in the "
        "hidden Full Details column.",
        "Network Automation",
    )
    details.conditional_formatting.add(
        f"A5:I{detail_last_row}",
        FormulaRule(formula=['$B5="Critical"'], fill=_fill(red)),
    )
    details.conditional_formatting.add(
        f"A5:I{detail_last_row}",
        FormulaRule(formula=['$B5="Warning"'], fill=_fill(amber)),
    )
    details.sheet_properties.pageSetUpPr.fitToPage = True
    details.page_setup.orientation = "landscape"
    details.page_setup.fitToWidth = 1
    details.page_setup.fitToHeight = 0
    details.print_title_rows = "1:4"
    details.print_area = f"A1:I{detail_last_row}"

    scoring_rows = [
        ("Parameter", "Value"),
        ("CPU warning", CPU_WARNING_PCT / 100),
        ("CPU critical", CPU_CRITICAL_PCT / 100),
        ("Minimum data coverage", MIN_DATA_COVERAGE_PCT / 100),
        ("Warning penalty", WARNING_PENALTY),
        ("Critical penalty", CRITICAL_PENALTY),
        ("Environment alert penalty", ENVIRONMENT_ALERT_PENALTY),
        ("Err-disabled penalty", ERR_DISABLED_PENALTY),
        ("Degraded EtherChannel penalty", DEGRADED_ETHERCHANNEL_PENALTY),
        ("Down EtherChannel penalty", DOWN_ETHERCHANNEL_PENALTY),
        ("Maximum count penalty", MAX_COUNT_PENALTY),
        ("Maximum EtherChannel penalty", MAX_ETHERCHANNEL_PENALTY),
        ("Healthy score", HEALTHY_SCORE),
        ("Watch score", WATCH_SCORE),
        ("Interface quality penalty", INTERFACE_QUALITY_PENALTY),
        ("Interface instability penalty", INTERFACE_INSTABILITY_PENALTY),
        ("Maximum interface penalty", MAX_INTERFACE_PENALTY),
        ("Critical EtherChannel down ratio", ETHERCHANNEL_CRITICAL_DOWN_RATIO),
    ]
    for row in scoring_rows:
        scoring.append(row)
    for row in range(2, 5):
        scoring.cell(row, 2).number_format = "0%"
    scoring.cell(18, 2).number_format = "0%"
    scoring.sheet_state = "veryHidden"

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
