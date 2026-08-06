'''This script currently goes through netbox, filter out devices based on unique "tags/tag" objects and then back up the running-config of those devices to a local dir'''

'''
Features in work:
 1.) Save to Netbox
 2.) Pipeline to save to a GitRepo
 3.) Filter based on multiple set parameters
 4.) Back up based on Hostname -> Date -> Config/Interface stats?/Health-Status
'''

import os
import re
from collections.abc import Callable
from datetime import datetime
from pathlib import Path
from typing import Any

from nornir import InitNornir
from nornir.core.filter import F
from nornir.core.task import Result, Task
from nornir_netmiko.tasks import netmiko_send_command
from nornir_utils.plugins.functions import print_result
from nornir.core.inventory import ConnectionOptions
try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    raise SystemExit(
        "ERROR: This script requires openpyxl. Install it with: pip install openpyxl"
    ) from exc
try:
    from rich.console import Console
    from rich.markup import escape
    from rich.progress import (
        BarColumn,
        Progress,
        SpinnerColumn,
        TaskID,
        TextColumn,
        TimeElapsedColumn,
    )
except ImportError as exc:
    raise SystemExit(
        "ERROR: This script requires Rich. Install it with: pip install rich"
    ) from exc



TARGET_TAG = "nornirtest"  #Tag used on Objects within Netbox.
BACKUP_ROOT = Path("./config_backups")  #Dir path for configuration backups
PROGRESS_STEPS = 10
console = Console()

INTERFACE_COMMANDS = {
    "ip_brief": "show ip interface brief",
    "stats": "show interfaces stats",
    "interfaces": "show interfaces",
    "switchport": "show interfaces switchport",
    "trunk": "show interfaces trunk",
}


def main() -> int:

    username = os.getenv("NORNIR_USERNAME") #exports your #USERNAME for logging into Network devices
    password = os.getenv("NORNIR_PASSWORD") #exports your #PASSWORD for logging into Network devices

   #Checks to confirm if VARs are present/set
    if not username or not password: 
        console.print(
            "ERROR: NORNIR_USERNAME and NORNIR_PASSWORD "
            "must be set in the environment."
        )
        return 1

    #Intializes Nornir
    try:
        nr = InitNornir(config_file="config.yaml") 
    except Exception as exc:
        console.print(
            f"[bold red]ERROR:[/] Could not initialize Nornir/NetBox inventory: {exc}"
        )
        return 1
    
    nr.inventory.defaults.username = username #Set's username as default vaule for logging across all devices
    nr.inventory.defaults.password = password #Set's password as default value for logigng across all devices

    #Check's to confirm it can reach Netbox's Inventory
    if not nr.inventory.hosts: 
        console.print(
            "No devices were loaded. Check the NetBox inventory plugin, "
            "API URL, token, permissions, and inventory configuration."
        )
        return 1

    # Inspects the tags then proceeds to pass them to the "normalize_tags" task/function.
    for host in nr.inventory.hosts.values():
        raw_tags = host.data.get("tags", [])
        normalized_tags = normalize_tags(raw_tags)
        host.data["tag_slugs"] = normalized_tags

        #Connection options - Configured currently for connecting to legacy SSH/slower devices.
        #Applied to all the objects that are pulled above from nr.inventory.hosts
        host.connection_options["netmiko"] = ConnectionOptions(extras= {"conn_timeout": 30,
                                                                        "banner_timeout": 60,
                                                                        "auth_timeout": 60,
                                                                        "fast_cli": False})

    #Performs a filter against Netbox's Inventory; using the TARGET_TAG to filter devices based on 'tags'
    targets = nr.filter(
        F(tag_slugs__contains=TARGET_TAG.casefold())
    )

    console.print("\n[bold]--- Filter results ---[/]")
    console.print(f"Target tag: {TARGET_TAG!r}")
    console.print(f"Matched devices: {len(targets.inventory.hosts)}")
    console.print("Devices selected:")
    for device_number, host_name in enumerate(targets.inventory.hosts, start=1):
        console.print(f"  {device_number:>3}. {escape(host_name)}")

    #Provides error output for if no devices are assigned to the specified parameter(Tag)
    if not targets.inventory.hosts:
        console.print(
            "\nNo devices matched the tag. Review the raw_tags and "
            "normalized_tags shown above."
        )
        return 0

    # Build the dated backup directory once for the entire backup run.
    # Result: /networkbackups/<year>/<month>/<day>/
    backup_date = datetime.now()

    dated_output_dir = (
        BACKUP_ROOT
        / backup_date.strftime("%Y")
        / backup_date.strftime("%m")
        / backup_date.strftime("%d")
    )

    dated_output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    console.print(f"\nOutput directory: {dated_output_dir.resolve()}")
    console.print("\n[bold]--- Starting configuration backups ---[/]")

    # Collect the running configuration and environment output
    # from each device filtered from NetBox.
    progress = Progress(
        SpinnerColumn(),
        TextColumn("[bold blue]{task.fields[host]}", justify="right"),
        BarColumn(bar_width=24),
        TextColumn("{task.completed:.0f}/{task.total:.0f}"),
        TextColumn("{task.fields[status]}", markup=True),
        TimeElapsedColumn(),
        console=console,
        transient=True,
        expand=True,
    )

    with progress:
        progress_tasks = {
            host_name: progress.add_task(
                "backup",
                total=PROGRESS_STEPS,
                host=host_name,
                status="[dim]Queued[/]",
                visible=False,
            )
            for host_name in targets.inventory.hosts
        }

        results = targets.run(
            name="Back up Cisco running configurations and environment",
            task=save_device_outputs,
            output_dir=dated_output_dir,
            progress_update=make_progress_updater(progress, progress_tasks),
        )

    # This is essential while troubleshooting.
    print_result(results)

    console.print("\n[bold]--- Backup summary ---[/]")

    for host_name in targets.inventory.hosts:
        if host_name in results.failed_hosts:
          console.print(f"[bold red]FAILED[/] {host_name}")
        else:
            config_file = (
                dated_output_dir
                / host_name
                / f"{host_name}.cfg")

            environment_file = (
                dated_output_dir
                / host_name
                / f"{host_name}_environment.txt")

            interface_file = (
                dated_output_dir
                / host_name
                / f"{host_name}_interfaces.xlsx")

            console.print(f"[bold green]SAVED[/] {host_name}")
            console.print(f"        Config:      {config_file.resolve()}")
            console.print(f"        Environment: {environment_file.resolve()}")
            console.print(f"        Interfaces:  {interface_file.resolve()}")

    failed_count = len(results.failed_hosts)
    successful_count = len(targets.inventory.hosts) - failed_count
    console.print(
        f"\nCompleted: [green]{successful_count} successful[/], "
        f"[red]{failed_count} failed[/]"
    )
    return 1 if failed_count else 0


def make_progress_updater(
    progress: Progress,
    progress_tasks: dict[str, TaskID],
) -> Callable[[str, int, str], None]:
    """Create a scroll-safe per-host progress callback for Nornir workers."""

    finished_hosts: set[str] = set()

    def update(host_name: str, completed: int, status: str) -> None:
        task_id = progress_tasks[host_name]
        progress.update(
            task_id,
            completed=completed,
            status=status,
            visible=True,
            refresh=True,
        )

        if completed >= PROGRESS_STEPS and host_name not in finished_hosts:
            finished_hosts.add(host_name)
            failed = "failed" in status.casefold()
            label = "[bold red]FAILED[/]" if failed else "[bold green]COMPLETE[/]"
            progress.console.print(f"{label} {escape(host_name)}")
            progress.update(task_id, visible=False, refresh=True)

    return update


def normalize_tags(tags: list[Any]) -> list[str]:
    
    #Converts NetBox tags into lowercase names/slugs.
    #Can handle:
    #  - strings
    #  - dictionaries
    #  - NetBox/tag objects
    
    normalized: list[str] = []

    for tag in tags:
        if isinstance(tag, str):
            values = [tag]

        elif isinstance(tag, dict):
            values = [
                tag.get("slug"),
                tag.get("name"),
            ]

        else:
            values = [
                getattr(tag, "slug", None),
                getattr(tag, "name", None),
            ]

        for value in values:
            if value:
                normalized.append(str(value).casefold())

    return list(set(normalized))


def canonical_interface_name(name: str) -> str:
    """Normalize common Cisco long and short interface names for joining."""

    compact = name.strip().replace(" ", "")
    replacements = {
        "HundredGigabitEthernet": "Hu",
        "FortyGigabitEthernet": "Fo",
        "TwentyFiveGigE": "Twe",
        "TenGigabitEthernet": "Te",
        "TwoGigabitEthernet": "Tw",
        "GigabitEthernet": "Gi",
        "FastEthernet": "Fa",
        "Ethernet": "Eth",
        "Port-channel": "Po",
        "PortChannel": "Po",
        "Loopback": "Lo",
        "Vlan": "Vl",
    }

    for long_name, short_name in replacements.items():
        if compact.casefold().startswith(long_name.casefold()):
            return short_name + compact[len(long_name):]

    return compact


def new_interface_row(interface: str) -> dict[str, Any]:
    """Return an empty report row with a stable schema."""

    return {
        "Interface": canonical_interface_name(interface),
        "IP Address": "",
        "Admin State": "",
        "Protocol State": "",
        "Description": "",
        "Duplex": "",
        "Speed": "",
        "Input Rate (bps)": None,
        "Input Rate (pps)": None,
        "Output Rate (bps)": None,
        "Output Rate (pps)": None,
        "RX Packets": None,
        "RX Bytes": None,
        "TX Packets": None,
        "TX Bytes": None,
        "Last Received": "",
        "Last Transmitted": "",
        "Switchport": "",
        "Administrative Mode": "",
        "Operational Mode": "",
        "Access VLAN": "",
        "Native VLAN": "",
        "Allowed VLANs": "",
        "Active VLANs": "",
        "Forwarding VLANs": "",
    }


def get_interface_row(
    rows: dict[str, dict[str, Any]],
    interface: str,
) -> dict[str, Any]:
    key = canonical_interface_name(interface)
    return rows.setdefault(key, new_interface_row(key))


def parse_ip_interface_brief(
    output: str,
    rows: dict[str, dict[str, Any]],
) -> None:
    pattern = re.compile(
        r"^\s*(?P<interface>\S+)\s+(?P<ip>\S+)\s+"
        r"\S+\s+\S+\s+(?P<status>.+?)\s+(?P<protocol>up|down)\s*$",
        re.IGNORECASE,
    )

    for line in output.splitlines():
        match = pattern.match(line)
        if not match or match.group("interface").casefold() == "interface":
            continue
        row = get_interface_row(rows, match.group("interface"))
        row["IP Address"] = match.group("ip")
        row["Admin State"] = match.group("status")
        row["Protocol State"] = match.group("protocol")


def parse_show_interfaces(
    output: str,
    rows: dict[str, dict[str, Any]],
) -> None:
    header_pattern = re.compile(
        r"^(?P<interface>\S+) is (?P<state>.+?), line protocol is "
        r"(?P<protocol>\S+)",
        re.MULTILINE,
    )
    matches = list(header_pattern.finditer(output))

    for index, match in enumerate(matches):
        block_end = matches[index + 1].start() if index + 1 < len(matches) else len(output)
        block = output[match.start():block_end]
        row = get_interface_row(rows, match.group("interface"))
        row["Admin State"] = match.group("state")
        row["Protocol State"] = match.group("protocol").rstrip(",")

        field_patterns = {
            "Description": r"^\s*Description:\s*(.+)$",
            "Last Received": r"Last input\s+([^,]+)",
            "Last Transmitted": r"Last input\s+[^,]+,\s*output\s+([^,]+)",
            "Input Rate (bps)": r"5 minute input rate\s+(\d+)\s+bits/sec",
            "Input Rate (pps)": r"5 minute input rate\s+\d+\s+bits/sec,\s+(\d+)\s+packets/sec",
            "Output Rate (bps)": r"5 minute output rate\s+(\d+)\s+bits/sec",
            "Output Rate (pps)": r"5 minute output rate\s+\d+\s+bits/sec,\s+(\d+)\s+packets/sec",
            "RX Packets": r"^\s*(\d+) packets input,",
            "RX Bytes": r"^\s*\d+ packets input,\s*(\d+) bytes",
            "TX Packets": r"^\s*(\d+) packets output,",
            "TX Bytes": r"^\s*\d+ packets output,\s*(\d+) bytes",
        }

        for field_name, pattern in field_patterns.items():
            value_match = re.search(pattern, block, re.MULTILINE | re.IGNORECASE)
            if value_match:
                value = value_match.group(1).strip()
                row[field_name] = int(value) if value.isdigit() else value

        link_match = re.search(
            r"\b(?P<duplex>(?:Full|Half|Auto)-duplex),\s*"
            r"(?P<speed>[^,\n]+)",
            block,
            re.IGNORECASE,
        )
        if link_match:
            row["Duplex"] = link_match.group("duplex")
            row["Speed"] = link_match.group("speed").strip()


def parse_show_interface_stats(
    output: str,
    rows: dict[str, dict[str, Any]],
) -> None:
    current_interface = ""
    for line in output.splitlines():
        interface_match = re.match(r"^\s*Interface\s+(\S+)\s*$", line, re.IGNORECASE)
        if interface_match:
            current_interface = interface_match.group(1)
            continue

        total_match = re.match(
            r"^\s*Total\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s*$",
            line,
            re.IGNORECASE,
        )
        if current_interface and total_match:
            row = get_interface_row(rows, current_interface)
            row["RX Packets"] = int(total_match.group(1))
            row["RX Bytes"] = int(total_match.group(2))
            row["TX Packets"] = int(total_match.group(3))
            row["TX Bytes"] = int(total_match.group(4))


def vlan_number(value: str) -> str:
    match = re.match(r"\s*(\d+|none|unassigned)", value, re.IGNORECASE)
    return match.group(1) if match else value.strip()


def parse_show_interfaces_switchport(
    output: str,
    rows: dict[str, dict[str, Any]],
) -> None:
    blocks = re.split(r"(?=^Name:\s*)", output, flags=re.MULTILINE)
    fields = {
        "Switchport": "Switchport",
        "Administrative Mode": "Administrative Mode",
        "Operational Mode": "Operational Mode",
        "Access Mode VLAN": "Access VLAN",
        "Trunking Native Mode VLAN": "Native VLAN",
        "Trunking VLANs Enabled": "Allowed VLANs",
    }

    for block in blocks:
        name_match = re.search(r"^Name:\s*(\S+)", block, re.MULTILINE)
        if not name_match:
            continue
        row = get_interface_row(rows, name_match.group(1))

        for label, field_name in fields.items():
            match = re.search(
                rf"^{re.escape(label)}:\s*(.+)$",
                block,
                re.MULTILINE | re.IGNORECASE,
            )
            if match:
                value = match.group(1).strip()
                row[field_name] = (
                    vlan_number(value)
                    if field_name in {"Access VLAN", "Native VLAN"}
                    else value
                )


def parse_show_interfaces_trunk(
    output: str,
    rows: dict[str, dict[str, Any]],
) -> None:
    section = ""
    section_headers = {
        "vlans allowed on trunk": "Allowed VLANs",
        "vlans allowed and active in management domain": "Active VLANs",
        "vlans in spanning tree forwarding state and not pruned": "Forwarding VLANs",
    }

    for line in output.splitlines():
        lowered = line.strip().casefold()
        if "native vlan" in lowered and lowered.startswith("port"):
            section = "trunk"
            continue

        changed_section = False
        for heading, field_name in section_headers.items():
            if heading in lowered:
                section = field_name
                changed_section = True
                break
        if changed_section or not lowered or lowered.startswith("port"):
            continue

        if section == "trunk":
            match = re.match(
                r"^\s*(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s*$",
                line,
            )
            if match:
                row = get_interface_row(rows, match.group(1))
                if not row["Administrative Mode"]:
                    row["Administrative Mode"] = match.group(2)
                row["Operational Mode"] = match.group(4)
                row["Native VLAN"] = match.group(5)
        elif section in section_headers.values():
            match = re.match(r"^\s*(\S+)\s+(.+?)\s*$", line)
            if match:
                get_interface_row(rows, match.group(1))[section] = match.group(2)


def interface_sort_key(name: str) -> tuple[str, tuple[int, ...], str]:
    prefix_match = re.match(r"([A-Za-z-]+)(.*)", name)
    prefix = prefix_match.group(1) if prefix_match else name
    numbers = tuple(int(value) for value in re.findall(r"\d+", name))
    return prefix.casefold(), numbers, name.casefold()


def build_interface_rows(outputs: dict[str, str]) -> list[dict[str, Any]]:
    """Parse and join all collected command output by interface name."""

    rows: dict[str, dict[str, Any]] = {}
    parse_ip_interface_brief(outputs.get("ip_brief", ""), rows)
    parse_show_interfaces(outputs.get("interfaces", ""), rows)
    parse_show_interface_stats(outputs.get("stats", ""), rows)
    parse_show_interfaces_switchport(outputs.get("switchport", ""), rows)
    parse_show_interfaces_trunk(outputs.get("trunk", ""), rows)
    return [rows[name] for name in sorted(rows, key=interface_sort_key)]


def create_interface_workbook(
    hostname: str,
    rows: list[dict[str, Any]],
    command_errors: dict[str, str],
    output_path: Path,
) -> None:
    """Create a formatted Excel interface inventory for one device."""

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    interfaces = workbook.create_sheet("Interfaces")
    headers = list(new_interface_row("").keys())
    data_start_row = 5
    data_end_row = max(data_start_row, data_start_row + len(rows) - 1)

    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:D1")
    summary["A1"] = f"{hostname} Interface Backup"
    summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="17365D")
    summary["A1"].alignment = Alignment(horizontal="left")
    summary["A3"] = "Backup time"
    summary["B3"] = datetime.now()
    summary["B3"].number_format = "yyyy-mm-dd hh:mm:ss"
    summary["A5"] = "Metric"
    summary["B5"] = "Count"
    summary["A6"] = "Total interfaces"
    summary["B6"] = f"=COUNTA('Interfaces'!A{data_start_row}:A{data_end_row})"
    summary["A7"] = "Operationally up"
    summary["B7"] = (
        f'=COUNTIF(\'Interfaces\'!D{data_start_row}:D{data_end_row},"up")'
    )
    summary["A8"] = "Operationally down"
    summary["B8"] = (
        f'=COUNTIF(\'Interfaces\'!D{data_start_row}:D{data_end_row},"down")'
    )
    summary["A10"] = "Collection warnings"
    summary["A10"].font = Font(bold=True)
    if command_errors:
        for row_number, (command, error) in enumerate(command_errors.items(), start=11):
            summary.cell(row_number, 1, command)
            summary.cell(row_number, 2, error)
    else:
        summary["A11"] = "None"

    for cell in summary[5]:
        if cell.column <= 2:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F75B5")
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 42
    summary.freeze_panes = "A5"

    interfaces.sheet_view.showGridLines = False
    interfaces.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    interfaces.cell(1, 1, f"{hostname} Interface Details")
    interfaces.cell(1, 1).font = Font(size=18, bold=True, color="FFFFFF")
    interfaces.cell(1, 1).fill = PatternFill("solid", fgColor="17365D")
    interfaces.cell(2, 1, "Collected")
    interfaces.cell(2, 2, datetime.now())
    interfaces.cell(2, 2).number_format = "yyyy-mm-dd hh:mm:ss"

    for column, header in enumerate(headers, start=1):
        cell = interfaces.cell(4, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F75B5")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_number, row in enumerate(rows, start=data_start_row):
        for column, header in enumerate(headers, start=1):
            value = row[header]
            interfaces.cell(
                row_number,
                column,
                "N/A" if value is None or value == "" else value,
            )

    if rows:
        table = Table(
            displayName="InterfaceInventory",
            ref=f"A4:Y{data_end_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        interfaces.add_table(table)

        interfaces.conditional_formatting.add(
            f"D{data_start_row}:D{data_end_row}",
            FormulaRule(
                formula=[f'LOWER(D{data_start_row})="up"'],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )
        interfaces.conditional_formatting.add(
            f"D{data_start_row}:D{data_end_row}",
            FormulaRule(
                formula=[f'LOWER(D{data_start_row})="down"'],
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )

    interfaces.freeze_panes = "A5"
    interfaces.auto_filter.ref = f"A4:Y{data_end_row}"
    widths = {
        "A": 20, "B": 16, "C": 22, "D": 16, "E": 32,
        "F": 14, "G": 14, "H": 18, "I": 18, "J": 18,
        "K": 18, "L": 16, "M": 16, "N": 16, "O": 16,
        "P": 18, "Q": 18, "R": 14, "S": 22, "T": 20,
        "U": 14, "V": 14, "W": 24, "X": 24, "Y": 28,
    }
    for column, width in widths.items():
        interfaces.column_dimensions[column].width = width

    for row in interfaces.iter_rows(min_row=data_start_row, max_row=data_end_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top")
    for column in range(8, 16):
        for row_number in range(data_start_row, data_end_row + 1):
            interfaces.cell(row_number, column).number_format = "#,##0"

    workbook.save(output_path)

def save_device_outputs(
    task: Task,
    output_dir: Path,
    progress_update: Callable[[str, int, str], None],
) -> Result:
    """
    Retrieve and save the running configuration and environment
    information for a single network device.
    """

    host_name = task.host.name

    def set_progress(completed: int, status: str) -> None:
        progress_update(host_name, completed, status)

    def failure(message: str, exception: Exception | None = None) -> Result:
        # Print a durable error above the live display. Keeping the progress-row
        # message short prevents long exceptions from wrapping over other hosts.
        console.print(
            f"[bold red]ERROR[/] [bold]{escape(host_name)}[/]: {escape(message)}"
        )
        set_progress(PROGRESS_STEPS, "[bold red]Failed — see error above[/]")
        return Result(
            host=task.host,
            failed=True,
            exception=exception,
            result=message,
        )

    set_progress(0, "[cyan]Connecting[/]")

    # Retrieve the running configuration.
    set_progress(1, "[cyan]Collecting configuration[/]")
    try:
        running_config_results = task.run(
            name="Get running configuration",
            task=netmiko_send_command,
            command_string="show running-config",
            read_timeout=120,
        )
    except Exception as exc:
        return failure(f"Connection/configuration error: {exc}", exc)

    running_config_result = running_config_results[-1]

    if running_config_result.failed:
        return failure(
            (
                "Failed to retrieve running configuration: "
                f"{running_config_result.exception or running_config_result.result}"
            )
        )

    running_config = str(running_config_result.result)

    if not running_config.strip():
        return failure("The device returned an empty running configuration.")

    set_progress(2, "[cyan]Collecting environment[/]")

    # Retrieve the environment information.
    try:
        environment_results = task.run(
            name="Get environment information",
            task=netmiko_send_command,
            command_string="show environment all",
            read_timeout=120,
        )
    except Exception as exc:
        return failure(f"Environment collection error: {exc}", exc)

    environment_result = environment_results[-1]

    if environment_result.failed:
        return failure(
            (
                "Failed to retrieve environment information: "
                f"{environment_result.exception or environment_result.result}"
            )
        )

    environment_output = str(environment_result.result)

    if not environment_output.strip():
        return failure("The device returned empty environment information.")

    interface_outputs: dict[str, str] = {}
    command_errors: dict[str, str] = {}

    for step, (output_name, command) in enumerate(
        INTERFACE_COMMANDS.items(),
        start=3,
    ):
        set_progress(step, f"[cyan]{command}[/]")

        try:
            command_results = task.run(
                name=f"Collect {command}",
                task=netmiko_send_command,
                command_string=command,
                read_timeout=180,
            )
            command_result = command_results[-1]
        except Exception as exc:
            return failure(f"Failed while running '{command}': {exc}", exc)

        if command_result.failed:
            return failure(
                f"Failed to run '{command}': "
                f"{command_result.exception or command_result.result}"
            )

        command_output = str(command_result.result)
        if not command_output.strip():
            command_errors[command] = "Device returned no output."
        else:
            interface_outputs[output_name] = command_output

    set_progress(8, "[cyan]Building interface report[/]")
    interface_rows = build_interface_rows(interface_outputs)

    # Create the hostname-specific backup directory.
    host_backup_dir = output_dir / task.host.name

    config_filename = (
        host_backup_dir
        / f"{task.host.name}.cfg"
    )

    environment_filename = (
        host_backup_dir
        / f"{task.host.name}_environment.txt"
    )

    interface_filename = (
        host_backup_dir
        / f"{task.host.name}_interfaces.xlsx"
    )

    try:
        set_progress(9, "[cyan]Saving files[/]")
        host_backup_dir.mkdir(
            parents=True,
            exist_ok=True,
        )

        config_filename.write_text(
            running_config.rstrip() + "\n",
            encoding="utf-8",
        )

        environment_filename.write_text(
            environment_output.rstrip() + "\n",
            encoding="utf-8",
        )

        create_interface_workbook(
            hostname=task.host.name,
            rows=interface_rows,
            command_errors=command_errors,
            output_path=interface_filename,
        )

    except OSError as exc:
        return failure(
            (
                f"Could not write backup files for "
                f"{task.host.name}: {exc}"
            ),
            exc,
        )

    set_progress(PROGRESS_STEPS, "[bold green]Complete[/]")

    return Result(
        host=task.host,
        changed=False,
        result=(
            f"Saved running configuration to "
            f"{config_filename.resolve()}\n"
            f"Saved environment information to "
            f"{environment_filename.resolve()}\n"
            f"Saved interface report to "
            f"{interface_filename.resolve()}"
        ),
    )

if __name__ == "__main__":
    raise SystemExit(main())
