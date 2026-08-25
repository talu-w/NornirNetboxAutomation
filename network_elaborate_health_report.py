"""Create an engineer-focused network-health workbook from Nornir and NetBox.

The script uses the repository's existing Nornir ``config.yaml``. NetBox devices
are loaded by that configuration, normalized, and filtered with Nornir's ``F``
filter. It collects device, interface quality, spanning-tree, deep
EtherChannel, security/control-plane, CPU, and environment state. No device
configuration is changed.
"""

from __future__ import annotations

import argparse
import os
import re
from collections.abc import Mapping, Sequence
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from nornir import InitNornir
from nornir.core.filter import F
from nornir.core.inventory import ConnectionOptions, Host
from nornir.core.task import Result, Task
from nornir_netmiko.tasks import netmiko_send_command

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.pagebreak import Break
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    raise SystemExit(
        "ERROR: This script requires openpyxl. Install it with: pip install openpyxl"
    ) from exc

try:
    from rich.console import Console
except ImportError as exc:
    raise SystemExit(
        "ERROR: This script requires Rich. Install it with: pip install rich"
    ) from exc


DEFAULT_CONFIG_FILE = "config.yaml"
DEFAULT_TARGET_TAG = "nornirtest"
DEFAULT_OUTPUT_FILE = (
    "Network_Elaborate_Health_Report_"
    f"{datetime.now().astimezone().strftime('%Y-%m-%d')}.xlsx"
)
DEFAULT_CONNECTION_TIMEOUT = 60.0
DEFAULT_AUTH_TIMEOUT = 120.0
DEFAULT_BANNER_TIMEOUT = 120.0
DEFAULT_READ_TIMEOUT = 180.0
DEFAULT_GLOBAL_DELAY_FACTOR = 2.0

# Internal scoring values. They are stored in hidden worksheet rows so the
# front-facing report stays clean while its formulas remain consistent.
CPU_WARNING_PCT = 75.0
CPU_CRITICAL_PCT = 90.0
MIN_DATA_COVERAGE_PCT = 60.0
WARNING_PENALTY = 10
CRITICAL_PENALTY = 25
ENVIRONMENT_ALERT_PENALTY = 10
ERR_DISABLED_PENALTY = 10
PORT_SECURITY_PENALTY = 10
DEGRADED_ETHERCHANNEL_PENALTY = 10
DOWN_ETHERCHANNEL_PENALTY = 25
MAX_COUNT_PENALTY = 30
MAX_ETHERCHANNEL_PENALTY = 50
INTERFACE_QUALITY_PENALTY = 3
INTERFACE_INSTABILITY_PENALTY = 5
STP_RECENT_CHANGE_PENALTY = 5
STP_INCONSISTENT_PORT_PENALTY = 15
SECURITY_EXCEPTION_PENALTY = 5
MAX_ENGINEERING_EXCEPTION_PENALTY = 30
HEALTHY_SCORE = 85
WATCH_SCORE = 70
HEALTH_COMPONENT_COUNT = 9
INTERFACE_ERROR_CRITICAL_COUNT = 10_000
RECENT_EVENT_SECONDS = 3_600

PHYSICAL_INTERFACE_RE = re.compile(
    r"^(?:Fa|FastEthernet|Gi|GigabitEthernet|Te|TenGigabitEthernet|"
    r"Tw|TwoGigabitEthernet|Twe|TwentyFiveGigE|Fo|FortyGigabitEthernet|"
    r"Hu|HundredGig(?:E|abitEthernet)|Eth|Ethernet|Et)\d",
    re.IGNORECASE,
)

INTERFACE_STATUS_RE = re.compile(
    r"\b(err-disabled|errdisabled|notconnect(?:ed)?|notconnec|connected|disabled|inactive|"
    r"monitoring|sfpAbsent|xcvrAbsent)\b",
    re.IGNORECASE,
)

ETHERCHANNEL_ROW_RE = re.compile(
    r"^\s*(?P<group>\d+)\s+"
    r"(?P<name>(?:Po|Port-?channel)\s*\d+)"
    r"(?:\((?P<flags>[^)]*)\))?\s*(?P<rest>.*)$",
    re.IGNORECASE,
)

ETHERCHANNEL_MEMBER_RE = re.compile(
    r"(?P<name>[A-Za-z][A-Za-z-]*\d[\w./:-]*)"
    r"\((?P<flags>[^)]+)\)",
)

NETWORK_INTERFACE_RE = re.compile(
    r"^(?:(?:Fa|FastEthernet|Gi|GigabitEthernet|Te|TenGigabitEthernet|"
    r"Tw|TwoGigabitEthernet|Twe|TwentyFiveGigE|Fo|FortyGigabitEthernet|"
    r"Hu|HundredGig(?:E|abitEthernet)|Eth|Ethernet|Et|Po|Port-?channel)\d)",
    re.IGNORECASE,
)

INTERFACE_COUNTER_ALIASES = {
    "alignerr": "alignment_errors",
    "fcserr": "crc_errors",
    "crc": "crc_errors",
    "xmterr": "output_errors",
    "rcverr": "input_errors",
    "inerrors": "input_errors",
    "outerrors": "output_errors",
    "undersize": "undersize",
    "outdiscards": "output_drops",
    "indiscards": "input_drops",
    "singlecol": "single_collisions",
    "multicol": "multiple_collisions",
    "latecol": "late_collisions",
    "excesscol": "excessive_collisions",
    "collisions": "collisions",
    "carrisen": "carrier_sense_errors",
    "runts": "runts",
    "runt": "runts",
    "giants": "giants",
    "frame": "frame_errors",
    "overrun": "overruns",
    "ignored": "ignored",
}

ETHERCHANNEL_MEMBER_FLAG_MEANINGS = {
    "P": "bundled",
    "H": "hot standby",
    "s": "suspended",
    "I": "stand-alone",
    "D": "down",
    "w": "waiting",
    "f": "failed aggregator allocation",
    "M": "minimum links not met",
    "m": "minimum links not met",
}

ERRDISABLE_REASON_RE = re.compile(
    r"\b(psecure-violation|security-violation|bpduguard|bpdu-guard|"
    r"link-flap|channel-misconfig|udld|storm-control|dhcp-rate-limit|"
    r"arp-inspection|loopback|l2ptguard|sfp-config-mismatch)\b",
    re.IGNORECASE,
)

INVALID_COMMAND_MARKERS = (
    "% invalid input",
    "% incomplete command",
    "% ambiguous command",
    "invalid command",
    "unknown command",
)

console = Console()


@dataclass(slots=True)
class InterfaceSummary:
    connected: int | None = None
    not_connected: int | None = None
    total: int | None = None
    err_disabled: list[dict[str, str]] = field(default_factory=list)


@dataclass(slots=True)
class EtherChannelState:
    name: str
    group: int
    protocol: str
    flags: str
    state: str
    members: list[str] = field(default_factory=list)
    bundled_members: list[str] = field(default_factory=list)
    standby_members: list[str] = field(default_factory=list)
    problem_members: list[str] = field(default_factory=list)
    member_details: list[dict[str, str]] = field(default_factory=list)
    min_links: int | None = None
    risk_reasons: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(slots=True)
class PortSecurityIssue:
    interface: str
    violation_count: int
    action: str = ""

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(slots=True)
class InterfaceQualityIssue:
    interface: str
    counters: dict[str, int] = field(default_factory=dict)
    last_link_flapped: str = ""
    recent_link_flap: bool = False
    interface_resets: int = 0
    carrier_transitions: int = 0

    @property
    def error_total(self) -> int:
        return sum(self.counters.values())

    @property
    def unstable(self) -> bool:
        return bool(
            self.recent_link_flap
            or self.interface_resets > 0
            or self.carrier_transitions > 0
        )

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["error_total"] = self.error_total
        data["unstable"] = self.unstable
        return data


@dataclass(slots=True)
class SpanningTreeInstance:
    instance: str
    root_bridge: str = ""
    root_port: str = ""
    topology_changes: int = 0
    last_change: str = ""
    change_source: str = ""
    recent_change: bool = False

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(slots=True)
class SecurityControlException:
    category: str
    object_name: str
    count: int
    details: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(slots=True)
class DeviceHealth:
    hostname: str
    location: str = ""
    site: str = ""
    role: str = ""
    model: str = ""
    platform: str = ""
    netbox_status: str = ""
    primary_ip: str = ""
    serial_number: str = ""
    firmware: str = ""
    connected_interfaces: int | None = None
    not_connected_interfaces: int | None = None
    total_interfaces: int | None = None
    interface_quality_issues: list[dict[str, Any]] = field(default_factory=list)
    interfaces_with_errors: int | None = None
    interface_error_total: int | None = None
    interface_instability_count: int | None = None
    cpu_pct: float | None = None
    environment_alerts: int | None = None
    err_disabled_interfaces: list[dict[str, str]] = field(default_factory=list)
    err_disabled_count: int | None = None
    port_security_issues: list[dict[str, Any]] = field(default_factory=list)
    port_security_interfaces: int | None = None
    port_security_violations: int | None = None
    etherchannels: list[dict[str, Any]] = field(default_factory=list)
    etherchannels_total: int | None = None
    etherchannels_healthy: int | None = None
    etherchannels_degraded: int | None = None
    etherchannels_down: int | None = None
    etherchannel_members_total: int | None = None
    etherchannel_members_bundled: int | None = None
    etherchannel_members_standby: int | None = None
    etherchannel_member_issues: int | None = None
    etherchannel_min_link_risks: int | None = None
    spanning_tree_instances: list[dict[str, Any]] = field(default_factory=list)
    spanning_tree_instance_count: int | None = None
    stp_topology_changes: int | None = None
    stp_recent_changes: int | None = None
    stp_inconsistent_ports: list[dict[str, str]] = field(default_factory=list)
    stp_inconsistent_port_count: int | None = None
    security_control_exceptions: list[dict[str, Any]] = field(default_factory=list)
    security_control_exception_count: int | None = None
    security_control_drop_count: int | None = None
    reachable: bool = False
    collected_at_utc: datetime | None = None
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def positive_float(value: str) -> float:
    """Return a positive numeric command-line value."""

    try:
        parsed = float(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"{value!r} is not a number") from exc
    if parsed <= 0:
        raise argparse.ArgumentTypeError("value must be greater than zero")
    return parsed


def environment_flag(name: str) -> bool:
    """Interpret a common true/false environment-variable value."""

    value = os.getenv(name, "").strip().casefold()
    return value in {"1", "true", "yes", "on"}


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Collect engineer-focused health data from NetBox-tagged devices with "
            "Nornir and create an elaborate Excel report."
        )
    )
    parser.add_argument(
        "--config",
        default=os.getenv("NORNIR_CONFIG_FILE", DEFAULT_CONFIG_FILE),
        help="Existing Nornir config file (default: config.yaml)",
    )
    parser.add_argument(
        "--tag",
        default=os.getenv("NORNIR_TARGET_TAG", DEFAULT_TARGET_TAG),
        help="NetBox tag to select (default: nornirtest)",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT_FILE,
        help=f"Excel output path (default: {DEFAULT_OUTPUT_FILE})",
    )
    parser.add_argument(
        "--connection-timeout",
        "--conn-timeout",
        dest="connection_timeout",
        type=positive_float,
        default=os.getenv("NORNIR_CONNECTION_TIMEOUT", str(DEFAULT_CONNECTION_TIMEOUT)),
        help=(
            "Seconds allowed to establish SSH transport "
            f"(default: {DEFAULT_CONNECTION_TIMEOUT:g})"
        ),
    )
    parser.add_argument(
        "--auth-timeout",
        type=positive_float,
        default=os.getenv("NORNIR_AUTH_TIMEOUT", str(DEFAULT_AUTH_TIMEOUT)),
        help=(
            "Seconds allowed for SSH authentication "
            f"(default: {DEFAULT_AUTH_TIMEOUT:g})"
        ),
    )
    parser.add_argument(
        "--banner-timeout",
        type=positive_float,
        default=os.getenv("NORNIR_BANNER_TIMEOUT", str(DEFAULT_BANNER_TIMEOUT)),
        help=(
            "Seconds allowed for an SSH identification banner "
            f"(default: {DEFAULT_BANNER_TIMEOUT:g})"
        ),
    )
    parser.add_argument(
        "--read-timeout",
        type=positive_float,
        default=os.getenv("NORNIR_READ_TIMEOUT", str(DEFAULT_READ_TIMEOUT)),
        help=(
            "Seconds allowed for each show command to finish "
            f"(default: {DEFAULT_READ_TIMEOUT:g})"
        ),
    )
    parser.add_argument(
        "--global-delay-factor",
        type=positive_float,
        default=os.getenv(
            "NORNIR_GLOBAL_DELAY_FACTOR", str(DEFAULT_GLOBAL_DELAY_FACTOR)
        ),
        help=(
            "Netmiko delay multiplier for slow devices "
            f"(default: {DEFAULT_GLOBAL_DELAY_FACTOR:g})"
        ),
    )
    parser.add_argument(
        "--legacy-ssh",
        action="store_true",
        default=environment_flag("NORNIR_LEGACY_SSH"),
        help=(
            "Enable Netmiko compatibility for older SSH servers that do not "
            "advertise RSA-SHA2 support"
        ),
    )
    return parser.parse_args(argv)


def build_netmiko_extras(args: argparse.Namespace) -> dict[str, Any]:
    """Build adjustable Netmiko settings for slow or legacy devices."""

    extras: dict[str, Any] = {
        "conn_timeout": args.connection_timeout,
        "banner_timeout": args.banner_timeout,
        "auth_timeout": args.auth_timeout,
        "read_timeout_override": args.read_timeout,
        "global_delay_factor": args.global_delay_factor,
        "fast_cli": False,
    }
    if args.legacy_ssh:
        extras["disable_sha2_fix"] = True
    return extras


def normalize_tags(tags: Sequence[Any] | None) -> list[str]:
    """Convert NetBox tag strings, dictionaries, or objects to lowercase values."""

    normalized: set[str] = set()
    for tag in tags or []:
        if isinstance(tag, str):
            values = [tag]
        elif isinstance(tag, Mapping):
            values = [tag.get("slug"), tag.get("name")]
        else:
            values = [getattr(tag, "slug", None), getattr(tag, "name", None)]

        for value in values:
            if value:
                normalized.add(str(value).casefold())
    return sorted(normalized)


def _nested_value(data: Mapping[str, Any], *paths: str) -> Any:
    for path in paths:
        value: Any = data
        for part in path.split("."):
            if isinstance(value, Mapping):
                value = value.get(part)
            else:
                value = getattr(value, part, None)
            if value is None:
                break
        if value not in (None, ""):
            return value
    return None


def _display_value(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, Mapping):
        for key in ("display", "label", "name", "model", "address", "value", "slug"):
            if value.get(key) not in (None, ""):
                return str(value[key])
        return ""
    for attribute in ("display", "label", "name", "model", "address", "value", "slug"):
        candidate = getattr(value, attribute, None)
        if candidate not in (None, ""):
            return str(candidate)
    return str(value)


def netbox_metadata(host: Host) -> dict[str, str]:
    """Read useful management fields already supplied by the NetBox inventory."""

    data = host.data
    site = _display_value(_nested_value(data, "site", "site_name"))
    location = _display_value(_nested_value(data, "location", "location_name")) or site
    return {
        "location": location,
        "site": site,
        "role": _display_value(_nested_value(data, "role", "device_role")),
        "model": _display_value(_nested_value(data, "device_type.model", "device_type", "model")),
        "platform": _display_value(host.platform or _nested_value(data, "platform")),
        "netbox_status": _display_value(_nested_value(data, "status")),
        "primary_ip": _display_value(
            _nested_value(data, "primary_ip4.address", "primary_ip4", "primary_ip.address", "primary_ip")
        )
        or str(host.hostname or ""),
        "serial_number": _display_value(_nested_value(data, "serial", "serial_number")),
    }


def _is_invalid_command(output: str) -> bool:
    lowered = output.casefold()
    return any(marker in lowered for marker in INVALID_COMMAND_MARKERS)


def run_first_supported(
    task: Task,
    label: str,
    commands: Sequence[str],
    read_timeout: float = DEFAULT_READ_TIMEOUT,
) -> tuple[str, str]:
    """Run commands in order and return the first usable output and command."""

    errors: list[str] = []
    for command in commands:
        try:
            results = task.run(
                name=f"{label}: {command}",
                task=netmiko_send_command,
                command_string=command,
                read_timeout=read_timeout,
            )
            result = results[-1]
        except Exception as exc:  # noqa: BLE001 - plugins raise platform-specific errors.
            errors.append(f"{command}: {exc}")
            continue

        if result.failed:
            errors.append(f"{command}: {result.exception or result.result}")
            continue

        output = str(result.result or "")
        if output.strip() and not _is_invalid_command(output):
            return output, command
        errors.append(f"{command}: contained an empty output or was unable to collect at this time!")

    raise RuntimeError("; ".join(errors) or f"No {label} command succeeded")


def command_profile(platform: str) -> dict[str, list[str]]:
    platform_name = platform.casefold()
    profile = {
        "version": ["show version"],
        "interfaces": ["show interfaces status", "show ip interface brief"],
        "cpu": ["show process cpu"],
        "environment": ["show environment all", "show environment",],
        "interface_errors": [
            "show interfaces counters errors",
            "show interface counters errors",
        ],
        "interface_detail": ["show interfaces"],
        "etherchannels": ["show etherchannel summary", "show port-channel summary"],
        "etherchannel_detail": [
            "show etherchannel detail",
            "show port-channel database",
        ],
        "spanning_tree": ["show spanning-tree detail"],
        "stp_inconsistent": [
            "show spanning-tree inconsistentports",
            "show spanning-tree inconsistent-ports",
        ],
        "err_disabled": [
            "show interfaces status err-disabled",
            "show interface status err-disabled",
        ],
        "port_security": ["show port-security"],
        "control_plane": [
            "show policy-map control-plane",
            "show policy-map interface control-plane",
        ],
        "arp_inspection": ["show ip arp inspection statistics"],
        "dhcp_snooping": ["show ip dhcp snooping statistics"],
    }

    if "nxos" in platform_name or "nx-os" in platform_name:
        profile.update(
            {
                "interfaces": ["show interface status"],
                "cpu": ["show system resources"],
                "environment": ["show env all"],
                "interface_errors": [
                    "show interface counters errors",
                    "show interfaces counters errors",
                ],
                "interface_detail": ["show interface"],
                "etherchannels": [
                    "show port-channel summary",
                    "show etherchannel summary",
                ],
                "etherchannel_detail": [
                    "show port-channel database",
                    "show port-channel summary",
                ],
                "spanning_tree": ["show spanning-tree detail"],
                "stp_inconsistent": [
                    "show spanning-tree inconsistentports",
                ],
                "err_disabled": [
                    "show interface status err-disabled",
                    "show interfaces status err-disabled",
                ],
                "port_security": ["show port-security"],
                "control_plane": [
                    "show policy-map interface control-plane",
                    "show policy-map control-plane",
                ],
            }
        )
    elif "ios" in platform_name:
        profile.update(
            {
                "interfaces": ["show interface status"],
                "cpu": ["show process cpu"],
                "environment": ["show env all"],
                "interface_errors": [
                    "show interfaces counters errors",
                    "show interface counters errors",
                ],
                "interface_detail": ["show interfaces"],
                "etherchannels": [
                    "show etherchannel summary",
                    "show port-channel summary",
                ],
                "etherchannel_detail": [
                    "show etherchannel detail",
                    "show port-channel database",
                ],
                "spanning_tree": ["show spanning-tree detail"],
                "stp_inconsistent": [
                    "show spanning-tree inconsistentports",
                    "show spanning-tree inconsistent-ports",
                ],
                "err_disabled": [
                    "show interfaces status err-disabled",
                    "show interface status err-disabled",
                ],
                "port_security": ["show port-security"],
                "control_plane": [
                    "show policy-map control-plane",
                    "show policy-map interface control-plane",
                ],
            }
        )
    elif "eos" in platform_name or "arista" in platform_name:
        profile.update(
            {
                "interfaces": ["show interfaces status"],
                "cpu": ["show processes top once"],
                "environment": ["show system environment all"],
                "interface_errors": [
                    "show interfaces counters errors",
                    "show interfaces counters discards",
                ],
                "interface_detail": ["show interfaces"],
                "etherchannels": ["show port-channel summary"],
                "etherchannel_detail": [
                    "show lacp neighbor",
                    "show port-channel summary",
                ],
                "spanning_tree": ["show spanning-tree detail"],
                "stp_inconsistent": [
                    "show spanning-tree inconsistentports",
                ],
                "err_disabled": [
                    "show interfaces status errdisabled",
                    "show interfaces status err-disabled",
                ],
                "port_security": ["show port-security"],
                "control_plane": ["show policy-map control-plane"],
            }
        )
    return profile


def parse_firmware(output: str) -> str:
    patterns = (
        r"Cisco IOS XE Software, Version\s+([^,\s]+)",
        r"Cisco IOS Software.*?Version\s+([^,\s]+)",
        r"NXOS:\s+version\s+([^\s]+)",
        r"system:\s+version\s+([^\s]+)",
        r"Software image version:\s*([^\s]+)",
        r"Arista .*? version\s+([^\s]+)",
    )
    for pattern in patterns:
        match = re.search(pattern, output, re.IGNORECASE | re.DOTALL)
        if match:
            return match.group(1).strip()
    return "Unknown"


def parse_cpu_pct(output: str) -> float | None:
    match = re.search(
        r"CPU utilization for five seconds:\s*(\d+(?:\.\d+)?)%",
        output,
        re.IGNORECASE,
    )
    if match:
        return float(match.group(1))

    idle_match = re.search(
        r"(\d+(?:\.\d+)?)%\s*(?:idle|id)\b", output, re.IGNORECASE
    )
    if idle_match:
        return max(0.0, min(100.0, 100.0 - float(idle_match.group(1))))

    user_match = re.search(r"(\d+(?:\.\d+)?)%\s*user", output, re.IGNORECASE)
    kernel_match = re.search(r"(\d+(?:\.\d+)?)%\s*kernel", output, re.IGNORECASE)
    if user_match and kernel_match:
        return min(100.0, float(user_match.group(1)) + float(kernel_match.group(1)))
    return None


def _counter_value(value: str) -> int | None:
    cleaned = value.replace(",", "").strip()
    return int(cleaned) if cleaned.isdigit() else None


def _counter_name(value: str) -> str | None:
    normalized = re.sub(r"[^a-z0-9]", "", value.casefold())
    return INTERFACE_COUNTER_ALIASES.get(normalized)


def _interface_key(value: str) -> str:
    """Normalize common long/short interface prefixes for deduplication."""

    compact = value.strip().replace(" ", "").casefold()
    prefixes = (
        ("twentyfivegige", "twe"),
        ("hundredgigabitethernet", "hu"),
        ("hundredgige", "hu"),
        ("fortygigabitethernet", "fo"),
        ("tengigabitethernet", "te"),
        ("twogigabitethernet", "tw"),
        ("gigabitethernet", "gi"),
        ("fastethernet", "fa"),
        ("port-channel", "po"),
        ("portchannel", "po"),
        ("ethernet", "eth"),
    )
    for long_name, short_name in prefixes:
        if compact.startswith(long_name):
            return short_name + compact[len(long_name) :]
    return compact


def _duration_seconds(value: str) -> int | None:
    """Convert common IOS/NX-OS elapsed-time strings to seconds."""

    text = value.strip().casefold().rstrip(".,")
    if not text or "never" in text:
        return None

    clock_match = re.fullmatch(r"(?:(\d+):)?(\d+):(\d+)", text)
    if clock_match:
        hours = int(clock_match.group(1) or 0)
        minutes = int(clock_match.group(2))
        seconds = int(clock_match.group(3))
        return hours * 3_600 + minutes * 60 + seconds

    units = {
        "w": 604_800,
        "week": 604_800,
        "weeks": 604_800,
        "d": 86_400,
        "day": 86_400,
        "days": 86_400,
        "h": 3_600,
        "hour": 3_600,
        "hours": 3_600,
        "m": 60,
        "minute": 60,
        "minutes": 60,
        "s": 1,
        "second": 1,
        "seconds": 1,
    }
    matches = re.findall(
        r"(\d+)\s*(weeks?|days?|hours?|minutes?|seconds?|[wdhms])", text
    )
    if not matches:
        return None
    return sum(int(number) * units[unit] for number, unit in matches)


def parse_interface_quality(output: str) -> list[InterfaceQualityIssue]:
    """Parse cumulative interface errors plus reset/link-transition indicators."""

    issues: dict[str, InterfaceQualityIssue] = {}

    def get_issue(interface: str) -> InterfaceQualityIssue:
        key = _interface_key(interface)
        if key not in issues:
            issues[key] = InterfaceQualityIssue(interface=interface)
        return issues[key]

    header: list[str | None] | None = None
    for line in output.splitlines():
        tokens = line.split()
        if not tokens:
            continue
        if tokens[0].casefold() in {"port", "interface"}:
            candidate = [_counter_name(token) for token in tokens[1:]]
            header = candidate if any(candidate) else None
            continue
        if header and PHYSICAL_INTERFACE_RE.match(tokens[0]):
            issue = get_issue(tokens[0])
            for name, raw_value in zip(header, tokens[1:], strict=False):
                value = _counter_value(raw_value)
                if name and value is not None:
                    issue.counters[name] = max(issue.counters.get(name, 0), value)

    current: InterfaceQualityIssue | None = None
    block_pattern = re.compile(
        r"^\s*(?P<interface>\S+)\s+is\s+.+?,\s*line protocol is\s+",
        re.IGNORECASE,
    )
    metric_patterns = {
        "input_errors": r"([\d,]+)\s+input errors",
        "crc_errors": r"([\d,]+)\s+CRC",
        "frame_errors": r"([\d,]+)\s+frame",
        "overruns": r"([\d,]+)\s+overrun",
        "ignored": r"([\d,]+)\s+ignored",
        "runts": r"([\d,]+)\s+runts",
        "giants": r"([\d,]+)\s+giants",
        "output_errors": r"([\d,]+)\s+output errors",
        "collisions": r"([\d,]+)\s+collisions",
        "output_drops": r"Total output drops:\s*([\d,]+)",
    }
    for line in output.splitlines():
        block_match = block_pattern.match(line)
        if block_match:
            interface = block_match.group("interface")
            current = get_issue(interface) if PHYSICAL_INTERFACE_RE.match(interface) else None
            continue
        if current is None:
            continue
        for name, pattern in metric_patterns.items():
            match = re.search(pattern, line, re.IGNORECASE)
            if match:
                value = _counter_value(match.group(1))
                if value is not None:
                    current.counters[name] = max(current.counters.get(name, 0), value)

        reset_match = re.search(r"([\d,]+)\s+interface resets", line, re.IGNORECASE)
        if reset_match:
            current.interface_resets = _counter_value(reset_match.group(1)) or 0
        transition_match = re.search(
            r"([\d,]+)\s+carrier transitions", line, re.IGNORECASE
        )
        if transition_match:
            current.carrier_transitions = _counter_value(transition_match.group(1)) or 0
        flap_match = re.search(
            r"Last link flapped\s+([^\s,()]+)", line, re.IGNORECASE
        )
        if flap_match:
            current.last_link_flapped = flap_match.group(1)
            elapsed = _duration_seconds(current.last_link_flapped)
            current.recent_link_flap = (
                elapsed is not None and elapsed <= RECENT_EVENT_SECONDS
            )

    return sorted(
        (
            issue
            for issue in issues.values()
            if issue.error_total > 0 or issue.unstable
        ),
        key=lambda item: item.interface.casefold(),
    )


def merge_interface_quality_issues(
    *issue_groups: Sequence[InterfaceQualityIssue],
) -> list[InterfaceQualityIssue]:
    merged: dict[str, InterfaceQualityIssue] = {}
    for group in issue_groups:
        for issue in group:
            key = _interface_key(issue.interface)
            target = merged.setdefault(key, InterfaceQualityIssue(issue.interface))
            for name, value in issue.counters.items():
                target.counters[name] = max(target.counters.get(name, 0), value)
            target.interface_resets = max(
                target.interface_resets, issue.interface_resets
            )
            target.carrier_transitions = max(
                target.carrier_transitions, issue.carrier_transitions
            )
            if issue.last_link_flapped:
                target.last_link_flapped = issue.last_link_flapped
            target.recent_link_flap = target.recent_link_flap or issue.recent_link_flap
    return sorted(merged.values(), key=lambda item: item.interface.casefold())


def parse_spanning_tree_instances(output: str) -> list[SpanningTreeInstance]:
    """Parse roots and topology-change data from spanning-tree detail output."""

    instances: list[SpanningTreeInstance] = []
    current: SpanningTreeInstance | None = None
    instance_pattern = re.compile(
        r"^\s*(?P<instance>\S+)\s+is executing .*Spanning Tree protocol",
        re.IGNORECASE,
    )
    for line in output.splitlines():
        instance_match = instance_pattern.match(line)
        if instance_match:
            current = SpanningTreeInstance(instance=instance_match.group("instance"))
            instances.append(current)
            continue
        if current is None:
            continue

        root_match = re.search(
            r"Current root has priority\s+\d+,\s+address\s+([0-9a-f.:-]+)",
            line,
            re.IGNORECASE,
        )
        if root_match:
            current.root_bridge = root_match.group(1)
        root_port_match = re.search(r"Root port is .*?\(([^)]+)\)", line, re.IGNORECASE)
        if root_port_match:
            current.root_port = root_port_match.group(1)
        change_match = re.search(
            r"Number of topology changes\s+(\d+)\s+last change occurred\s+(.+?)\s+ago",
            line,
            re.IGNORECASE,
        )
        if change_match:
            current.topology_changes = int(change_match.group(1))
            current.last_change = change_match.group(2).strip()
            elapsed = _duration_seconds(current.last_change)
            current.recent_change = (
                elapsed is not None and elapsed <= RECENT_EVENT_SECONDS
            )
        source_match = re.match(r"^\s*from\s+(\S+)", line, re.IGNORECASE)
        if source_match and current.last_change:
            current.change_source = source_match.group(1)
    return instances


def parse_stp_inconsistent_ports(output: str) -> list[dict[str, str]]:
    """Return STP inconsistent interfaces and their reported reason/state."""

    issues: dict[tuple[str, str], dict[str, str]] = {}
    for line in output.splitlines():
        tokens = line.split()
        interface_index = next(
            (
                index
                for index, token in enumerate(tokens)
                if NETWORK_INTERFACE_RE.match(token)
            ),
            None,
        )
        if interface_index is None:
            continue
        interface = tokens[interface_index]
        instance = tokens[0] if interface_index > 0 else "Spanning Tree"
        reason = " ".join(tokens[interface_index + 1 :]) or "Inconsistent"
        issues[(instance.casefold(), interface.casefold())] = {
            "instance": instance,
            "interface": interface,
            "reason": reason,
        }
    return sorted(
        issues.values(),
        key=lambda item: (item["instance"].casefold(), item["interface"].casefold()),
    )


def parse_security_control_exceptions(
    outputs: Mapping[str, str],
) -> list[SecurityControlException]:
    """Parse nonzero CoPP, Dynamic ARP Inspection, and DHCP-snooping drops."""

    exceptions: dict[tuple[str, str, str], SecurityControlException] = {}

    def add(category: str, object_name: str, count: int, details: str) -> None:
        if count <= 0:
            return
        key = (category.casefold(), object_name.casefold(), details.casefold())
        exceptions[key] = SecurityControlException(
            category=category,
            object_name=object_name,
            count=count,
            details=details,
        )

    control_plane = outputs.get("control_plane", "")
    current_class = "Control plane"
    for line in control_plane.splitlines():
        class_match = re.search(r"Class-map:\s*([^\s(]+)", line, re.IGNORECASE)
        if class_match:
            current_class = class_match.group(1)
        drop_match = re.search(r"drop packets\s+([\d,]+)", line, re.IGNORECASE)
        if drop_match:
            add(
                "Control Plane",
                current_class,
                _counter_value(drop_match.group(1)) or 0,
                "Cumulative control-plane policy drops",
            )

    arp_output = outputs.get("arp_inspection", "")
    for line in arp_output.splitlines():
        tokens = line.split()
        if len(tokens) >= 3 and tokens[0].isdigit():
            dropped = _counter_value(tokens[2])
            if dropped is not None:
                add(
                    "Dynamic ARP Inspection",
                    f"VLAN {tokens[0]}",
                    dropped,
                    "Cumulative DAI drops",
                )

    for source, category in (
        ("arp_inspection", "Dynamic ARP Inspection"),
        ("dhcp_snooping", "DHCP Snooping"),
    ):
        for line in outputs.get(source, "").splitlines():
            match = re.search(
                r"^\s*(?P<label>[^:=]*drop[^:=]*?)\s*(?:=|:)\s*"
                r"(?P<count>[\d,]+)\s*$",
                line,
                re.IGNORECASE,
            )
            if not match:
                continue
            label = " ".join(match.group("label").split())
            add(
                category,
                label or category,
                _counter_value(match.group("count")) or 0,
                "Cumulative security drop counter",
            )
    return sorted(
        exceptions.values(),
        key=lambda item: (item.category.casefold(), item.object_name.casefold()),
    )


def parse_interface_summary(output: str) -> InterfaceSummary:
    connected = 0
    total = 0
    err_disabled: list[dict[str, str]] = []

    for line in output.splitlines():
        tokens = line.split()
        if not tokens or not PHYSICAL_INTERFACE_RE.match(tokens[0]):
            continue
        status_match = INTERFACE_STATUS_RE.search(line)
        if not status_match:
            continue
        status = status_match.group(1).casefold()
        total += 1
        connected += int(status == "connected")
        if status in {"err-disabled", "errdisabled"}:
            err_disabled.extend(parse_err_disabled_interfaces(line))

    if total:
        return InterfaceSummary(
            connected=connected,
            not_connected=total - connected,
            total=total,
            err_disabled=err_disabled,
        )

    # Router-style fallback for "show ip interface brief".
    brief_pattern = re.compile(
        r"^\s*(?P<interface>\S+)\s+\S+\s+\S+\s+\S+\s+"
        r"(?P<status>administratively down|up|down)\s+(?P<protocol>up|down)\s*$",
        re.IGNORECASE,
    )
    for line in output.splitlines():
        match = brief_pattern.match(line)
        if not match or not PHYSICAL_INTERFACE_RE.match(match.group("interface")):
            continue
        total += 1
        connected += int(
            match.group("status").casefold() == "up"
            and match.group("protocol").casefold() == "up"
        )

    return InterfaceSummary(
        connected=connected if total else None,
        not_connected=(total - connected) if total else None,
        total=total if total else None,
        err_disabled=[],
    )


def parse_err_disabled_interfaces(output: str) -> list[dict[str, str]]:
    """Return err-disabled physical interfaces and any reported reason."""

    issues: dict[str, dict[str, str]] = {}
    for line in output.splitlines():
        tokens = line.split()
        if not tokens or not PHYSICAL_INTERFACE_RE.match(tokens[0]):
            continue
        lowered = line.casefold()
        if "err-disabled" not in lowered and "errdisabled" not in lowered:
            continue
        reason_match = ERRDISABLE_REASON_RE.search(line)
        interface = tokens[0]
        issues[interface.casefold()] = {
            "interface": interface,
            "reason": reason_match.group(1) if reason_match else "Reason not reported",
        }
    return sorted(issues.values(), key=lambda item: item["interface"].casefold())


def merge_interface_issues(
    *issue_groups: Sequence[Mapping[str, str]],
) -> list[dict[str, str]]:
    """Deduplicate interface issues while retaining the most useful reason."""

    merged: dict[str, dict[str, str]] = {}
    for group in issue_groups:
        for issue in group:
            interface = str(issue.get("interface", "")).strip()
            if not interface:
                continue
            reason = str(issue.get("reason", "Reason not reported")).strip()
            key = interface.casefold()
            if key not in merged or merged[key]["reason"] == "Reason not reported":
                merged[key] = {"interface": interface, "reason": reason}
    return sorted(merged.values(), key=lambda item: item["interface"].casefold())


def parse_port_security_issues(output: str) -> list[PortSecurityIssue]:
    """Parse affected interfaces from the Cisco port-security summary table."""

    issues: dict[str, PortSecurityIssue] = {}
    row_pattern = re.compile(
        r"^\s*(?P<interface>\S+)\s+"
        r"(?P<maximum>\d+)\s+(?P<current>\d+)\s+"
        r"(?P<violations>\d+)\s+(?P<action>\S+)",
        re.IGNORECASE,
    )
    for line in output.splitlines():
        match = row_pattern.match(line)
        if not match or not PHYSICAL_INTERFACE_RE.match(match.group("interface")):
            continue
        violations = int(match.group("violations"))
        if violations <= 0:
            continue
        issue = PortSecurityIssue(
            interface=match.group("interface"),
            violation_count=violations,
            action=match.group("action"),
        )
        issues[issue.interface.casefold()] = issue
    return sorted(issues.values(), key=lambda item: item.interface.casefold())


def parse_etherchannels(output: str) -> list[EtherChannelState]:
    """Parse channel state and member flags from EtherChannel summaries."""

    raw_channels: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for line in output.splitlines():
        row_match = ETHERCHANNEL_ROW_RE.match(line)
        if row_match:
            if current is not None:
                raw_channels.append(current)
            current = {
                "group": int(row_match.group("group")),
                "name": row_match.group("name").replace(" ", ""),
                "flags": row_match.group("flags") or "",
                "text": row_match.group("rest"),
            }
            continue
        if current is not None and ETHERCHANNEL_MEMBER_RE.search(line):
            current["text"] = f"{current['text']} {line.strip()}"
    if current is not None:
        raw_channels.append(current)

    channels: list[EtherChannelState] = []
    for raw in raw_channels:
        text = str(raw["text"])
        protocol = "Unknown"
        for token in re.findall(
            r"\b(?:LACP|PAgP|static|on|none)\b", text, re.IGNORECASE
        ):
            protocol = token
            break

        members: list[str] = []
        bundled: list[str] = []
        standby: list[str] = []
        problem: list[str] = []
        member_details: list[dict[str, str]] = []
        seen: set[str] = set()
        for member_match in ETHERCHANNEL_MEMBER_RE.finditer(text):
            member = member_match.group("name")
            flags = member_match.group("flags")
            key = member.casefold()
            if key in seen:
                continue
            seen.add(key)
            members.append(member)
            meanings = [
                ETHERCHANNEL_MEMBER_FLAG_MEANINGS[flag]
                for flag in flags
                if flag in ETHERCHANNEL_MEMBER_FLAG_MEANINGS
            ]
            if "P" in flags:
                bundled.append(member)
                member_state = "Bundled"
                is_problem = False
            elif "H" in flags:
                standby.append(member)
                member_state = "Hot standby"
                is_problem = False
            else:
                member_state = ", ".join(meanings) or "Not bundled"
                is_problem = True
                problem.append(f"{member} ({flags}: {member_state})")
            member_details.append(
                {
                    "interface": member,
                    "flags": flags,
                    "state": member_state,
                    "problem": "Yes" if is_problem else "No",
                }
            )

        channel_flags = str(raw["flags"])
        channel_is_up = "U" in channel_flags or (
            not channel_flags and bool(bundled)
        )
        risk_reasons: list[str] = []
        if not channel_is_up:
            state = "Down"
            risk_reasons.append("port-channel is not in use")
        elif problem or not members:
            state = "Degraded"
            if problem:
                risk_reasons.append("one or more members are not bundled")
            if not members:
                risk_reasons.append("no member interfaces were parsed")
        else:
            state = "Up"
        if "M" in channel_flags or "m" in channel_flags:
            if "minimum links not met" not in risk_reasons:
                risk_reasons.append("minimum links not met")
            state = "Down"

        channels.append(
            EtherChannelState(
                name=str(raw["name"]),
                group=int(raw["group"]),
                protocol=protocol,
                flags=channel_flags,
                state=state,
                members=members,
                bundled_members=bundled,
                standby_members=standby,
                problem_members=problem,
                member_details=member_details,
                risk_reasons=risk_reasons,
            )
        )
    return sorted(channels, key=lambda item: (item.group, item.name.casefold()))


def parse_etherchannel_min_links(output: str) -> dict[int, int]:
    """Parse configured minimum-link requirements from channel detail output."""

    minimums: dict[int, int] = {}
    current_group: int | None = None
    for line in output.splitlines():
        group_match = re.search(r"\bGroup\s*:?\s*(\d+)\b", line, re.IGNORECASE)
        if group_match:
            current_group = int(group_match.group(1))
        min_match = re.search(
            r"\b(?:Minimum Links|Min-?links)\s*:?\s*(\d+)\b",
            line,
            re.IGNORECASE,
        )
        if current_group is not None and min_match:
            minimums[current_group] = int(min_match.group(1))
    return minimums


def enrich_etherchannels(
    channels: Sequence[EtherChannelState], detail_output: str
) -> None:
    """Apply minimum-link requirements to parsed EtherChannel state."""

    minimums = parse_etherchannel_min_links(detail_output)
    for channel in channels:
        channel.min_links = minimums.get(channel.group)
        if not channel.min_links:
            continue
        bundled = len(channel.bundled_members)
        if bundled < channel.min_links:
            reason = (
                f"{bundled}/{channel.min_links} required members are bundled; "
                "minimum links not met"
            )
            if reason not in channel.risk_reasons:
                channel.risk_reasons.append(reason)
            channel.state = "Down"


def count_environment_alerts(output: str) -> int:
    alert_terms = re.compile(
        r"\b(fail(?:ed|ure)?|critical|alarm|shutdown|overtemp)\b", re.IGNORECASE
    )
    healthy_terms = re.compile(
        r"\b(no alarms?|normal|ok|good|passed|not present)\b", re.IGNORECASE
    )
    return sum(
        1
        for line in output.splitlines()
        if alert_terms.search(line) and not healthy_terms.search(line)
    )


def collect_device_health(
    task: Task, read_timeout: float = DEFAULT_READ_TIMEOUT
) -> Result:
    metadata = netbox_metadata(task.host)
    record = DeviceHealth(hostname=task.host.name, **metadata)
    profile = command_profile(record.platform)

    try:
        version_output, _ = run_first_supported(
            task, "Firmware", profile["version"], read_timeout
        )
    except Exception as exc:  # noqa: BLE001 - collection errors must remain in the report.
        record.notes.append(f"Connection or show version failed: {exc}")
        return Result(host=task.host, changed=False, result=record.to_dict())

    record.reachable = True
    record.firmware = parse_firmware(version_output)
    record.collected_at_utc = datetime.now(UTC).replace(tzinfo=None)

    try:
        interface_output, _ = run_first_supported(
            task, "Interfaces", profile["interfaces"], read_timeout
        )
        summary = parse_interface_summary(interface_output)
        record.connected_interfaces = summary.connected
        record.not_connected_interfaces = summary.not_connected
        record.total_interfaces = summary.total
        record.err_disabled_interfaces = summary.err_disabled
        if summary.total is None:
            record.notes.append(
                "Interface status output was returned but no physical ports were parsed."
            )
    except Exception as exc:  # noqa: BLE001 - collection errors must remain in the report.
        record.notes.append(f"Interface statistics unavailable: {exc}")

    quality_outputs: list[str] = []
    quality_errors: list[str] = []
    for label, profile_key in (
        ("Interface error counters", "interface_errors"),
        ("Interface detail", "interface_detail"),
    ):
        try:
            output, _ = run_first_supported(
                task, label, profile[profile_key], read_timeout
            )
            quality_outputs.append(output)
        except Exception:  # noqa: BLE001 - command support varies by platform.
            quality_errors.append(label)

    if quality_outputs:
        quality_issues = merge_interface_quality_issues(
            *(parse_interface_quality(output) for output in quality_outputs)
        )
        record.interface_quality_issues = [
            issue.to_dict() for issue in quality_issues
        ]
        record.interfaces_with_errors = sum(
            issue.error_total > 0 for issue in quality_issues
        )
        record.interface_error_total = sum(
            issue.error_total for issue in quality_issues
        )
        record.interface_instability_count = sum(
            issue.unstable for issue in quality_issues
        )
        if quality_errors:
            record.notes.append(
                "Some interface-quality detail was unavailable: "
                + ", ".join(quality_errors)
                + "."
            )
    else:
        record.notes.append(
            "Interface quality and instability commands were unsupported or "
            "returned no usable output."
        )

    try:
        errdisabled_output, _ = run_first_supported(
            task, "Err-disabled interfaces", profile["err_disabled"], read_timeout
        )
        record.err_disabled_interfaces = merge_interface_issues(
            record.err_disabled_interfaces,
            parse_err_disabled_interfaces(errdisabled_output),
        )
        record.err_disabled_count = len(record.err_disabled_interfaces)
    except Exception as exc:  # noqa: BLE001 - unsupported commands vary by platform.
        if record.total_interfaces is not None:
            record.err_disabled_count = len(record.err_disabled_interfaces)
        record.notes.append(f"Err-disabled interface detail unavailable: {exc}")

    try:
        etherchannel_output, _ = run_first_supported(
            task, "EtherChannels", profile["etherchannels"], read_timeout
        )
        channels = parse_etherchannels(etherchannel_output)
        try:
            etherchannel_detail, _ = run_first_supported(
                task,
                "EtherChannel detail",
                profile["etherchannel_detail"],
                read_timeout,
            )
            enrich_etherchannels(channels, etherchannel_detail)
        except Exception:  # noqa: BLE001 - deep detail is platform-dependent.
            if channels:
                record.notes.append(
                    "EtherChannel detail was unavailable; summary member flags "
                    "were still collected."
                )
        record.etherchannels = [channel.to_dict() for channel in channels]
        record.etherchannels_total = len(channels)
        record.etherchannels_healthy = sum(
            channel.state == "Up" for channel in channels
        )
        record.etherchannels_degraded = sum(
            channel.state == "Degraded" for channel in channels
        )
        record.etherchannels_down = sum(
            channel.state == "Down" for channel in channels
        )
        record.etherchannel_members_total = sum(
            len(channel.members) for channel in channels
        )
        record.etherchannel_members_bundled = sum(
            len(channel.bundled_members) for channel in channels
        )
        record.etherchannel_members_standby = sum(
            len(channel.standby_members) for channel in channels
        )
        record.etherchannel_member_issues = sum(
            len(channel.problem_members) for channel in channels
        )
        record.etherchannel_min_link_risks = sum(
            any("minimum links" in reason for reason in channel.risk_reasons)
            for channel in channels
        )
    except Exception as exc:  # noqa: BLE001 - unsupported commands vary by platform.
        record.notes.append(f"EtherChannel status unavailable: {exc}")

    try:
        spanning_tree_output, _ = run_first_supported(
            task, "Spanning tree", profile["spanning_tree"], read_timeout
        )
        instances = parse_spanning_tree_instances(spanning_tree_output)
        record.spanning_tree_instances = [item.to_dict() for item in instances]
        record.spanning_tree_instance_count = len(instances)
        record.stp_topology_changes = sum(
            item.topology_changes for item in instances
        )
        record.stp_recent_changes = sum(item.recent_change for item in instances)
        if not instances and "no spanning tree" not in spanning_tree_output.casefold():
            record.notes.append(
                "Spanning-tree output was returned but no instances were parsed."
            )
        try:
            inconsistent_output, _ = run_first_supported(
                task,
                "STP inconsistent ports",
                profile["stp_inconsistent"],
                read_timeout,
            )
            record.stp_inconsistent_ports = parse_stp_inconsistent_ports(
                inconsistent_output
            )
            record.stp_inconsistent_port_count = len(
                record.stp_inconsistent_ports
            )
        except Exception:  # noqa: BLE001 - not all platforms expose this view.
            record.notes.append(
                "STP inconsistent-port detail was unavailable; instance health was "
                "still collected."
            )
    except Exception as exc:  # noqa: BLE001 - routers may not support spanning tree.
        record.notes.append(f"Spanning-tree health unavailable: {exc}")

    try:
        port_security_output, _ = run_first_supported(
            task, "Port security", profile["port_security"], read_timeout
        )
        security_issues = parse_port_security_issues(port_security_output)
        record.port_security_issues = [issue.to_dict() for issue in security_issues]
        record.port_security_interfaces = len(security_issues)
        record.port_security_violations = sum(
            issue.violation_count for issue in security_issues
        )
    except Exception as exc:  # noqa: BLE001 - unsupported commands vary by platform.
        record.notes.append(f"Port-security status unavailable: {exc}")

    security_outputs: dict[str, str] = {}
    security_errors: list[str] = []
    for profile_key, label in (
        ("control_plane", "Control-plane policy"),
        ("arp_inspection", "Dynamic ARP Inspection"),
        ("dhcp_snooping", "DHCP snooping"),
    ):
        try:
            output, _ = run_first_supported(
                task, label, profile[profile_key], read_timeout
            )
            security_outputs[profile_key] = output
        except Exception:  # noqa: BLE001 - features are frequently optional.
            security_errors.append(label)

    if security_outputs:
        exceptions = parse_security_control_exceptions(security_outputs)
        record.security_control_exceptions = [
            exception.to_dict() for exception in exceptions
        ]
        record.security_control_exception_count = len(exceptions)
        record.security_control_drop_count = sum(
            exception.count for exception in exceptions
        )
        if security_errors:
            record.notes.append(
                "Some security/control-plane views were unavailable: "
                + ", ".join(security_errors)
                + "."
            )
    else:
        record.notes.append(
            "Security/control-plane exception commands were unsupported or returned "
            "no usable output."
        )

    try:
        cpu_output, _ = run_first_supported(
            task, "CPU", profile["cpu"], read_timeout
        )
        record.cpu_pct = parse_cpu_pct(cpu_output)
        if record.cpu_pct is None:
            record.notes.append(
                "CPU output was returned but utilization could not be parsed."
            )
    except Exception as exc:  # noqa: BLE001 - collection errors must remain in the report.
        record.notes.append(f"CPU utilization unavailable: {exc}")

    try:
        environment_output, _ = run_first_supported(
            task, "Environment", profile["environment"], read_timeout
        )
        record.environment_alerts = count_environment_alerts(environment_output)
    except Exception as exc:  # noqa: BLE001 - collection errors must remain in the report.
        record.notes.append(f"Environment status unavailable: {exc}")

    return Result(host=task.host, changed=False, result=record.to_dict())


def _extract_records(results: Any, hosts: Mapping[str, Host]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for hostname, host in hosts.items():
        record: dict[str, Any] | None = None
        multi_result = results.get(hostname)
        if multi_result is not None:
            for item in multi_result:
                if isinstance(item.result, dict) and item.result.get("hostname") == hostname:
                    record = item.result
                    break
        if record is None:
            record = DeviceHealth(
                hostname=hostname,
                **netbox_metadata(host),
                notes=["Nornir did not return a collection result for this device."],
            ).to_dict()
        records.append(record)
    return sorted(records, key=lambda item: item["hostname"].casefold())


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _compact_detail_preview(value: Any, max_chars: int = 110) -> str:
    """Return a single compact preview while retaining room for a note hint."""

    text = " ".join(str(value or "").split())
    if len(text) <= max_chars:
        return text
    suffix = "... [see note]"
    available = max(20, max_chars - len(suffix))
    preview = text[:available].rsplit(" ", 1)[0].rstrip(" ,;:")
    return f"{preview}{suffix}"


def _set_compact_detail_cell(
    cell: Any,
    full_text: Any,
    *,
    max_chars: int = 110,
) -> None:
    """Show a short cell preview and retain unabridged text in an Excel note."""

    text = str(full_text or "").strip()
    cell.value = _compact_detail_preview(text, max_chars=max_chars)
    if not text:
        return
    note_text = re.sub(r"\s*\|\s*", "\n", text)
    note = Comment(
        "Full unabridged details:\n\n"
        f"{note_text}\n\n"
        "Tip: hover over or select this noted cell in Excel to review the full text.",
        "Network Automation",
    )
    note.width = 620
    note.height = 360
    cell.comment = note


def build_collection_notes(record: Mapping[str, Any]) -> list[str]:
    """Combine collection warnings with the conditions that reduced health."""

    notes = [str(note) for note in record.get("notes", []) if str(note).strip()]

    def add(note: str) -> None:
        if note not in notes:
            notes.append(note)

    if not record.get("reachable"):
        add(
            "Health impact: device was unreachable; live health data could not "
            "be collected."
        )
        return notes

    cpu_pct = record.get("cpu_pct")
    if isinstance(cpu_pct, (int, float)):
        if cpu_pct >= CPU_CRITICAL_PCT:
            add(
                f"Health impact: CPU utilization {cpu_pct:.1f}% met or exceeded the "
                f"critical threshold of {CPU_CRITICAL_PCT:.0f}%."
            )
        elif cpu_pct >= CPU_WARNING_PCT:
            add(
                f"Health impact: CPU utilization {cpu_pct:.1f}% met or exceeded the "
                f"warning threshold of {CPU_WARNING_PCT:.0f}%."
            )

    environment_alerts = record.get("environment_alerts")
    if isinstance(environment_alerts, (int, float)) and environment_alerts > 0:
        add(
            f"Health impact: {int(environment_alerts)} environmental alert"
            f"{'s were' if environment_alerts != 1 else ' was'} detected."
        )

    err_disabled = record.get("err_disabled_interfaces", [])
    if isinstance(err_disabled, Sequence) and err_disabled:
        details = ", ".join(
            f"{issue.get('interface', 'Unknown')} ({issue.get('reason', 'Reason not reported')})"
            for issue in err_disabled
            if isinstance(issue, Mapping)
        )
        add(f"Health impact: err-disabled: {details}.")

    port_security_issues = record.get("port_security_issues", [])
    if isinstance(port_security_issues, Sequence) and port_security_issues:
        details = ", ".join(
            f"{issue.get('interface', 'Unknown')} "
            f"({issue.get('violation_count', 0)} violations / "
            f"{issue.get('action', 'Unknown')})"
            for issue in port_security_issues
            if isinstance(issue, Mapping)
        )
        add(f"Health impact: port security: {details}.")

    interface_quality = record.get("interface_quality_issues", [])
    if isinstance(interface_quality, Sequence) and interface_quality:
        quality_details: list[str] = []
        for issue in interface_quality[:8]:
            if not isinstance(issue, Mapping):
                continue
            parts: list[str] = []
            error_total = issue.get("error_total")
            if isinstance(error_total, (int, float)) and error_total > 0:
                parts.append(f"{int(error_total):,} cumulative error counters")
            if issue.get("interface_resets"):
                parts.append(f"interface resets {issue.get('interface_resets')}")
            if issue.get("carrier_transitions"):
                parts.append(
                    f"carrier transitions {issue.get('carrier_transitions')}"
                )
            if issue.get("recent_link_flap"):
                parts.append(
                    f"last link flap {issue.get('last_link_flapped', 'recent')} ago"
                )
            quality_details.append(f"{issue.get('interface', 'Unknown')} ({', '.join(parts)})")
        if len(interface_quality) > 8:
            quality_details.append(
                f"{len(interface_quality) - 8} additional interfaces; see Issue Details"
            )
        add(
            "Health impact: interface quality/instability: "
            + "; ".join(quality_details)
            + "."
        )

    etherchannels = record.get("etherchannels", [])
    if isinstance(etherchannels, Sequence):
        for channel in etherchannels:
            if not isinstance(channel, Mapping) or channel.get("state") == "Up":
                continue
            members = channel.get("members", [])
            bundled = channel.get("bundled_members", [])
            problem = channel.get("problem_members", [])
            detail = (
                f"Health impact: {channel.get('name', 'Unknown channel')} "
                f"{channel.get('state', 'Unknown')}; {len(bundled)}/{len(members)} "
                "members bundled"
            )
            if problem:
                detail += f"; affected: {', '.join(map(str, problem))}"
            risks = channel.get("risk_reasons", [])
            if risks:
                concise_risks = [
                    "minimum links not met" if "minimum links" in str(risk) else str(risk)
                    for risk in risks
                ]
                detail += f"; {', '.join(dict.fromkeys(concise_risks))}"
            add(f"{detail}.")

    recent_stp = [
        item
        for item in record.get("spanning_tree_instances", [])
        if isinstance(item, Mapping) and item.get("recent_change")
    ]
    if recent_stp:
        details = ", ".join(
            f"{item.get('instance', 'Unknown')} changed "
            f"{item.get('last_change', 'recently')} ago"
            + (
                f" from {item.get('change_source')}"
                if item.get("change_source")
                else ""
            )
            for item in recent_stp
        )
        add(f"Health impact: recent STP changes: {details}.")

    inconsistent_ports = record.get("stp_inconsistent_ports", [])
    if isinstance(inconsistent_ports, Sequence) and inconsistent_ports:
        details = ", ".join(
            f"{issue.get('interface', 'Unknown')} "
            f"({issue.get('instance', 'STP')}: {issue.get('reason', 'Inconsistent')})"
            for issue in inconsistent_ports
            if isinstance(issue, Mapping)
        )
        add(f"Health impact: STP inconsistent ports: {details}.")

    security_exceptions = record.get("security_control_exceptions", [])
    if isinstance(security_exceptions, Sequence) and security_exceptions:
        object_names = [
            f"{issue.get('category', 'Security')} / {issue.get('object_name', 'Unknown')}"
            for issue in security_exceptions[:6]
            if isinstance(issue, Mapping)
        ]
        drop_total = sum(
            int(issue.get("count", 0))
            for issue in security_exceptions
            if isinstance(issue, Mapping)
            and isinstance(issue.get("count", 0), (int, float))
        )
        add(
            f"Health impact: {len(security_exceptions)} security/control-plane "
            f"exceptions with {drop_total:,} cumulative drops: "
            f"{', '.join(object_names)}."
        )

    available_components = 1 + sum(
        isinstance(record.get(field), (int, float))
        for field in (
            "cpu_pct",
            "environment_alerts",
            "err_disabled_count",
            "port_security_interfaces",
            "etherchannels_total",
            "interfaces_with_errors",
            "spanning_tree_instance_count",
            "security_control_exception_count",
        )
    )
    coverage_pct = available_components / HEALTH_COMPONENT_COUNT * 100
    if coverage_pct < MIN_DATA_COVERAGE_PCT:
        add(
            f"Health impact: only {coverage_pct:.0f}% of health inputs were available; "
            "health data is incomplete."
        )

    return notes


def _create_compact_health_workbook(
    records: Sequence[Mapping[str, Any]],
    target_tag: str,
    output_path: Path,
) -> None:
    """Retain the compact scorecard layout as an internal compatibility helper."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Network Health"
    sheet.sheet_view.showGridLines = False

    navy = "17365D"
    teal = "147D74"
    green = "C6EFCE"
    amber = "FFEB9C"
    red = "FFC7CE"
    pale_blue = "D9EAF7"
    white = "FFFFFF"
    thin_gray = Side(style="thin", color="D9E1F2")

    last_column = max(2, len(records) + 1)
    last_column_letter = get_column_letter(last_column)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet["A1"] = "Nornir Networking Report - Networking Infrastructure Health"
    sheet["A1"].font = Font(size=18, bold=True, color=white)
    sheet["A1"].fill = _fill(navy)
    sheet["A1"].alignment = Alignment(vertical="center", wrap_text=True, shrink_to_fit=True)
    sheet.row_dimensions[1].height = 42

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    generated = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    sheet["A2"] = (
        f"Scope: devices from the existing Nornir/NetBox inventory tagged "
        f"'{target_tag}' | Generated {generated}"
    )
    sheet["A2"].font = Font(italic=True, color="44546A")
    sheet["A2"].fill = _fill(pale_blue)
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True, shrink_to_fit=True)
    sheet.row_dimensions[2].height = 32

    metric_rows = {
        "Health Status": 5,
        "Health Score": 6,
        "Data Coverage": 7,
        "Location": 8,
        "Site": 9,
        "Role": 10,
        "Model": 11,
        "Platform": 12,
        "NetBox Status": 13,
        "Primary IP": 14,
        "Serial Number": 15,
        "Current Firmware": 16,
        "Interface Usage": 17,
        "Connected Interfaces": 18,
        "Not Connected Interfaces": 19,
        "Total Interfaces": 20,
        "CPU Utilization": 21,
        "Environment Alerts": 22,
        "Err-disabled Interfaces": 23,
        "Reachable": 24,
        "Collected UTC": 25,
        "Collection Notes": 26,
    }

    sheet["A4"] = "Metric"
    for label, row in metric_rows.items():
        sheet.cell(row, 1, label)

    for row in range(4, 27):
        cell = sheet.cell(row, 1)
        cell.fill = _fill(navy)
        cell.font = Font(bold=True, color=white)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    field_map = {
        "Location": "location",
        "Site": "site",
        "Role": "role",
        "Model": "model",
        "Platform": "platform",
        "NetBox Status": "netbox_status",
        "Primary IP": "primary_ip",
        "Serial Number": "serial_number",
        "Current Firmware": "firmware",
        "Connected Interfaces": "connected_interfaces",
        "Not Connected Interfaces": "not_connected_interfaces",
        "Total Interfaces": "total_interfaces",
        "Environment Alerts": "environment_alerts",
        "Err-disabled Interfaces": "err_disabled_count",
        "Collected UTC": "collected_at_utc",
    }

    for index, record in enumerate(records, start=2):
        column = get_column_letter(index)
        sheet.cell(4, index, record.get("hostname", ""))
        for label, field_name in field_map.items():
            sheet.cell(metric_rows[label], index, record.get(field_name))

        cpu_pct = record.get("cpu_pct")
        sheet.cell(21, index, cpu_pct / 100 if isinstance(cpu_pct, (int, float)) else None)
        sheet.cell(24, index, "Yes" if record.get("reachable") else "No")
        sheet.cell(26, index, " | ".join(build_collection_notes(record)))

        sheet.cell(17, index, f'=IF({column}20=0,"",{column}18/{column}20)')
        sheet.cell(7, index, f"=(1+COUNT({column}21:{column}23))/4")
        sheet.cell(
            6,
            index,
            (
                f'=IF({column}24<>"Yes",0,MAX(0,100'
                f'-IF(ISNUMBER({column}21),IF({column}21>=$B$32,$B$35,'
                f'IF({column}21>=$B$31,$B$34,0)),0)'
                f'-IF(ISNUMBER({column}22),MIN({column}22*$B$36,$B$38),0)'
                f'-IF(ISNUMBER({column}23),MIN({column}23*$B$37,$B$38),0)))'
            ),
        )
        sheet.cell(
            5,
            index,
            (
                f'=IF({column}24<>"Yes","Unreachable",'
                f'IF({column}7<$B$33,"Insufficient Data",'
                f'IF({column}6>=$B$39,"Healthy",'
                f'IF({column}6>=$B$40,"Watch","Critical"))))'
            ),
        )

        header = sheet.cell(4, index)
        header.fill = _fill(teal)
        header.font = Font(bold=True, color=white)
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[column].width = 24

        for row in range(5, 27):
            cell = sheet.cell(row, index)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin_gray)

    sheet["A29"] = "Health scoring method"
    sheet.merge_cells("A29:B29")
    sheet["A29"].fill = _fill(navy)
    sheet["A29"].font = Font(bold=True, color=white)
    sheet["A30"] = "Parameter"
    sheet["B30"] = "Value"
    settings = [
        ("CPU warning", CPU_WARNING_PCT / 100),
        ("CPU critical", CPU_CRITICAL_PCT / 100),
        ("Minimum data coverage", MIN_DATA_COVERAGE_PCT / 100),
        ("Warning penalty", WARNING_PENALTY),
        ("Critical penalty", CRITICAL_PENALTY),
        ("Environment alert penalty", ENVIRONMENT_ALERT_PENALTY),
        ("Err-disabled penalty", ERR_DISABLED_PENALTY),
        ("Maximum count penalty", MAX_COUNT_PENALTY),
        ("Healthy score", HEALTHY_SCORE),
        ("Watch score", WATCH_SCORE),
    ]
    for row, (label, value) in enumerate(settings, start=31):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
    for cell in sheet[30][:2]:
        cell.fill = _fill(teal)
        cell.font = Font(bold=True, color=white)
    for row in range(31, 34):
        sheet.cell(row, 2).number_format = "0%"

    sheet["A42"] = "Definitions"
    sheet["A42"].font = Font(bold=True)
    sheet.merge_cells(start_row=42, start_column=2, end_row=42, end_column=last_column)
    sheet.merge_cells(start_row=43, start_column=2, end_row=43, end_column=last_column)
    sheet["B42"] = (
        "Interface Usage = connected physical interfaces divided by all parsed physical "
        "interfaces. Unconnected, disabled, and err-disabled ports count as not connected."
    )
    sheet["B43"] = (
        "Health Score begins at 100 and applies visible CPU, environment, and "
        "err-disabled penalties. Missing metrics reduce Data Coverage instead of being "
        "treated as healthy or unhealthy."
    )
    sheet["B42"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet["B43"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[42].height = 58
    sheet.row_dimensions[43].height = 72

    sheet["A45"] = "Fleet summary"
    sheet.merge_cells("A45:B45")
    sheet["A45"].fill = _fill(navy)
    sheet["A45"].font = Font(bold=True, color=white)
    summary_labels = [
        "Devices",
        "Healthy",
        "Watch",
        "Critical",
        "Unreachable",
        "Insufficient Data",
        "Average Health Score",
        "Average Interface Usage",
    ]
    for row, label in enumerate(summary_labels, start=46):
        sheet.cell(row, 1, label)

    device_header_range = f"B4:{last_column_letter}4"
    status_range = f"B5:{last_column_letter}5"
    score_range = f"B6:{last_column_letter}6"
    usage_range = f"B17:{last_column_letter}17"
    sheet["B46"] = f"=COUNTA({device_header_range})"
    for row, status in zip(range(47, 52), summary_labels[1:6], strict=True):
        sheet.cell(row, 2, f'=COUNTIF({status_range},"{status}")')
    sheet["B52"] = f'=IFERROR(AVERAGE({score_range}),"")'
    sheet["B53"] = f'=IFERROR(AVERAGE({usage_range}),"")'
    sheet["B53"].number_format = "0%"

    sheet.column_dimensions["A"].width = 29
    sheet.column_dimensions["B"].width = max(sheet.column_dimensions["B"].width or 0, 24)
    sheet.row_dimensions[4].height = 30
    sheet.row_dimensions[26].height = 160
    sheet.freeze_panes = "A5"

    # Keep scoring mechanics available to formulas without showing them to
    # management users. Hidden rows collapse the method/definition section so
    # Fleet Summary follows the scorecard when scrolling.
    for row in range(29, 45):
        sheet.row_dimensions[row].hidden = True

    for column in range(2, last_column + 1):
        sheet.cell(6, column).number_format = "0"
        sheet.cell(7, column).number_format = "0%"
        sheet.cell(17, column).number_format = "0.0%"
        sheet.cell(21, column).number_format = "0.0%"
        sheet.cell(25, column).number_format = "yyyy-mm-dd hh:mm"

    status_cells = f"B5:{last_column_letter}5"
    for status, color in {
        "Healthy": green,
        "Watch": amber,
        "Critical": red,
        "Unreachable": red,
        "Insufficient Data": amber,
    }.items():
        sheet.conditional_formatting.add(
            status_cells,
            FormulaRule(formula=[f'B5="{status}"'], fill=_fill(color)),
        )
    sheet.conditional_formatting.add(
        f"B6:{last_column_letter}6",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=70,
            mid_color="FFEB84",
            end_type="num",
            end_value=100,
            end_color="63BE7B",
        ),
    )

    sheet["B17"].comment = Comment(
        "Connected physical interfaces divided by all parsed physical interfaces.",
        "Network Automation",
    )

    for row in range(29, 54):
        for column in range(1, min(last_column, 2) + 1):
            sheet.cell(row, column).alignment = Alignment(vertical="center", wrap_text=True)

    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_title_rows = "1:4"
    sheet.print_area = f"A1:{last_column_letter}26"

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _interface_issue_text(record: Mapping[str, Any]) -> str:
    issues = record.get("err_disabled_interfaces", [])
    if not isinstance(issues, Sequence):
        return ""
    return ", ".join(
        f"{issue.get('interface', 'Unknown')} ({issue.get('reason', 'Reason not reported')})"
        for issue in issues
        if isinstance(issue, Mapping)
    )


def _port_security_text(record: Mapping[str, Any]) -> str:
    issues = record.get("port_security_issues", [])
    if not isinstance(issues, Sequence):
        return ""
    return ", ".join(
        f"{issue.get('interface', 'Unknown')}: "
        f"{issue.get('violation_count', 0)} violations / "
        f"{issue.get('action', 'Unknown')}"
        for issue in issues
        if isinstance(issue, Mapping)
    )


def _interface_quality_text(record: Mapping[str, Any]) -> str:
    issues = record.get("interface_quality_issues", [])
    if not isinstance(issues, Sequence):
        return ""
    details: list[str] = []
    for issue in issues:
        if not isinstance(issue, Mapping):
            continue
        parts: list[str] = []
        counters = issue.get("counters", {})
        if isinstance(counters, Mapping):
            values = [
                f"{name.replace('_', ' ')}={value}"
                for name, value in counters.items()
                if isinstance(value, (int, float)) and value > 0
            ]
            if values:
                parts.append(", ".join(values))
        if issue.get("interface_resets"):
            parts.append(f"resets={issue.get('interface_resets')}")
        if issue.get("carrier_transitions"):
            parts.append(f"carrier transitions={issue.get('carrier_transitions')}")
        if issue.get("recent_link_flap"):
            parts.append(f"last flap={issue.get('last_link_flapped', 'recent')} ago")
        details.append(f"{issue.get('interface', 'Unknown')}: {'; '.join(parts)}")
    return " | ".join(details)


def _spanning_tree_text(record: Mapping[str, Any]) -> str:
    details: list[str] = []
    instances = record.get("spanning_tree_instances", [])
    if isinstance(instances, Sequence):
        for item in instances:
            if not isinstance(item, Mapping):
                continue
            text = (
                f"{item.get('instance', 'Unknown')}: root "
                f"{item.get('root_bridge') or 'not parsed'}, root port "
                f"{item.get('root_port') or 'local/not parsed'}, "
                f"topology changes {item.get('topology_changes', 0)}"
            )
            if item.get("last_change"):
                text += f", last {item.get('last_change')} ago"
            if item.get("change_source"):
                text += f" from {item.get('change_source')}"
            details.append(text)
    inconsistent = record.get("stp_inconsistent_ports", [])
    if isinstance(inconsistent, Sequence):
        for issue in inconsistent:
            if isinstance(issue, Mapping):
                details.append(
                    f"{issue.get('instance', 'STP')} / "
                    f"{issue.get('interface', 'Unknown')}: "
                    f"{issue.get('reason', 'Inconsistent')}"
                )
    return " | ".join(details)


def _etherchannel_issue_text(record: Mapping[str, Any]) -> str:
    channels = record.get("etherchannels", [])
    if not isinstance(channels, Sequence):
        return ""
    details: list[str] = []
    for channel in channels:
        if not isinstance(channel, Mapping) or channel.get("state") == "Up":
            continue
        members = channel.get("members", [])
        bundled = channel.get("bundled_members", [])
        problem = channel.get("problem_members", [])
        text = (
            f"{channel.get('name', 'Unknown')} {channel.get('state', 'Unknown')} "
            f"({len(bundled)}/{len(members)} bundled)"
        )
        if problem:
            text += f": {', '.join(map(str, problem))}"
        standby = channel.get("standby_members", [])
        if standby:
            text += f"; hot standby: {', '.join(map(str, standby))}"
        if channel.get("min_links") is not None:
            text += f"; minimum links: {channel.get('min_links')}"
        risks = channel.get("risk_reasons", [])
        if risks:
            text += f"; risks: {', '.join(map(str, risks))}"
        details.append(text)
    return " | ".join(details)


def _etherchannel_member_state_text(record: Mapping[str, Any]) -> str:
    channels = record.get("etherchannels", [])
    if not isinstance(channels, Sequence):
        return ""
    details: list[str] = []
    for channel in channels:
        if not isinstance(channel, Mapping):
            continue
        members = channel.get("member_details", [])
        member_text = ", ".join(
            f"{member.get('interface', 'Unknown')}={member.get('state', 'Unknown')}"
            for member in members
            if isinstance(member, Mapping)
        )
        text = (
            f"{channel.get('name', 'Unknown')} [{channel.get('protocol', 'Unknown')}] "
            f"{channel.get('state', 'Unknown')}"
        )
        if channel.get("min_links") is not None:
            text += f", min-links {channel.get('min_links')}"
        if member_text:
            text += f": {member_text}"
        details.append(text)
    return " | ".join(details)


def _security_control_text(record: Mapping[str, Any]) -> str:
    issues = record.get("security_control_exceptions", [])
    if not isinstance(issues, Sequence):
        return ""
    return " | ".join(
        f"{issue.get('category', 'Security')} / "
        f"{issue.get('object_name', 'Unknown')}: {issue.get('count', 0)} — "
        f"{issue.get('details', '')}"
        for issue in issues
        if isinstance(issue, Mapping)
    )


def build_issue_rows(records: Sequence[Mapping[str, Any]]) -> list[list[Any]]:
    """Create one engineer-actionable row for every detected issue."""

    rows: list[list[Any]] = []
    for record in records:
        hostname = str(record.get("hostname", "Unknown"))
        location = str(record.get("location", ""))
        primary_ip = str(record.get("primary_ip", ""))
        collected = record.get("collected_at_utc")

        def add(
            severity: str,
            category: str,
            object_name: str,
            state: str,
            details: str,
            *,
            _hostname: str = hostname,
            _location: str = location,
            _primary_ip: str = primary_ip,
            _collected: Any = collected,
        ) -> None:
            rows.append(
                [
                    _hostname,
                    severity,
                    category,
                    object_name,
                    state,
                    details,
                    _location,
                    _primary_ip,
                    _collected,
                ]
            )

        if not record.get("reachable"):
            add(
                "Critical",
                "Reachability",
                hostname,
                "Unreachable",
                "SSH collection did not complete.",
            )
            continue

        cpu_pct = record.get("cpu_pct")
        if isinstance(cpu_pct, (int, float)) and cpu_pct >= CPU_WARNING_PCT:
            severity = "Critical" if cpu_pct >= CPU_CRITICAL_PCT else "Warning"
            add(
                severity,
                "CPU",
                hostname,
                f"{cpu_pct:.1f}%",
                "CPU utilization exceeded the configured health threshold.",
            )

        environment_alerts = record.get("environment_alerts")
        if isinstance(environment_alerts, (int, float)) and environment_alerts > 0:
            add(
                "Warning",
                "Environment",
                hostname,
                f"{int(environment_alerts)} alerts",
                "Environmental command output contained active alert terms.",
            )

        for issue in record.get("err_disabled_interfaces", []):
            if isinstance(issue, Mapping):
                add(
                    "Critical",
                    "Err-disabled",
                    str(issue.get("interface", "Unknown")),
                    "Err-disabled",
                    f"Reason: {issue.get('reason', 'Reason not reported')}",
                )

        for issue in record.get("port_security_issues", []):
            if isinstance(issue, Mapping):
                add(
                    "Critical",
                    "Port Security",
                    str(issue.get("interface", "Unknown")),
                    str(issue.get("action", "Violation")),
                    f"Violation counter: {issue.get('violation_count', 0)}",
                )

        for issue in record.get("interface_quality_issues", []):
            if not isinstance(issue, Mapping):
                continue
            error_total = issue.get("error_total", 0)
            error_total = error_total if isinstance(error_total, (int, float)) else 0
            states: list[str] = []
            detail_parts: list[str] = []
            counters = issue.get("counters", {})
            if error_total > 0:
                states.append("Quality errors")
            if isinstance(counters, Mapping):
                detail_parts.extend(
                    f"{name.replace('_', ' ')}={value}"
                    for name, value in counters.items()
                    if isinstance(value, (int, float)) and value > 0
                )
            if issue.get("interface_resets"):
                states.append("Instability")
                detail_parts.append(f"resets={issue.get('interface_resets')}")
            if issue.get("carrier_transitions"):
                if "Instability" not in states:
                    states.append("Instability")
                detail_parts.append(
                    f"carrier transitions={issue.get('carrier_transitions')}"
                )
            if issue.get("recent_link_flap"):
                if "Instability" not in states:
                    states.append("Instability")
                detail_parts.append(
                    f"last link flap={issue.get('last_link_flapped', 'recent')} ago"
                )
            add(
                "Critical"
                if error_total >= INTERFACE_ERROR_CRITICAL_COUNT
                else "Warning",
                "Interface Quality",
                str(issue.get("interface", "Unknown")),
                " / ".join(states) or "Exception",
                "; ".join(detail_parts)
                + "; counters are cumulative unless cleared on the device",
            )

        for channel in record.get("etherchannels", []):
            if not isinstance(channel, Mapping) or channel.get("state") == "Up":
                continue
            state = str(channel.get("state", "Unknown"))
            problem = channel.get("problem_members", [])
            bundled = channel.get("bundled_members", [])
            standby = channel.get("standby_members", [])
            members = channel.get("members", [])
            details = (
                f"Protocol {channel.get('protocol', 'Unknown')}; "
                f"{len(bundled)}/{len(members)} members bundled"
            )
            if problem:
                details += f"; affected: {', '.join(map(str, problem))}"
            if standby:
                details += f"; hot standby: {', '.join(map(str, standby))}"
            if channel.get("min_links") is not None:
                details += f"; minimum links: {channel.get('min_links')}"
            risks = channel.get("risk_reasons", [])
            if risks:
                details += f"; risks: {', '.join(map(str, risks))}"
            add(
                "Critical" if state == "Down" else "Warning",
                "EtherChannel",
                str(channel.get("name", "Unknown")),
                state,
                details,
            )

        for instance in record.get("spanning_tree_instances", []):
            if not isinstance(instance, Mapping) or not instance.get("recent_change"):
                continue
            details = (
                f"{instance.get('topology_changes', 0)} cumulative topology changes; "
                f"last change {instance.get('last_change', 'recently')} ago"
            )
            if instance.get("change_source"):
                details += f" from {instance.get('change_source')}"
            if instance.get("root_bridge"):
                details += f"; current root {instance.get('root_bridge')}"
            add(
                "Warning",
                "Spanning Tree",
                str(instance.get("instance", "Unknown")),
                "Recent topology change",
                details,
            )

        for issue in record.get("stp_inconsistent_ports", []):
            if isinstance(issue, Mapping):
                add(
                    "Critical",
                    "Spanning Tree",
                    str(issue.get("interface", "Unknown")),
                    "Inconsistent",
                    f"{issue.get('instance', 'STP')}: "
                    f"{issue.get('reason', 'Inconsistent')}",
                )

        for issue in record.get("security_control_exceptions", []):
            if isinstance(issue, Mapping):
                add(
                    "Warning",
                    str(issue.get("category", "Security / Control Plane")),
                    str(issue.get("object_name", "Unknown")),
                    f"{issue.get('count', 0)} drops",
                    str(issue.get("details", "Cumulative exception counter")),
                )

        for note in record.get("notes", []):
            add("Warning", "Telemetry", hostname, "Collection note", str(note))
    return rows


def create_elaborate_health_workbook(
    records: Sequence[Mapping[str, Any]],
    target_tag: str,
    output_path: Path,
) -> None:
    """Create an engineer overview plus a normalized issue-detail worksheet."""

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Engineer Health"
    overview.sheet_view.showGridLines = False
    details = workbook.create_sheet("Issue Details")
    details.sheet_view.showGridLines = False
    scoring = workbook.create_sheet("Scoring")

    navy = "17365D"
    teal = "147D74"
    green = "C6EFCE"
    amber = "FFEB9C"
    red = "FFC7CE"
    pale_blue = "D9EAF7"
    pale_teal = "DDEBF7"
    white = "FFFFFF"
    gray = "44546A"
    thin_gray = Side(style="thin", color="D9E1F2")

    last_column = max(2, len(records) + 1)
    last_column_letter = get_column_letter(last_column)
    generated = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")

    overview.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=last_column
    )
    overview["A1"] = "Nornir Engineer Report — Network Infrastructure Health"
    overview["A1"].font = Font(size=18, bold=True, color=white)
    overview["A1"].fill = _fill(navy)
    overview["A1"].alignment = Alignment(
        vertical="center", wrap_text=True, shrink_to_fit=True
    )
    overview.row_dimensions[1].height = 42

    overview.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=last_column
    )
    overview["A2"] = (
        f"Scope: Nornir/NetBox devices tagged '{target_tag}' | Generated {generated}"
    )
    overview["A2"].font = Font(italic=True, color=gray)
    overview["A2"].fill = _fill(pale_blue)
    overview["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    overview.row_dimensions[2].height = 30

    metric_rows = {
        "Health Status": 5,
        "Health Score": 6,
        "Data Coverage": 7,
        "Location": 8,
        "Site": 9,
        "Role": 10,
        "Model": 11,
        "Platform": 12,
        "NetBox Status": 13,
        "Primary IP": 14,
        "Serial Number": 15,
        "Current Firmware": 16,
        "Reachable": 17,
        "Collected UTC": 18,
        "Interface Health": 20,
        "Interface Usage": 21,
        "Connected Interfaces": 22,
        "Not Connected Interfaces": 23,
        "Total Interfaces": 24,
        "Err-disabled Interfaces": 25,
        "Err-disabled Details": 26,
        "Port-security Interfaces": 27,
        "Port-security Violations": 28,
        "Port-security Details": 29,
        "EtherChannel Health": 31,
        "EtherChannels Total": 32,
        "EtherChannels Healthy": 33,
        "EtherChannels Degraded": 34,
        "EtherChannels Down": 35,
        "Member Availability": 36,
        "Bundled Members": 37,
        "Total Members": 38,
        "Degraded / Down Details": 39,
        "Min-link Risks": 40,
        "Hot-standby Members": 41,
        "Member State Details": 42,
        "System and Collection Health": 44,
        "CPU Utilization": 45,
        "Environment Alerts": 46,
        "Interface Quality and Instability": 48,
        "Interfaces with Errors": 49,
        "Error Counter Sum": 50,
        "Instability Indicators": 51,
        "Interface Quality Details": 52,
        "Spanning Tree Health": 54,
        "STP Instances": 55,
        "Topology Changes": 56,
        "Recent Topology Changes": 57,
        "Inconsistent Ports": 58,
        "STP Details": 59,
        "Security and Control Plane Exceptions": 61,
        "Additional Exceptions": 62,
        "Drop Counter Sum": 63,
        "Security / Control Plane Details": 64,
        "Engineering Notes": 66,
    }
    section_rows = {20, 31, 44, 48, 54, 61}

    overview["A4"] = "Metric"
    for label, row in metric_rows.items():
        overview.cell(row, 1, label)

    for row in range(4, 67):
        cell = overview.cell(row, 1)
        if row in section_rows:
            cell.fill = _fill(teal)
        else:
            cell.fill = _fill(navy)
        cell.font = Font(bold=True, color=white)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in section_rows:
        for column in range(2, last_column + 1):
            cell = overview.cell(row, column)
            cell.fill = _fill(pale_teal)
            cell.border = Border(bottom=thin_gray)

    field_map = {
        "Location": "location",
        "Site": "site",
        "Role": "role",
        "Model": "model",
        "Platform": "platform",
        "NetBox Status": "netbox_status",
        "Primary IP": "primary_ip",
        "Serial Number": "serial_number",
        "Current Firmware": "firmware",
        "Collected UTC": "collected_at_utc",
        "Connected Interfaces": "connected_interfaces",
        "Not Connected Interfaces": "not_connected_interfaces",
        "Total Interfaces": "total_interfaces",
        "Err-disabled Interfaces": "err_disabled_count",
        "Port-security Interfaces": "port_security_interfaces",
        "Port-security Violations": "port_security_violations",
        "EtherChannels Total": "etherchannels_total",
        "EtherChannels Healthy": "etherchannels_healthy",
        "EtherChannels Degraded": "etherchannels_degraded",
        "EtherChannels Down": "etherchannels_down",
        "Bundled Members": "etherchannel_members_bundled",
        "Total Members": "etherchannel_members_total",
        "Environment Alerts": "environment_alerts",
        "Interfaces with Errors": "interfaces_with_errors",
        "Error Counter Sum": "interface_error_total",
        "Instability Indicators": "interface_instability_count",
        "STP Instances": "spanning_tree_instance_count",
        "Topology Changes": "stp_topology_changes",
        "Recent Topology Changes": "stp_recent_changes",
        "Inconsistent Ports": "stp_inconsistent_port_count",
        "Min-link Risks": "etherchannel_min_link_risks",
        "Hot-standby Members": "etherchannel_members_standby",
        "Additional Exceptions": "security_control_exception_count",
        "Drop Counter Sum": "security_control_drop_count",
    }

    for index, record in enumerate(records, start=2):
        column = get_column_letter(index)
        overview.cell(4, index, record.get("hostname", ""))
        for label, field_name in field_map.items():
            overview.cell(metric_rows[label], index, record.get(field_name))

        overview.cell(17, index, "Yes" if record.get("reachable") else "No")
        for row, full_text, max_chars in (
            (26, _interface_issue_text(record), 100),
            (29, _port_security_text(record), 100),
            (39, _etherchannel_issue_text(record), 120),
            (42, _etherchannel_member_state_text(record), 120),
            (52, _interface_quality_text(record), 120),
            (59, _spanning_tree_text(record), 120),
            (64, _security_control_text(record), 120),
            (66, "\n".join(build_collection_notes(record)), 150),
        ):
            _set_compact_detail_cell(
                overview.cell(row, index),
                full_text,
                max_chars=max_chars,
            )

        cpu_pct = record.get("cpu_pct")
        overview.cell(
            45,
            index,
            cpu_pct / 100 if isinstance(cpu_pct, (int, float)) else None,
        )
        overview.cell(21, index, f'=IF({column}24=0,"",{column}22/{column}24)')
        overview.cell(36, index, f'=IF({column}38=0,"",{column}37/{column}38)')
        overview.cell(
            7,
            index,
            (
                f'=(IF({column}17="Yes",1,0)+'
                f'COUNT({column}45,{column}46,{column}25,{column}27,'
                f'{column}32,{column}49,{column}55,{column}62))/'
                f"{HEALTH_COMPONENT_COUNT}"
            ),
        )
        overview.cell(
            6,
            index,
            (
                f"=IF({column}17<>\"Yes\",0,MAX(0,100"
                f"-IF(ISNUMBER({column}45),IF({column}45>='Scoring'!$B$3,"
                f"'Scoring'!$B$6,IF({column}45>='Scoring'!$B$2,"
                f"'Scoring'!$B$5,0)),0)"
                f"-IF(ISNUMBER({column}46),MIN({column}46*'Scoring'!$B$7,"
                f"'Scoring'!$B$12),0)"
                f"-IF(ISNUMBER({column}25),MIN({column}25*'Scoring'!$B$8,"
                f"'Scoring'!$B$12),0)"
                f"-IF(ISNUMBER({column}27),MIN({column}27*'Scoring'!$B$9,"
                f"'Scoring'!$B$12),0)"
                f"-MIN(IF(ISNUMBER({column}34),{column}34*'Scoring'!$B$10,0)"
                f"+IF(ISNUMBER({column}35),{column}35*'Scoring'!$B$11,0),"
                f"'Scoring'!$B$13)"
                f"-MIN(IF(ISNUMBER({column}49),{column}49*'Scoring'!$B$16,0)"
                f"+IF(ISNUMBER({column}51),{column}51*'Scoring'!$B$17,0),"
                f"'Scoring'!$B$21)"
                f"-MIN(IF(ISNUMBER({column}57),{column}57*'Scoring'!$B$18,0)"
                f"+IF(ISNUMBER({column}58),{column}58*'Scoring'!$B$19,0),"
                f"'Scoring'!$B$21)"
                f"-IF(ISNUMBER({column}62),MIN({column}62*'Scoring'!$B$20,"
                f"'Scoring'!$B$21),0)))"
            ),
        )
        overview.cell(
            5,
            index,
            (
                f"=IF({column}17<>\"Yes\",\"Unreachable\","
                f"IF({column}7<'Scoring'!$B$4,\"Insufficient Data\","
                f"IF(OR({column}35>0,{column}58>0,"
                f"{column}6<'Scoring'!$B$15),\"Critical\","
                f"IF(OR({column}6<'Scoring'!$B$14,{column}34>0,"
                f"{column}25>0,{column}27>0,{column}46>0,{column}49>0,"
                f"{column}51>0,{column}57>0,{column}40>0,{column}62>0),"
                f"\"Watch\",\"Healthy\"))))"
            ),
        )

        header = overview.cell(4, index)
        header.fill = _fill(teal)
        header.font = Font(bold=True, color=white)
        header.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        overview.column_dimensions[column].width = 29

        for row in range(5, 67):
            if row in section_rows:
                continue
            cell = overview.cell(row, index)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin_gray)

    overview.column_dimensions["A"].width = 31
    overview.row_dimensions[4].height = 32
    overview.row_dimensions[26].height = 45
    overview.row_dimensions[29].height = 45
    overview.row_dimensions[39].height = 60
    overview.row_dimensions[42].height = 60
    overview.row_dimensions[52].height = 60
    overview.row_dimensions[59].height = 60
    overview.row_dimensions[64].height = 60
    overview.row_dimensions[66].height = 72
    overview.freeze_panes = "A5"

    for column in range(2, last_column + 1):
        overview.cell(6, column).number_format = "0"
        overview.cell(7, column).number_format = "0%"
        overview.cell(18, column).number_format = "yyyy-mm-dd hh:mm"
        overview.cell(21, column).number_format = "0.0%"
        overview.cell(36, column).number_format = "0.0%"
        overview.cell(45, column).number_format = "0.0%"

    status_cells = f"B5:{last_column_letter}5"
    for status, color in {
        "Healthy": green,
        "Watch": amber,
        "Critical": red,
        "Unreachable": red,
        "Insufficient Data": amber,
    }.items():
        overview.conditional_formatting.add(
            status_cells,
            FormulaRule(formula=[f'B5="{status}"'], fill=_fill(color)),
        )
    overview.conditional_formatting.add(
        f"B6:{last_column_letter}6",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=70,
            mid_color="FFEB84",
            end_type="num",
            end_value=100,
            end_color="63BE7B",
        ),
    )
    for row, color in (
        (25, red),
        (27, red),
        (34, amber),
        (35, red),
        (46, amber),
        (49, amber),
        (51, amber),
        (57, amber),
        (58, red),
        (40, red),
        (62, amber),
        (63, amber),
    ):
        overview.conditional_formatting.add(
            f"B{row}:{last_column_letter}{row}",
            FormulaRule(formula=[f"B{row}>0"], fill=_fill(color)),
        )

    overview["A21"].comment = Comment(
        "Connected physical interfaces divided by all parsed physical interfaces.",
        "Network Automation",
    )
    overview["A36"].comment = Comment(
        "Bundled EtherChannel members divided by all parsed channel members.",
        "Network Automation",
    )

    overview["A50"].comment = Comment(
        "Sum of parsed cumulative error/drop counters; use Issue Details to see "
        "the counter type and interface.",
        "Network Automation",
    )
    overview["A56"].comment = Comment(
        "Cumulative spanning-tree topology changes reported by all parsed instances.",
        "Network Automation",
    )
    overview["A63"].comment = Comment(
        "Cumulative parsed CoPP, DAI, and DHCP-snooping drop counters.",
        "Network Automation",
    )
    overview["A66"].comment = Comment(
        "Device cells in detail and Engineering Notes rows show a short preview. "
        "Hover over or select the noted device cell to read its complete text.",
        "Network Automation",
    )

    summary_start = 69
    overview[f"A{summary_start}"] = "Fleet Summary"
    overview.merge_cells(
        start_row=summary_start,
        start_column=1,
        end_row=summary_start,
        end_column=2,
    )
    overview[f"A{summary_start}"].fill = _fill(navy)
    overview[f"A{summary_start}"].font = Font(bold=True, color=white)
    summary_labels = [
        "Devices",
        "Healthy",
        "Watch",
        "Critical",
        "Unreachable",
        "Insufficient Data",
        "Average Health Score",
        "Total EtherChannels",
        "Degraded EtherChannels",
        "Down EtherChannels",
        "Err-disabled Interfaces",
        "Port-security Interfaces",
        "Interfaces with Quality Errors",
        "Instability Indicators",
        "Recent STP Changes",
        "STP Inconsistent Ports",
        "EtherChannel Min-link Risks",
        "Security / Control Exceptions",
    ]
    for row, label in enumerate(summary_labels, start=summary_start + 1):
        overview.cell(row, 1, label)

    status_range = f"B5:{last_column_letter}5"
    overview.cell(summary_start + 1, 2, f"=COUNTA(B4:{last_column_letter}4)")
    for offset, status in enumerate(summary_labels[1:6], start=2):
        overview.cell(
            summary_start + offset,
            2,
            f'=COUNTIF({status_range},"{status}")',
        )
    overview.cell(
        summary_start + 7,
        2,
        f'=IFERROR(AVERAGE(B6:{last_column_letter}6),"")',
    )
    for offset, source_row in enumerate(
        (32, 34, 35, 25, 27, 49, 51, 57, 58, 40, 62),
        start=8,
    ):
        overview.cell(
            summary_start + offset,
            2,
            f"=SUM(B{source_row}:{last_column_letter}{source_row})",
        )
    overview.cell(summary_start + 7, 2).number_format = "0"
    for row in range(summary_start + 1, summary_start + len(summary_labels) + 1):
        overview.cell(row, 1).font = Font(bold=True)
        overview.cell(row, 1).fill = _fill(pale_blue)
        overview.cell(row, 1).alignment = Alignment(wrap_text=True)
        overview.cell(row, 2).alignment = Alignment(horizontal="right")

    overview.sheet_properties.pageSetUpPr.fitToPage = True
    overview.page_setup.orientation = "landscape"
    overview.page_setup.fitToWidth = 1
    overview.page_setup.fitToHeight = 0
    overview.print_title_rows = "1:4"
    overview.print_area = (
        f"A1:{last_column_letter}{summary_start + len(summary_labels)}"
    )
    overview.row_breaks.append(Break(id=65))
    overview.row_breaks.append(Break(id=summary_start - 1))

    details.merge_cells("A1:I1")
    details["A1"] = "Engineer Issue Details"
    details["A1"].font = Font(size=18, bold=True, color=white)
    details["A1"].fill = _fill(navy)
    details["A1"].alignment = Alignment(vertical="center")
    details.row_dimensions[1].height = 42
    details.merge_cells("A2:I2")
    details["A2"] = (
        "One row per actionable device, interface, spanning-tree, security, "
        f"EtherChannel, or telemetry issue | Generated {generated}"
    )
    details["A2"].font = Font(italic=True, color=gray)
    details["A2"].fill = _fill(pale_blue)
    details["A2"].alignment = Alignment(wrap_text=True)
    details.row_dimensions[2].height = 30

    headers = [
        "Device",
        "Severity",
        "Category",
        "Object",
        "State",
        "Details",
        "Location",
        "Primary IP",
        "Collected UTC",
        "Full Details",
    ]
    for column, header in enumerate(headers, start=1):
        details.cell(4, column, header)
    issue_rows = build_issue_rows(records)
    if not issue_rows:
        issue_rows = [
            [
                "Fleet",
                "Info",
                "Health",
                "All devices",
                "No active issues",
                "No engineer-actionable conditions were detected.",
                "",
                "",
                None,
            ]
        ]
    for row_index, row in enumerate(issue_rows, start=5):
        full_detail = str(row[5] or "")
        for column_index, value in enumerate(row, start=1):
            cell = details.cell(row_index, column_index)
            if column_index == 6:
                _set_compact_detail_cell(cell, full_detail, max_chars=105)
            else:
                cell.value = value
        details.cell(row_index, 10, full_detail)
        details.row_dimensions[row_index].height = 36

    detail_last_row = 4 + len(issue_rows)
    detail_table = Table(displayName="EngineerIssues", ref=f"A4:J{detail_last_row}")
    detail_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    details.add_table(detail_table)
    details.freeze_panes = "A5"
    details.column_dimensions["A"].width = 24
    details.column_dimensions["B"].width = 12
    details.column_dimensions["C"].width = 18
    details.column_dimensions["D"].width = 20
    details.column_dimensions["E"].width = 18
    details.column_dimensions["F"].width = 52
    details.column_dimensions["G"].width = 20
    details.column_dimensions["H"].width = 18
    details.column_dimensions["I"].width = 20
    details.column_dimensions["J"].width = 90
    details.column_dimensions["J"].hidden = True
    for row in range(5, detail_last_row + 1):
        details.cell(row, 9).number_format = "yyyy-mm-dd hh:mm"
    for row in details.iter_rows(min_row=5, max_row=detail_last_row, max_col=10):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    details["F4"].comment = Comment(
        "Visible Details cells are compact previews. Hover over or select a noted "
        "cell for the complete text. The same unabridged value is retained in the "
        "hidden Full Details column.",
        "Network Automation",
    )
    details.conditional_formatting.add(
        f"A5:I{detail_last_row}",
        FormulaRule(formula=['$B5="Critical"'], fill=_fill(red)),
    )
    details.conditional_formatting.add(
        f"A5:I{detail_last_row}",
        FormulaRule(formula=['$B5="Warning"'], fill=_fill(amber)),
    )
    details.sheet_properties.pageSetUpPr.fitToPage = True
    details.page_setup.orientation = "landscape"
    details.page_setup.fitToWidth = 1
    details.page_setup.fitToHeight = 0
    details.print_title_rows = "1:4"
    details.print_area = f"A1:I{detail_last_row}"

    scoring_rows = [
        ("Parameter", "Value"),
        ("CPU warning", CPU_WARNING_PCT / 100),
        ("CPU critical", CPU_CRITICAL_PCT / 100),
        ("Minimum data coverage", MIN_DATA_COVERAGE_PCT / 100),
        ("Warning penalty", WARNING_PENALTY),
        ("Critical penalty", CRITICAL_PENALTY),
        ("Environment alert penalty", ENVIRONMENT_ALERT_PENALTY),
        ("Err-disabled penalty", ERR_DISABLED_PENALTY),
        ("Port-security penalty", PORT_SECURITY_PENALTY),
        ("Degraded EtherChannel penalty", DEGRADED_ETHERCHANNEL_PENALTY),
        ("Down EtherChannel penalty", DOWN_ETHERCHANNEL_PENALTY),
        ("Maximum count penalty", MAX_COUNT_PENALTY),
        ("Maximum EtherChannel penalty", MAX_ETHERCHANNEL_PENALTY),
        ("Healthy score", HEALTHY_SCORE),
        ("Watch score", WATCH_SCORE),
        ("Interface quality penalty", INTERFACE_QUALITY_PENALTY),
        ("Interface instability penalty", INTERFACE_INSTABILITY_PENALTY),
        ("Recent STP change penalty", STP_RECENT_CHANGE_PENALTY),
        ("STP inconsistent-port penalty", STP_INCONSISTENT_PORT_PENALTY),
        ("Security/control-plane exception penalty", SECURITY_EXCEPTION_PENALTY),
        ("Maximum engineering exception penalty", MAX_ENGINEERING_EXCEPTION_PENALTY),
    ]
    for row in scoring_rows:
        scoring.append(row)
    for row in range(2, 5):
        scoring.cell(row, 2).number_format = "0%"
    scoring.sheet_state = "veryHidden"

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    config_path = Path(args.config).expanduser()
    output_path = Path(args.output).expanduser()
    target_tag = str(args.tag).strip().casefold()

    if not config_path.is_file():
        console.print(
            f"[bold red]ERROR:[/] Existing Nornir config was not found: {config_path.resolve()}"
        )
        return 1
    if output_path.suffix.casefold() != ".xlsx":
        console.print("[bold red]ERROR:[/] --output must end in .xlsx")
        return 1
    if not target_tag:
        console.print("[bold red]ERROR:[/] --tag cannot be empty")
        return 1

    username = os.getenv("NORNIR_USERNAME")
    password = os.getenv("NORNIR_PASSWORD")
    if bool(username) != bool(password):
        console.print(
            "[bold red]ERROR:[/] Set both NORNIR_USERNAME and NORNIR_PASSWORD, or neither."
        )
        return 1

    try:
        nr = InitNornir(config_file=str(config_path))
    except Exception as exc:  # noqa: BLE001 - inventory plugins raise varied errors.
        console.print(f"[bold red]ERROR:[/] Could not initialize Nornir: {exc}")
        return 1

    if username and password:
        nr.inventory.defaults.username = username
        nr.inventory.defaults.password = password

    netmiko_extras = build_netmiko_extras(args)
    for host in nr.inventory.hosts.values():
        host.data["tag_slugs"] = normalize_tags(host.data.get("tags", []))
        if "netmiko" not in host.connection_options:
            host.connection_options["netmiko"] = ConnectionOptions(
                extras=dict(netmiko_extras)
            )
        else:
            existing_extras = host.connection_options["netmiko"].extras or {}
            existing_extras.update(netmiko_extras)
            host.connection_options["netmiko"].extras = existing_extras

    targets = nr.filter(F(tag_slugs__contains=target_tag))
    console.print(f"Target tag: [bold]{target_tag}[/]")
    console.print(f"Matched devices: [bold]{len(targets.inventory.hosts)}[/]")
    console.print(
        "SSH timing: "
        f"connect {args.connection_timeout:g}s | "
        f"authenticate {args.auth_timeout:g}s | "
        f"banner {args.banner_timeout:g}s | "
        f"command {args.read_timeout:g}s | "
        f"delay factor {args.global_delay_factor:g}"
    )
    if args.legacy_ssh:
        console.print(
            "[yellow]Legacy SSH compatibility is enabled for this run.[/]"
        )

    if not targets.inventory.hosts:
        console.print("No devices matched the requested NetBox tag. No workbook was created.")
        return 0

    console.print(
        "Collecting firmware, interface quality/instability, spanning-tree, "
        "EtherChannel details, security/control-plane, CPU, and environment health..."
    )
    results = targets.run(
        name="Collect network health",
        task=collect_device_health,
        read_timeout=args.read_timeout,
    )
    records = _extract_records(results, targets.inventory.hosts)
    create_elaborate_health_workbook(records, target_tag, output_path)

    reachable = sum(bool(record.get("reachable")) for record in records)
    console.print(f"[bold green]Created:[/] {output_path.resolve()}")
    console.print(f"Reachable devices: {reachable}/{len(records)}")
    if reachable < len(records):
        console.print(
            "Unreachable devices remain in the workbook with their NetBox information and notes."
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
